# Standard library imports
import io
import json
import logging
import os
import random
import shutil
import subprocess
import sys
import warnings
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Type, Union
from urllib.error import HTTPError
from webbrowser import get

# Third party imports
import brickschema.brickify.src.handlers.Handler.TableHandler as table_handler
import pandas as pd
import pydotplus
import rdflib.tools.csv2rdf
import typer
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from rdflib import RDF, RDFS, Graph, Literal, Namespace, URIRef
from rdflib.tools.rdf2dot import rdf2dot

# Local application imports
import twin4build.core as core
from twin4build.utils.get_obj_attr import get_obj_attr
from twin4build.utils.mkdir_in_root import mkdir_in_root
from twin4build.utils.print_progress import LOGGER, autoreset_print
from twin4build.utils.uppath import uppath

DYNAMIC_PARSING = True
IGNORE_PARSING_FOR_NAMESPACES = ["XSD", "FPO"]


def parse_wrapper(graph, source=None, **kwargs):
    """
    Wrapper for rdflib.graph.parse that handles XSD namespace separately.
    We don't parse XSD namespace because it isn't defined in rdf.
    """
    namespaces = [
        str(getattr(core.namespace, s)) for s in IGNORE_PARSING_FOR_NAMESPACES
    ]

    if isinstance(source, (str, URIRef)):
        if str(source) in namespaces:  # Dont parse XSD
            return
    graph.parse(source, **kwargs)


def get_short_name(uri: Union[str, URIRef], namespaces: Dict[str, Namespace]):
    for namespace in namespaces.values():
        if namespace in str(uri):
            return str(uri).split(namespace)[-1]
    return None


class SemanticEntity:
    """Base class for all semantic entities (types, instances, predicates).

    Provides shared URI handling, namespace resolution, and comparison methods.
    """

    def __init__(self, uri: Union[str, URIRef, Literal], model: "SemanticModel"):
        self.uri = URIRef(uri) if isinstance(uri, str) else uri
        self.model = model
        self._namespace = (None, None)

    def get_short_name(self) -> Optional[str]:
        """Get the local name (without namespace prefix)"""
        for namespace in self.model.namespaces.values():
            if str(namespace) in str(self.uri):
                return str(self.uri).split(str(namespace))[-1]
        return None

    def get_namespace(self) -> Tuple[Optional[str], Optional[Namespace]]:
        """Get the namespace prefix and URI for this entity"""
        if self._namespace == (None, None):
            d = get_obj_attr(core.namespace)
            tup = [(k, v) for k, v in d.items() if v in str(self.uri)]
            if len(tup) > 0:
                self._namespace = tup[0]
                return self._namespace

            tup = [
                (prefix, namespace)
                for prefix, namespace in self.model.ontology_graph.namespaces()
                if str(namespace) in str(self.uri)
            ]
            if len(tup) > 0:
                self._namespace = tup[0]
                return self._namespace

        return self._namespace

    def __str__(self) -> str:
        return str(self.uri)

    def __repr__(self) -> str:
        return f"{self.__class__.__name__}({str(self.uri)})"

    def __hash__(self) -> int:
        return hash(str(self.uri))

    def __eq__(self, other) -> bool:
        if isinstance(other, SemanticEntity):
            return str(self.uri) == str(other.uri)
        elif isinstance(other, (str, URIRef, Literal)):
            return str(self.uri) == str(other)
        return False


@autoreset_print
class SemanticPredicate(SemanticEntity):
    """Represents an ontology predicate (property used in relationships)

    This class wraps predicates (properties) used in RDF triples and provides
    access to their characteristics like domain, range, inverse properties, etc.

    Note: In RDF, predicates are essentially properties. This class is a specialized
    wrapper for properties when they are used as predicates in relationships.
    """

    def __init__(self, uri: Union[str, URIRef], model: "SemanticModel"):
        super().__init__(uri, model)
        self._ontology_graph = model.ontology_graph
        self._domain = None
        self._range = None
        self._inverse_properties = None
        self._super_properties = None
        self._sub_properties = None
        self._equivalent_properties = None
        self._is_symmetric = None
        self._is_transitive = None
        self._is_functional = None
        self._is_inverse_functional = None

    @property
    def domain(self) -> Set[URIRef]:
        """Get the domain (valid subject types) of this predicate"""
        if self._domain is None:
            self._domain = set(self._ontology_graph.objects(self.uri, RDFS.domain))
        return self._domain

    @property
    def range(self) -> Set[URIRef]:
        """Get the range (valid object types) of this predicate"""
        if self._range is None:
            self._range = set(self._ontology_graph.objects(self.uri, RDFS.range))
        return self._range

    @property
    def inverse_properties(self) -> List["SemanticPredicate"]:
        """Get all inverse properties of this predicate"""
        if self._inverse_properties is None:
            self.parse_ontology()
            self._inverse_properties = []
            owl_inverse_of = URIRef("http://www.w3.org/2002/07/owl#inverseOf")

            # Check both directions since inverseOf is symmetric
            for inv_uri in self._ontology_graph.objects(self.uri, owl_inverse_of):
                self._inverse_properties.append(self.model.get_predicate(inv_uri))
            for inv_uri in self._ontology_graph.subjects(owl_inverse_of, self.uri):
                self._inverse_properties.append(self.model.get_predicate(inv_uri))

        return self._inverse_properties

    @property
    def super_properties(self) -> List["SemanticPredicate"]:
        """Get all parent properties (including indirect) using RDFS reasoning"""
        if self._super_properties is None:
            self.parse_ontology()
            self._super_properties = []
            rdfs_sub_property = URIRef(
                "http://www.w3.org/2000/01/rdf-schema#subPropertyOf"
            )

            for parent in self._ontology_graph.transitive_objects(
                self.uri, rdfs_sub_property
            ):
                if parent != self.uri:  # Exclude self
                    self._super_properties.append(self.model.get_predicate(parent))

        return self._super_properties

    @property
    def sub_properties(self) -> List["SemanticPredicate"]:
        """Get all sub properties (including indirect) using RDFS reasoning"""
        if self._sub_properties is None:
            self.parse_ontology()
            self._sub_properties = []
            rdfs_sub_property = URIRef(
                "http://www.w3.org/2000/01/rdf-schema#subPropertyOf"
            )

            for sub in self._ontology_graph.transitive_subjects(
                rdfs_sub_property, self.uri
            ):
                if sub != self.uri:  # Exclude self
                    self._sub_properties.append(self.model.get_predicate(sub))

        return self._sub_properties

    @property
    def equivalent_properties(self) -> List["SemanticPredicate"]:
        """Get all properties that are equivalent to this property"""
        if self._equivalent_properties is None:
            self.parse_ontology()
            self._equivalent_properties = []
            owl_equivalent_property = URIRef(
                "http://www.w3.org/2002/07/owl#equivalentProperty"
            )

            equivalent_uris = set()

            # Forward direction
            for equiv in self._ontology_graph.transitive_objects(
                self.uri, owl_equivalent_property
            ):
                if equiv != self.uri:
                    equivalent_uris.add(equiv)

            # Backward direction
            for equiv in self._ontology_graph.transitive_subjects(
                owl_equivalent_property, self.uri
            ):
                if equiv != self.uri:
                    equivalent_uris.add(equiv)

            # Convert to SemanticPredicate objects
            self._equivalent_properties = [
                self.model.get_predicate(uri) for uri in equivalent_uris
            ]

        return self._equivalent_properties

    @property
    def is_symmetric(self) -> bool:
        """Check if this predicate is symmetric"""
        if self._is_symmetric is None:
            self.parse_ontology()
            owl_symmetric = URIRef("http://www.w3.org/2002/07/owl#SymmetricProperty")
            self._is_symmetric = (
                self.uri,
                RDF.type,
                owl_symmetric,
            ) in self._ontology_graph
        return self._is_symmetric

    @property
    def is_transitive(self) -> bool:
        """Check if this predicate is transitive"""
        if self._is_transitive is None:
            self.parse_ontology()
            owl_transitive = URIRef("http://www.w3.org/2002/07/owl#TransitiveProperty")
            self._is_transitive = (
                self.uri,
                RDF.type,
                owl_transitive,
            ) in self._ontology_graph
        return self._is_transitive

    @property
    def is_functional(self) -> bool:
        """Check if this predicate is functional"""
        if self._is_functional is None:
            self.parse_ontology()
            owl_functional = URIRef("http://www.w3.org/2002/07/owl#FunctionalProperty")
            self._is_functional = (
                self.uri,
                RDF.type,
                owl_functional,
            ) in self._ontology_graph
        return self._is_functional

    @property
    def is_inverse_functional(self) -> bool:
        """Check if this predicate is inverse functional"""
        if self._is_inverse_functional is None:
            self.parse_ontology()
            owl_inverse_functional = URIRef(
                "http://www.w3.org/2002/07/owl#InverseFunctionalProperty"
            )
            self._is_inverse_functional = (
                self.uri,
                RDF.type,
                owl_inverse_functional,
            ) in self._ontology_graph
        return self._is_inverse_functional

    def isproperty(
        self,
        cls: Union[
            str,
            "SemanticPredicate",
            Tuple[Union[str, "SemanticPredicate"], ...],
            List[Union[str, "SemanticPredicate"]],
        ],
    ) -> bool:
        """Backward-compatible alias for ispredicate().

        Previously on the now-removed SemanticProperty class.
        """
        return self.ispredicate(cls)

    def parse_ontology(self):
        """Parse the ontology for this predicate's namespace

        This ensures that property characteristics (inverse, symmetric, transitive,
        equivalent properties, etc.) are available in the ontology_graph for reasoning.
        """
        if DYNAMIC_PARSING:
            prefix, namespace = self.get_namespace()

            if prefix is not None:
                # Parse the namespace into ontology_graph if not already parsed
                self.model.parse_namespaces(namespaces={prefix: namespace})
            else:
                message = f"Failed to parse namespace for predicate {self.uri}. Reasoning involving this predicate will not be possible."
                warnings.warn(message)

    def ispredicate(
        self,
        predicate: Union[
            str,
            "SemanticPredicate",
            URIRef,
            Tuple[Union[str, "SemanticPredicate", URIRef], ...],
            List[Union[str, "SemanticPredicate", URIRef]],
        ],
    ) -> bool:
        """Check if this predicate is any of the given predicates (including inheritance)

        Args:
            predicate: Single predicate or tuple/list of predicates to check against

        Returns:
            True if this predicate matches any of the specified predicates
        """
        if not isinstance(predicate, (tuple, list)):
            predicate = (predicate,)

        # Check each predicate in the tuple
        for p in predicate:
            if str(p) == str(self.uri):
                return True
            # Check super properties
            elif str(p) in [str(s.uri) for s in self.super_properties]:
                return True
            # Check equivalent properties
            elif str(p) in [str(e.uri) for e in self.equivalent_properties]:
                return True

        return False


# Backward compatibility: SemanticProperty is now SemanticPredicate
SemanticProperty = SemanticPredicate


@autoreset_print
class SemanticType(SemanticEntity):
    """Represents an ontology class with inheritance"""

    def __init__(self, uri: Union[str, URIRef], model: "SemanticModel", validate=False):
        super().__init__(uri, model)
        self._ontology_graph = model.ontology_graph

        if validate:

            # Built-in RDF/RDFS classes that are always valid
            BUILT_IN_CLASSES = {
                URIRef("http://www.w3.org/2000/01/rdf-schema#Resource"),
                URIRef("http://www.w3.org/2000/01/rdf-schema#Class"),
                URIRef("http://www.w3.org/2002/07/owl#Class"),
            }

            # Debug: Print all type triples for this URI
            # print(f"\nDebug - Checking type declarations for {self.uri}")
            # print("All triples where this URI is subject:")
            # for s, p, o in self.graph.triples((self.uri, None, None)):
            #     print(f"  {p} -> {o}")
            # print("All triples where this URI is object:")
            # for s, p, o in self.graph.triples((None, None, self.uri)):
            #     print(f"  {s} -> {p}")

            # Check if URI represents a valid type
            is_owl_class = (
                self.uri,
                RDF.type,
                URIRef("http://www.w3.org/2002/07/owl#Class"),
            ) in self._ontology_graph
            is_rdfs_class = (self.uri, RDF.type, RDFS.Class) in self._ontology_graph
            is_built_in = self.uri in BUILT_IN_CLASSES

            # Additional checks for class-like behavior
            has_subclass = any(
                self._ontology_graph.triples((None, RDFS.subClassOf, self.uri))
            )
            is_subclass = any(
                self._ontology_graph.triples((self.uri, RDFS.subClassOf, None))
            )
            has_instances = any(
                self.model.instance_graph.triples((None, RDF.type, self.uri))
            )

            # Check if it's a property (which should not be treated as a class)
            property_types = {
                URIRef("http://www.w3.org/2002/07/owl#ObjectProperty"),
                URIRef("http://www.w3.org/1999/02/22-rdf-syntax-ns#Property"),
                URIRef("http://www.w3.org/2002/07/owl#DatatypeProperty"),
                URIRef("http://www.w3.org/2002/07/owl#AnnotationProperty"),
                URIRef("http://www.w3.org/2002/07/owl#FunctionalProperty"),
            }

            is_property = any(
                (self.uri, RDF.type, prop_type) in self._ontology_graph
                for prop_type in property_types
            )
            is_used_as_predicate = any(
                self._ontology_graph.triples((None, self.uri, None))
            )

            # Check if it's used in domain/range declarations (suggesting it's a class)
            is_in_domain = any(
                self._ontology_graph.triples((None, RDFS.domain, self.uri))
            )
            is_in_range = any(
                self._ontology_graph.triples((None, RDFS.range, self.uri))
            )

            # print(f"Debug - Class checks for {self.uri}:")
            # print(f"  is_owl_class: {is_owl_class}")
            # print(f"  is_rdfs_class: {is_rdfs_class}")
            # print(f"  is_built_in: {is_built_in}")
            # print(f"  has_subclass: {has_subclass}")
            # print(f"  is_subclass: {is_subclass}")
            # print(f"  has_instances: {has_instances}")
            # print(f"  is_property: {is_property}")
            # print(f"  is_used_as_predicate: {is_used_as_predicate}")
            # print(f"  is_in_domain: {is_in_domain}")
            # print(f"  is_in_range: {is_in_range}")

            if is_property or is_used_as_predicate:
                raise ValueError(f"URI '{self.uri}' is a property, not a class")

            # Consider it a valid class if any of these conditions are true
            if not (
                is_owl_class
                or is_rdfs_class
                or is_built_in
                or has_subclass
                or is_subclass
                or has_instances
                or is_in_domain
                or is_in_range
            ):
                raise ValueError(
                    f"URI '{self.uri}' is not declared as a valid class/type in the ontology"
                )

        self._super_classes = None
        self._sub_classes = None
        self._equivalent_classes = None
        self._attributes = None
        self._namespace = (None, None)

    @property
    def super_classes(self) -> List["SemanticType"]:
        """Get all parent classes (including indirect) using RDFS reasoning.

        Note: This requires that ontology definitions (rdfs:subClassOf triples) are loaded
        into self._ontology_graph, typically by loading ontologies during initialization.
        """
        if self._super_classes is None:
            self.parse_ontology()
            self._super_classes = []

            # Collect parent classes using transitive closure
            # parent_uris = set()
            for parent in self._ontology_graph.transitive_objects(
                self.uri, RDFS.subClassOf
            ):
                if parent != self.uri:  # Exclude self
                    parent_type = self.model.get_type(parent)
                    self._super_classes.append(parent_type)
            for parent in self._super_classes:
                for equivalent in parent.equivalent_classes:
                    if equivalent not in self._super_classes:
                        self._super_classes.append(equivalent)

        return self._super_classes

    @property
    def sub_classes(self) -> List["SemanticType"]:
        """Get all sub classes (including indirect) using RDFS reasoning.

        Note: This requires that ontology definitions (rdfs:subClassOf triples) are loaded
        into self._ontology_graph, typically by loading ontologies during initialization.
        """

        if self._sub_classes is None:
            self.parse_ontology()
            self._sub_classes = []
            # Note: transitive_subjects has signature (predicate, object) not (object, predicate)!
            sub_uris = set()
            for sub in self._ontology_graph.transitive_subjects(
                RDFS.subClassOf, self.uri
            ):
                if sub != self.uri:  # Exclude self
                    sub_uris.add(sub)

            # Convert URIs to SemanticType objects
            for sub_uri in sub_uris:
                self._sub_classes.append(self.model.get_type(sub_uri))

        return self._sub_classes

    @property
    def equivalent_classes(self) -> List["SemanticType"]:
        """Get all classes that are equivalent to this class using owl:equivalentClass.

        Since owl:equivalentClass is bidirectional, we traverse in both directions using
        transitive_objects (forward) and transitive_subjects (backward).

        Note: This requires that ontology definitions (owl:equivalentClass triples) are loaded
        into self._ontology_graph, typically by loading ontologies during initialization.
        """

        if self._equivalent_classes is None:
            self.parse_ontology()
            owl_equivalent_class = URIRef(
                "http://www.w3.org/2002/07/owl#equivalentClass"
            )

            # Collect all equivalent classes by traversing in both directions
            equivalent_uris = set()

            # Forward direction: classes reachable from self.uri
            for equiv in self._ontology_graph.transitive_objects(
                self.uri, owl_equivalent_class
            ):
                if equiv != self.uri:
                    equivalent_uris.add(equiv)

            # Backward direction: classes that reach self.uri
            for equiv in self._ontology_graph.transitive_subjects(
                owl_equivalent_class, self.uri
            ):
                if equiv != self.uri:
                    equivalent_uris.add(equiv)

            # Convert to SemanticType objects
            self._equivalent_classes = [
                self.model.get_type(uri) for uri in equivalent_uris
            ]

        return self._equivalent_classes

    def get_type_attributes(self) -> Dict[str, List[Any]]:
        """Find all possible attributes (properties) that can be used with instances of this class.

        This method looks for all ObjectProperties defined in the ontology that could be used
        with instances of this class (self.uri). These are properties that either:
        1. Have no domain restrictions (can be used with any class)
        2. Have this class or any of its parent classes in their domain

        Returns:
            Dictionary mapping property names to lists of their allowed range values
        """
        if self._attributes is None:
            self.parse_ontology()
            self._attributes = {}

            # Find all ObjectProperties in the ontology
            for prop, _, _ in self._ontology_graph.triples(
                (None, RDF.type, URIRef("http://www.w3.org/2002/07/owl#ObjectProperty"))
            ):  # We are looking explicitly for ObjectProperties. This could maybe be generalized?
                # pred_name = str(prop).split('#')[-1]

                # Get the domains (if any) for this property
                domains = list(self._ontology_graph.objects(prop, RDFS.domain))

                super_classes = [str(s) for s in self.super_classes]

                # Property is valid if it has no domain restrictions or if this class/parents are in its domain
                if not domains or any(
                    str(domain) in super_classes or domain == self.uri
                    for domain in domains
                ):
                    # Get the ranges for this property
                    ranges = list(self._ontology_graph.objects(prop, RDFS.range))
                    self._attributes[prop] = ranges

        return self._attributes

    def istype(
        self,
        cls: Union[
            str,
            "SemanticType",
            Tuple[Union[str, "SemanticType"], ...],
            List[Union[str, "SemanticType"]],
        ],
    ) -> bool:
        """Check if this instance is of any of the given class types (including inheritance)

        Args:
            cls: Single class or tuple/list of classes to check against

        Returns:
            True if instance matches any of the specified classes
        """
        # Convert single class to tuple for consistent handling
        if not isinstance(cls, (tuple, list)):
            cls = (cls,)

        # Check each class in the tuple against all instance types
        for c in cls:
            if str(c) == str(self.uri):
                return True
            elif str(c) in [str(s) for s in self.super_classes]:
                return True
        return False

    def has_subclasses(self) -> bool:
        """Check if this type has any subclasses"""
        self.parse_ontology()
        return any(self._ontology_graph.triples((None, RDFS.subClassOf, self.uri)))

    def parse_ontology(self):
        # print(f"\n[DEBUG parse_ontology] Called for URI: {self.uri}")
        # print(f"[DEBUG parse_ontology] DYNAMIC_PARSING = {DYNAMIC_PARSING}")

        if DYNAMIC_PARSING:
            t = self.model.get_type(self.uri)
            # for p, n in self.model.ontology_graph.namespaces():
            # print(f"[DEBUG parse_ontology] Namespace: {p} -> {n}")
            prefix, namespace = t.get_namespace()
            # print(f"[DEBUG parse_ontology] Got namespace: prefix={prefix}, namespace={namespace}")

            if prefix is not None:
                # print(f"[DEBUG parse_ontology] Calling parse_namespaces for {prefix}: {namespace}")
                self.model.parse_namespaces(
                    namespaces={prefix: namespace}
                )  # Will only parse if not already parsed - now parses into ontology_graph
                # print(f"[DEBUG parse_ontology] parse_namespaces returned: {success}")
            else:
                # print(f"[DEBUG parse_ontology] ✗ prefix is None, skipping parse_namespaces")
                # success = False
                message = f"Failed to parse namespace for type {self.uri}. Reasoning involving this type will not be possible."
                # print(f"[DEBUG parse_ontology] ✗ ERROR: {message}")
                warnings.warn(message)
                # raise ValueError(message)
        else:
            print("[DEBUG parse_ontology] DYNAMIC PARSING IS DISABLED")


class SemanticObject(SemanticEntity):
    """Base class for semantic instances and literals.

    This is the base type used in type annotations throughout the codebase.
    Use SemanticInstance for URI resources and SemanticLiteral for literal values.
    The factory method SemanticModel.get_instance() returns the appropriate subclass.

    To check if an object is a literal, use isinstance(obj, SemanticLiteral).
    """

    @property
    def types(self) -> Set[SemanticType]:
        """Get all types of this object (overridden by subclasses)"""
        return set()

    @property
    def direct_types(self) -> Set[SemanticType]:
        """Get only the explicitly asserted types (no superclass expansion)"""
        return set()

    def get_predicate_object_pairs(
        self,
    ) -> Dict["SemanticPredicate", List[Union["SemanticObject", "SemanticType"]]]:
        """Return all attributes of this instance (empty for literals)"""
        return {}

    def isinstance(
        self,
        cls: Union[
            str,
            SemanticType,
            Tuple[Union[str, SemanticType], ...],
            List[Union[str, SemanticType]],
        ],
    ) -> bool:
        """Check if this object is of any of the given class types (including inheritance)"""
        return False

    def get_most_specific_type(self) -> Optional["SemanticType"]:
        """Get the most specific type of this object (overridden by subclasses)"""
        return None


@autoreset_print
class SemanticInstance(SemanticObject):
    """Represents a URI resource / OWL Individual in the semantic model."""

    def __init__(self, uri: Union[str, URIRef], model: "SemanticModel"):
        super().__init__(uri, model)
        self._types = None
        self._direct_types = None
        self._attributes = None

    @property
    def direct_types(self) -> Set[SemanticType]:
        """Get only the explicitly asserted rdf:type values (no superclass expansion).

        This is cheaper than types and sufficient when you just need the declared type.
        """
        if self._direct_types is None:
            direct = set(self.model.instance_graph.objects(self.uri, RDF.type))

            # Also check types through owl:sameAs relations
            same_as = set(
                self.model.instance_graph.objects(
                    self.uri, URIRef("http://www.w3.org/2002/07/owl#sameAs")
                )
            )
            for same_as_uri in same_as:
                direct.update(
                    set(self.model.instance_graph.objects(same_as_uri, RDF.type))
                )

            self._direct_types = {self.model.get_type(t) for t in direct}
        return self._direct_types

    @property
    def types(self) -> Set[SemanticType]:
        """Get all types of this instance (including superclasses and equivalent classes)"""
        if self._types is None:
            types_ = set()
            for type_obj in self.direct_types:
                types_.add(type_obj)

                for super_class in type_obj.super_classes:
                    types_.add(self.model.get_type(super_class.uri))

                for equivalent_class in type_obj.equivalent_classes:
                    types_.add(self.model.get_type(equivalent_class.uri))

            self._types = types_
        return self._types

    def _is_class_uri(self, obj: URIRef) -> bool:
        """Check if a URI represents a class rather than an instance.

        Checks both instance and ontology graphs, and whether the URI is used as a type.
        """
        OWL_CLASS = URIRef("http://www.w3.org/2002/07/owl#Class")
        return (
            any(self.model.instance_graph.triples((obj, RDF.type, OWL_CLASS)))
            or any(self.model.instance_graph.triples((obj, RDF.type, RDFS.Class)))
            or any(self.model.ontology_graph.triples((obj, RDF.type, OWL_CLASS)))
            or any(self.model.ontology_graph.triples((obj, RDF.type, RDFS.Class)))
        )

    def get_predicate_object_pairs(
        self,
    ) -> Dict[
        "SemanticPredicate", List[Union["SemanticObject", "SemanticType"]]
    ]:  # TODO: if the graph is changed dynamically, the caching done here wont work anymore. Make sure to update the cache when the graph is changed.
        """Return all attributes of this instance

        Returns:
            Dictionary mapping SemanticPredicate objects to lists of SemanticObject or SemanticType instances
        """
        if self._attributes is None:
            self._attributes = {}

            # Parse ontologies for all predicates used by this instance
            predicates_used = set()

            for pred, _ in self.model.instance_graph.predicate_objects(self.uri):
                predicates_used.add(pred)

            for _, pred, _ in self.model.instance_graph.triples((None, None, self.uri)):
                predicates_used.add(pred)

            for pred in predicates_used:
                pred_obj = self.model.get_predicate(pred)
                pred_obj.parse_ontology()

            #########################################################
            # On-demand reasoning for this specific instance
            #########################################################

            inferred_pairs = []

            owl_inverse_of = URIRef("http://www.w3.org/2002/07/owl#inverseOf")
            owl_symmetric = URIRef("http://www.w3.org/2002/07/owl#SymmetricProperty")
            owl_transitive = URIRef("http://www.w3.org/2002/07/owl#TransitiveProperty")
            owl_equivalent_property = URIRef(
                "http://www.w3.org/2002/07/owl#equivalentProperty"
            )

            # 1. Inverse properties
            for other, pred, _ in self.model.instance_graph.triples(
                (None, None, self.uri)
            ):
                for inverse_pred in self.model.ontology_graph.objects(
                    pred, owl_inverse_of
                ):
                    inferred_pairs.append((inverse_pred, other))
                for inverse_pred in self.model.ontology_graph.subjects(
                    owl_inverse_of, pred
                ):
                    inferred_pairs.append((inverse_pred, other))

            # 2. Symmetric properties
            for other, pred, _ in self.model.instance_graph.triples(
                (None, None, self.uri)
            ):
                is_symmetric = (
                    pred,
                    RDF.type,
                    owl_symmetric,
                ) in self.model.ontology_graph
                if is_symmetric:
                    inferred_pairs.append((pred, other))

            # 3. Transitive properties
            for pred, obj in self.model.instance_graph.predicate_objects(self.uri):
                is_transitive = (
                    pred,
                    RDF.type,
                    owl_transitive,
                ) in self.model.ontology_graph
                if is_transitive:
                    for transitive_obj in self.model.instance_graph.transitive_objects(
                        obj, pred
                    ):
                        if transitive_obj != obj:
                            inferred_pairs.append((pred, transitive_obj))

            # 4. Equivalent properties
            for pred, obj in self.model.instance_graph.predicate_objects(self.uri):
                for equiv_pred in self.model.ontology_graph.objects(
                    pred, owl_equivalent_property
                ):
                    if equiv_pred != pred:
                        inferred_pairs.append((equiv_pred, obj))
                for equiv_pred in self.model.ontology_graph.subjects(
                    owl_equivalent_property, pred
                ):
                    if equiv_pred != pred:
                        inferred_pairs.append((equiv_pred, obj))

            ####################################################

            # Collect direct predicate-object pairs from the graph
            for pred, obj in self.model.instance_graph.predicate_objects(self.uri):
                if self._is_class_uri(obj):
                    obj_instance = self.model.get_type(obj)
                else:
                    obj_instance = self.model.get_instance(obj)

                pred_obj = self.model.get_predicate(pred)

                if pred_obj in self._attributes:
                    self._attributes[pred_obj].append(obj_instance)
                else:
                    self._attributes[pred_obj] = [obj_instance]

            # Add inferred predicate-object pairs from reasoning
            for pred, obj in inferred_pairs:
                if self._is_class_uri(obj):
                    obj_instance = self.model.get_type(obj)
                else:
                    obj_instance = self.model.get_instance(obj)

                pred_obj = self.model.get_predicate(pred)

                if pred_obj in self._attributes:
                    if obj_instance not in self._attributes[pred_obj]:
                        self._attributes[pred_obj].append(obj_instance)
                else:
                    self._attributes[pred_obj] = [obj_instance]

        return self._attributes

    def isinstance(
        self,
        cls: Union[
            str,
            SemanticType,
            Tuple[Union[str, SemanticType], ...],
            List[Union[str, SemanticType]],
        ],
    ) -> bool:
        """Check if this instance is of any of the given class types (including inheritance)"""
        if not isinstance(cls, (tuple, list)):
            cls = (cls,)

        if self.types:
            for c in cls:
                for instance_type in self.types:
                    if str(c) == str(instance_type.uri):
                        return True
        return False

    def get_most_specific_type(self) -> Optional["SemanticType"]:
        """
        Get the most specific type of this instance.
        We find the class in self.types that has ALL the other classes in self.types as super classes.

        This is not the same as finding the class in self.types that has none of the other
        classes in self.types as subclasses. (as initially implemented).
        This is because ontologies reuse classes from other ontologies by declaring
        ChildClass -> rdfs:subClassOf -> ParentClass. However, the ParentClass in the parent
        ontology does not necessarily have ChildClass as a subclass as this should be declared
        by the inheriting ontology.

        Note: Equivalent classes (owl:equivalentClass) are filtered out when determining the
        most specific type, as they are at the same level of specificity.

        Anonymous / Skolemized blank-node types (those without a recognized namespace) are
        excluded from the comparison, as they represent OWL restrictions or similar constructs
        that don't participate in the named class hierarchy.
        """
        if not self.types:
            return None

        # Filter to named types only (those with a recognized namespace)
        named_types = {t for t in self.types if t.get_short_name() is not None}
        if not named_types:
            return None

        for t in named_types:
            types_excluding_this = named_types - {t}

            equivalent_types_set = set(t.equivalent_classes)
            types_excluding_this = types_excluding_this - equivalent_types_set

            if all(t_ in t.super_classes for t_ in types_excluding_this):
                return t

        warnings.warn(f"No most specific type found for {self.get_short_name()}")
        return None


class SemanticLiteral(SemanticObject):
    """Represents an RDF Literal value in the semantic model."""

    def __init__(
        self,
        value: Union[str, Literal],
        model: "SemanticModel",
        datatype: Optional[URIRef] = None,
        lang: Optional[str] = None,
    ):
        # Don't call SemanticEntity.__init__ since we need a Literal, not a URIRef
        self.model = model
        self._namespace = (None, None)

        if isinstance(value, Literal):
            # Special handling for rdf:JSON literals
            if value.datatype and value.datatype == core.namespace.RDF.JSON:
                datatype = value.datatype
                lexical_form = str(value)
                if lexical_form.endswith("^^rdf:JSON"):
                    json_str = lexical_form[:-9]
                else:
                    json_str = lexical_form
                value = json.loads(json_str)

            elif str(value) == "None":
                value = None
                datatype = value.datatype if hasattr(value, "datatype") else None

            self.uri = Literal(value, datatype=datatype)

        elif datatype is not None or lang is not None:
            if datatype and datatype == core.namespace.RDF.JSON:
                lexical_form = str(value)
                if lexical_form.endswith("^^rdf:JSON"):
                    json_str = lexical_form[:-9]
                else:
                    json_str = lexical_form
                value = json.loads(json_str)
            elif str(value) == "None":
                value = None

            self.uri = Literal(value, datatype=datatype, lang=lang)
        else:
            self.uri = Literal(value)

        self._types = None

    @property
    def direct_types(self) -> Set[SemanticType]:
        """For literals, the direct type is the datatype."""
        return self.types

    @property
    def types(self) -> Set[SemanticType]:
        """For literals, return the datatype as the type."""
        if self._types is None:
            if self.uri.datatype:
                self._types = {self.model.get_type(self.uri.datatype)}
            else:
                self._types = {self.model.get_type(RDFS.Literal)}
        return self._types

    def isinstance(
        self,
        cls: Union[
            str,
            SemanticType,
            Tuple[Union[str, SemanticType], ...],
            List[Union[str, SemanticType]],
        ],
    ) -> bool:
        """Check if this literal's datatype matches any of the given types."""
        if not isinstance(cls, (tuple, list)):
            cls = (cls,)

        for c in cls:
            c_str = str(c)
            if self.uri.datatype and c_str == str(self.uri.datatype):
                return True
            if self.uri.language and c_str == str(core.namespace.XSD.string):
                return True
            if (
                not self.uri.datatype
                and not self.uri.language
                and c_str == str(core.namespace.XSD.string)
            ):
                return True
        return False

    def get_most_specific_type(self) -> Optional["SemanticType"]:
        """For literals, the most specific type is the datatype."""
        if self.types:
            return next(iter(self.types))
        return None

    def __repr__(self) -> str:
        return f"SemanticLiteral({repr(self.uri.value)})"


@autoreset_print
class SemanticModel:
    def __init__(
        self,
        rdf_file: Optional[str] = None,
        namespaces: Optional[Dict[str, str]] = None,
        format: Optional[str] = None,
        verbose=None,
        id: str = "semantic_model",
        dir_conf: List[str] = None,
    ):
        """
        Initialize the semantic model with separate graphs for instances and ontologies

        Args:
            rdf_file: Path or URL to the RDF file containing instance data
            namespaces: Optional additional namespace prefix-URI pairs
            format: Optional format specification ('xml', 'turtle', 'n3', 'nt', 'json-ld', etc.)
            verbose: Verbosity level
            id: Identifier for the semantic model
            dir_conf: Directory configuration for file storage
        """
        if verbose is not None:
            LOGGER.verbose = verbose

        self.id = id
        self.rdf_file = rdf_file
        if namespaces is None:
            namespaces = {}
        else:
            assert isinstance(
                namespaces, dict
            ), 'The "namespaces" argument must be a dictionary.'
            namespaces_ = {}
            for prefix, uri in namespaces.items():
                namespaces_[prefix.upper()] = Namespace(uri)
                setattr(self, prefix.upper(), namespaces_[prefix.upper()])
            namespaces = namespaces_
            # self.namespaces = namespaces
        self.format = format
        self.parsed_namespaces = set()
        self.error_namespaces = set()

        # Cache for instances
        self._instances = {}
        self._types = {}
        self._properties = {}
        self._predicates = {}

        if dir_conf is None:
            self.dir_conf = ["generated_files", "models", self.id, "semantic_model"]
        else:
            self.dir_conf = dir_conf

        if rdf_file is not None:
            if verbose:
                self._instance_graph, self._ontology_graph = self.get_graphs(
                    self.rdf_file, self.format
                )
                # self._ontology_graph = Graph()
            else:
                logging.disable(
                    sys.maxsize
                )  # https://stackoverflow.com/questions/2266646/how-to-disable-logging-on-the-standard-error-stream
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    self._instance_graph, self._ontology_graph = self.get_graphs(
                        self.rdf_file, self.format
                    )
                    # self._ontology_graph = Graph()
            filename_instance_graph = "raw_instance_graph.ttl"
            filename_ontology_graph = "raw_ontology_graph.ttl"
            self.serialize(
                filename_instance_graph=filename_instance_graph,
                filename_ontology_graph=filename_ontology_graph,
            )
            namespaces.update(dict(self._instance_graph.namespaces()))
        else:
            # Initialize both graphs
            self._instance_graph = Graph()
            self._ontology_graph = Graph()

        self.add_namespaces(namespaces)
        # logging.disable(logging.NOTSET)

    @property
    def namespaces(self):
        # print("Called namespaces property")
        n = {p.lower(): n for p, n in self._instance_graph.namespaces()}
        n.update({p.lower(): n for p, n in self._ontology_graph.namespaces()})
        # print("Namespaces: ", n)
        return n

    def __getattr__(self, prefix: str) -> URIRef:
        """
        Define how attributes are accessed that are not defined yet.
        https://stackoverflow.com/questions/4295678/understanding-the-difference-between-getattr-and-getattribute
        """
        if prefix.startswith("__"):  # ignore any special Python names!
            raise AttributeError(f"Prefixes cannot start with __")

        prefix = prefix.lower()

        if prefix not in self.namespaces:
            raise AttributeError(f"Prefix {prefix} not found in namespaces")

        return self.namespaces[prefix]

    @property
    def instance_graph(self):
        """Backwards compatibility property that returns instance_graph.

        Deprecated: Use instance_graph directly instead.
        """
        return self._instance_graph

    @property
    def ontology_graph(self):
        """Backwards compatibility property that returns instance_graph.

        Deprecated: Use ontology_graph directly instead.
        """
        return self._ontology_graph

    @staticmethod
    def get_graph_copy(graph):
        """Create a complete copy of the graph including namespace bindings.

        Returns:
            Graph: A new graph instance with all triples and namespace bindings copied
        """
        new_graph = Graph()

        # Copy all namespace bindings
        for prefix, namespace in graph.namespaces():
            new_graph.bind(prefix, namespace)

        # Copy all triples
        for s, p, o in graph.triples((None, None, None)):
            new_graph.add((s, p, o))
        return new_graph

    def add_namespaces(self, namespaces: Dict[str, Union[str, Namespace]]):
        # print("---------------------------------add_namespaces---------------------------------")
        for prefix, namespace in namespaces.items():
            if isinstance(namespace, str):
                namespace = Namespace(namespace)
            self._instance_graph.bind(prefix.lower(), namespace)
            self._ontology_graph.bind(prefix.lower(), namespace)
            # print(f"Added namespace {prefix}: {namespace}")

    def parse_namespaces(self, namespaces=None):
        """Parse namespaces dynamically on demand into ontology_graph.

        This method is called when ontology information is needed (e.g., class hierarchies).
        It first tries to use fallback ontologies from core.ontology, then parses
        namespace URIs directly if no fallback is available.

        Args:
            namespaces: Dict of {prefix: namespace} to parse. If None, parses all namespaces
                       from the ontology_graph.
        """
        # print(f"\n[DEBUG parse_namespaces] Called with namespaces: {namespaces}")
        # print(f"[DEBUG parse_namespaces] Already parsed: {self.parsed_namespaces}")
        # print(f"[DEBUG parse_namespaces] Already failed: {self.error_namespaces}")

        # LOGGER.verbose = 0 # TODO: Remove this

        overall_success = True

        if namespaces is None:
            namespaces = self.namespaces
            # print(f"[DEBUG parse_namespaces] Using all namespaces from ontology_graph: {list(namespaces.keys())}")

        for prefix, namespace in namespaces.items():
            uri = str(namespace)
            # print(f"\n[DEBUG parse_namespaces] Processing {prefix}: {uri}")

            # Skip if already parsed or previously failed
            if uri in self.parsed_namespaces or uri in self.error_namespaces:
                continue

            success = False

            LOGGER.info("Parsing namespace: %s (%s)", prefix.upper(), uri)
            LOGGER.add_level()

            # First, try to use fallback from core.ontology
            if hasattr(core.ontology, prefix.upper()):
                fallback_ontology_uri = getattr(core.ontology, prefix.upper())
                LOGGER.info(
                    "Attempting to parse namespace from core.ontology using URI: %s",
                    fallback_ontology_uri,
                )
                try:
                    parse_wrapper(self._ontology_graph, source=fallback_ontology_uri)
                    self.parsed_namespaces.add(uri)
                    LOGGER.ok(
                        "Attempting to parse namespace from core.ontology using URI: %s",
                        fallback_ontology_uri,
                        change_status=True,
                    )
                    success = True
                except Exception as e:
                    status = "[ERROR]"
                    success = False
                    LOGGER.add_level()
                    LOGGER.error("Error: %s", str(e))
                    LOGGER.remove_level()
                    LOGGER.error(
                        f"Attempting to parse namespace from core.ontology using URI: {fallback_ontology_uri}",
                        change_status=True,
                    )

            # If no fallback or fallback failed, try parsing namespace directly
            if not success:
                LOGGER.info(f"Attempting to parse namespace directly using URI: {uri}")
                try:
                    parse_wrapper(self._ontology_graph, source=namespace)
                    self.parsed_namespaces.add(uri)
                    success = True
                except HTTPError as http_err:
                    success = False
                    LOGGER.add_level()
                    LOGGER.error("HTTPError: %s", str(http_err))
                    LOGGER.error(
                        f"Sometimes this error occurs when the ontology is not available at the same address as the namespace.",
                    )
                    LOGGER.remove_level()

                except Exception as e:
                    success = False
                    LOGGER.add_level()
                    LOGGER.error("Error: %s", str(e))
                    LOGGER.remove_level()

                if success:
                    LOGGER.ok(
                        "Attempting to parse namespace directly using URI: %s",
                        uri,
                        change_status=True,
                    )
                else:
                    LOGGER.error(
                        "Attempting to parse namespace directly using URI: %s",
                        uri,
                        change_status=True,
                    )

            # If both methods failed, add to error namespaces
            if not success:
                self.error_namespaces.add(uri)
                # return False

            overall_success = overall_success and success

            if not overall_success:
                LOGGER.error(
                    "Parsing namespace: %s (%s)",
                    prefix.upper(),
                    uri,
                    change_status=True,
                )
            else:
                LOGGER.ok(
                    "Parsing namespace: %s (%s)",
                    prefix.upper(),
                    uri,
                    change_status=True,
                )
            LOGGER.remove_level()

        return overall_success

    def get_dir(
        self, folder_list: List[str] = None, filename: Optional[str] = None
    ) -> Tuple[str, bool]:
        """
        Get the directory path for storing model-related files.

        Args:
            folder_list (List[str]): List of folder names to create.
            filename (Optional[str]): Name of the file to create.

        Returns:
            Tuple[str, bool]: The full path to the directory or file, and a boolean indicating if the file exists.
        """
        if folder_list is None:
            folder_list = []
        folder_list_ = self.dir_conf.copy()
        folder_list_.extend(folder_list)
        filename, isfile = mkdir_in_root(folder_list=folder_list_, filename=filename)
        return filename, isfile

    def get_graphs(
        self,
        filename: str,
        format: Optional[str] = None,
        mappings_dir: Optional[str] = None,
    ) -> Graph:
        ext = os.path.splitext(filename)[1][1:].lower()
        if ext == "xlsx" or ext == "xlsm":

            if os.path.isfile(filename):
                # Get extension
                # ext = os.path.splitext(filename)[1][1:].lower()
                instance_graph, ontology_graph = self.parse_spreadsheet(
                    filename, mappings_dir=mappings_dir
                )
            else:
                raise FileNotFoundError(f"File {filename} not found.")
        else:
            instance_graph = Graph()
            ontology_graph = Graph()
            if format is None:
                parse_wrapper(instance_graph, source=filename)
            else:
                parse_wrapper(instance_graph, source=filename, format=format)

            # Copy schema definitions to ontology_graph
            # This ensures that even if schema and instances are mixed in the input file,
            # the ontology_graph contains the necessary schema definitions for reasoning.
            # We copy instead of move to preserve the integrity of the input file representation in instance_graph.
            # This was added because we want the serialized simulation model graph to be self-contained in 1 graph - the instance_graph.
            # If we dont have this copy logic, we would need to pass around 2 graphs, the ontology_graph AND the instance graph.

            # 1. Class definitions and hierarchy
            for s, p, o in instance_graph.triples((None, RDFS.subClassOf, None)):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples(
                (None, URIRef("http://www.w3.org/2002/07/owl#equivalentClass"), None)
            ):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples(
                (None, RDF.type, URIRef("http://www.w3.org/2002/07/owl#Class"))
            ):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples((None, RDF.type, RDFS.Class)):
                ontology_graph.add((s, p, o))

            # 2. Property definitions
            for s, p, o in instance_graph.triples(
                (None, RDF.type, URIRef("http://www.w3.org/2002/07/owl#ObjectProperty"))
            ):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples(
                (
                    None,
                    RDF.type,
                    URIRef("http://www.w3.org/2002/07/owl#DatatypeProperty"),
                )
            ):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples(
                (
                    None,
                    RDF.type,
                    URIRef("http://www.w3.org/1999/02/22-rdf-syntax-ns#Property"),
                )
            ):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples((None, RDFS.subPropertyOf, None)):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples(
                (None, URIRef("http://www.w3.org/2002/07/owl#equivalentProperty"), None)
            ):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples(
                (None, URIRef("http://www.w3.org/2002/07/owl#inverseOf"), None)
            ):
                ontology_graph.add((s, p, o))

            # 3. Domain and Range
            for s, p, o in instance_graph.triples((None, RDFS.domain, None)):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples((None, RDFS.range, None)):
                ontology_graph.add((s, p, o))

            # 4. Property characteristics
            for s, p, o in instance_graph.triples(
                (
                    None,
                    RDF.type,
                    URIRef("http://www.w3.org/2002/07/owl#TransitiveProperty"),
                )
            ):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples(
                (
                    None,
                    RDF.type,
                    URIRef("http://www.w3.org/2002/07/owl#SymmetricProperty"),
                )
            ):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples(
                (
                    None,
                    RDF.type,
                    URIRef("http://www.w3.org/2002/07/owl#FunctionalProperty"),
                )
            ):
                ontology_graph.add((s, p, o))

            for s, p, o in instance_graph.triples(
                (
                    None,
                    RDF.type,
                    URIRef("http://www.w3.org/2002/07/owl#InverseFunctionalProperty"),
                )
            ):
                ontology_graph.add((s, p, o))

        return instance_graph, ontology_graph

    def _get_node_alias(self, node, node_aliases):
        """Create a sequential node alias (Node1, Node2, etc.) in order of first visit.

        Args:
            node: The node URI to create an alias for
            node_aliases: Dictionary mapping node URIs to their aliases

        Returns:
            A sequential node alias
        """
        if node not in node_aliases:
            # Assign next sequential number
            next_num = len(node_aliases) + 1
            node_aliases[node] = f"Node{next_num}"
        return node_aliases[node]

    def _traverse(
        self,
        start_nodes: set,
        initial_triples: set,
        triple_limit: int = None,
        node_limit: int = None,
        mode: str = "bfs",
    ) -> set:
        r"""Traverse through initial_triples in either BFS or DFS order starting from given nodes.
        
        Example graph to illustrate traversal:
        ```
                D --> G --> I
                ^     ^
                |     |
                B --> E
              ↗       \
            /          \
          A             -> H
            \          /
              ↘       /
                C --> F
        ```
        
        The traversal is done in two passes:
        1. First process all subject relationships (where node is the subject)
           Starting at A:
           - A has outgoing edges to B and C
           - B has outgoing edges to D and E
           - C has outgoing edges to F
           - D has outgoing edge to G
           - E has outgoing edges to G and H
           - F has outgoing edge to H
           - G has outgoing edge to I
        
        2. Then process all object relationships (where node is the object)
           Starting at A:
           - Nothing points to A
           - B is pointed to by A
           - C is pointed to by A
           - D is pointed to by B
           - E is pointed to by B
           - F is pointed to by C
           - G is pointed to by D and E
           - H is pointed to by E and F
           - I is pointed to by G
        
        For each pass, the traversal behavior depends on mode:
        
        BFS (Breadth-First Search):
        - Explores level by level, visiting all nodes at current distance before moving deeper
        - New nodes are added to END of queue (append)
        - First pass (subject relationships):
            A -> discovers B, C     Queue: [A] -> [B, C]                Level 1
            B -> discovers D, E     Queue: [C] -> [C, D, E]             Level 2
            C -> discovers F        Queue: [D, E] -> [D, E, F]          Level 2
            D -> discovers G        Queue: [E, F] -> [E, F, G]          Level 3
            E -> discovers G, H     Queue: [F, G] -> [F, G, H]          Level 3
            F -> discovers H        Queue: [G, H] -> [G, H]             Level 3
            G -> discovers I        Queue: [H] -> [H, I]                Level 4
            Order: A, B, C, D, E, F, G, H, I (strict level order)
        - Second pass (object relationships):
            Similar pattern but following incoming edges, discovering multiple
            parents before moving to next level
        
        DFS (Depth-First Search):
        - Explores each path completely before backtracking
        - New nodes are added to FRONT of queue (insert(0))
        - First pass (subject relationships):
            A: pop A, add B, C -> [C, B]                # Path 1 start
            B: pop B, add D, E -> [C, E, D]            # Start path B
            D: pop D, add G -> [C, E, G]               # Follow D
            G: pop G, add I -> [C, E, I]               # Follow G
            I: pop I, no children -> [C, E]            # End G's path
            E: pop E, add G, H -> [C, H, G]            # Start path E
            G: already visited -> [C, H]               # Skip visited
            H: pop H, no children -> [C]               # End E's path
            C: pop C, add F -> [F]                     # Start path C
            F: pop F, add H -> [H]                     # Follow F
            H: already visited -> []                   # Done
            Order: A, B, D, G, I, E, H, C, F (follows each path to end)
        - Second pass (object relationships):
            Similar pattern but following incoming edges, completing each
            incoming path before moving to siblings
        
        Args:
            start_nodes: Initial set of nodes (subjects/objects) to start traversal from
            initial_triples: Set of triples from the SPARQL query to traverse through
            triple_limit: Maximum number of triples to collect
            node_limit: Maximum number of nodes to visit
            mode: Traversal mode - either "bfs" for breadth-first or "dfs" for depth-first
            
        Returns:
            Set of triples collected in specified traversal order
        """
        if not start_nodes or not initial_triples:
            return set()

        collected = set()  # Triples we've collected
        visited_nodes = set()  # Nodes we've seen

        subject_triples = {}
        object_triples = {}

        # Organize triples by subject and object
        for triple in initial_triples:
            s, _, o = triple
            if s not in subject_triples:
                subject_triples[s] = []
            subject_triples[s].append(triple)

            if o not in object_triples:
                object_triples[o] = []
            object_triples[o].append(triple)

        # # First pass: Process all subject relationships
        # nodes_to_visit = list(start_nodes)
        # triples_leading_to = []
        # traversed_nodes = set()

        # iteration = 0
        # while len(nodes_to_visit) > 0:
        #     iteration += 1
        #     node = nodes_to_visit.pop(0) if mode == "bfs" else nodes_to_visit.pop()
        #     if len(triples_leading_to) > 0:
        #         triple = triples_leading_to.pop(0) if mode == "bfs" else triples_leading_to.pop()
        #     else:
        #         triple = None

        #     visited_nodes.add(node)
        #     if triple is not None:
        #         collected.add(triple)

        #     if node_limit and len(visited_nodes) >= node_limit:
        #         return collected

        #     if triple_limit and len(collected) >= triple_limit:
        #         return collected

        #     if node not in traversed_nodes:
        #         traversed_nodes.add(node)

        #         # Process only subject relationships
        #         discovered_triples = []
        #         for triple_ in subject_triples.get(node, []):
        #             if triple_ not in collected:
        #                 s, p, o = triple_
        #                 discovered_triples.append((triple_, o))
        #                 # if o not in traversed_nodes:
        #                 if mode == "bfs":
        #                     nodes_to_visit.append(o)  # For BFS, add to end
        #                     triples_leading_to.append(triple_)
        #                 else:
        #                     # nodes_to_visit.insert(0, o)  # For DFS, add to front
        #                     # triples_leading_to.insert(0, triple_)
        #                     nodes_to_visit.append(o)  # For DFS, add to front
        #                     triples_leading_to.append(triple_)

        # # Second pass: Process all object relationships
        # nodes_to_visit = list(visited_nodes) #start_nodes
        # triples_leading_to = []
        # traversed_nodes = set()  # Reset traversed nodes for object pass

        # iteration = 0
        # while len(nodes_to_visit) > 0:
        #     iteration += 1
        #     node = nodes_to_visit.pop(0) if mode == "bfs" else nodes_to_visit.pop()

        #     if len(triples_leading_to) > 0:
        #         triple = triples_leading_to.pop(0) if mode == "bfs" else triples_leading_to.pop()
        #     else:
        #         triple = None

        #     visited_nodes.add(node)
        #     if triple is not None:
        #         collected.add(triple)

        #     if node_limit and len(visited_nodes) >= node_limit:
        #         return collected

        #     if triple_limit and len(collected) >= triple_limit:
        #         return collected

        #     if node not in traversed_nodes:
        #         traversed_nodes.add(node)

        #         # Process only object relationships
        #         discovered_triples = []
        #         for triple_ in object_triples.get(node, []):
        #             if triple_ not in collected:
        #                 s, p, o = triple_
        #                 discovered_triples.append((triple_, s))
        #                 if mode == "bfs":
        #                     nodes_to_visit.append(s)  # For BFS, add to end
        #                     triples_leading_to.append(triple_)
        #                 else:
        #                     nodes_to_visit.append(s)  # For DFS, add to front
        #                     triples_leading_to.append(triple_)

        # Third pass: Process all relationships
        nodes_to_visit = list(start_nodes)  # start_nodes
        triples_leading_to = []
        traversed_nodes = set()  # Reset traversed nodes for object pass

        iteration = 0
        while len(nodes_to_visit) > 0:
            iteration += 1
            node = nodes_to_visit.pop(0) if mode == "bfs" else nodes_to_visit.pop()

            if len(triples_leading_to) > 0:
                triple = (
                    triples_leading_to.pop(0)
                    if mode == "bfs"
                    else triples_leading_to.pop()
                )
            else:
                triple = None

            visited_nodes.add(node)
            if triple is not None:
                collected.add(triple)

            if node_limit and len(visited_nodes) >= node_limit:
                return collected

            if triple_limit and len(collected) >= triple_limit:
                return collected

            if node not in traversed_nodes:
                traversed_nodes.add(node)

                # Process only object relationships
                discovered_triples = []
                for triple_ in subject_triples.get(node, []):
                    if triple_ not in collected:
                        s, p, o = triple_
                        discovered_triples.append((triple_, o))
                        if mode == "bfs":
                            nodes_to_visit.append(o)
                            triples_leading_to.append(triple_)
                        else:
                            nodes_to_visit.append(o)
                            triples_leading_to.append(triple_)

                for triple_ in object_triples.get(node, []):
                    if triple_ not in collected:
                        s, p, o = triple_
                        discovered_triples.append((triple_, s))
                        if mode == "bfs":
                            nodes_to_visit.append(s)  # For BFS, add to end
                            triples_leading_to.append(triple_)
                        else:
                            nodes_to_visit.append(s)  # For DFS, add to front
                            triples_leading_to.append(triple_)

        print(f"Number of nodes visited: {len(visited_nodes)}")
        print(f"Number of traversed nodes: {len(traversed_nodes)}")
        print(f"Number of triples collected: {len(collected)}")
        return collected

    def filter_graph(
        self,
        query: str,
        triple_limit: int = None,
        node_limit: int = None,
        traversal_mode: str = None,
        initial_node: Optional[Union[str, URIRef]] = None,
        random_seed: Optional[int] = None,
    ) -> Graph:
        """Filter the graph based on class and predicate filters.
        The filtering is done using OR(class_filter) and OR(predicate_filter).

        Args:
            query: SPARQL CONSTRUCT query to filter the graph
            triple_limit: Maximum number of triples to include in the filtered graph
            node_limit: Maximum number of nodes to traverse when using BFS/DFS traversal.
                       This can help control the "spread" of the traversal
            traversal_mode: How to traverse the graph when collecting triples:
                          None - Use query result order (default)
                          "bfs" - Breadth-first search (explores nodes level by level)
                          "dfs" - Depth-first search (explores as far as possible along each branch)
            initial_node: Optional specific node to start traversal from. If not provided and traversal_mode
                         is used, a random node will be selected from the query results
            random_seed: Seed for random node selection when initial_node is not provided. Use this to get
                        reproducible results when doing random traversal

        Returns:
            Filtered graph

        Note:
            When using traversal modes (BFS/DFS), you can control the size of the result in two ways:
            1. triple_limit: Stops after collecting this many triples (relationships)
            2. node_limit: Stops after visiting this many nodes (entities)

            Using node_limit can be more intuitive as it directly controls how many entities are included,
            while triple_limit controls how many relationships are included.
        """
        assert traversal_mode in (
            "bfs",
            "dfs",
            None,
        ), "Traversal mode must be either 'bfs' or 'dfs' or None"

        new_graph = SemanticModel.get_graph_copy(self._instance_graph)
        # self.get_graph(self.rdf_file, self.format)
        keep_triples = set()

        if query is not None:
            if not query.strip().upper().startswith("CONSTRUCT"):
                raise ValueError("Query must start with CONSTRUCT")

            # Get initial triples from query
            result = self._instance_graph.query(query)
            initial_triples = set()
            candidate_nodes = set()

            # Convert initial_node to URIRef if provided as string
            if initial_node is not None:
                initial_node = (
                    URIRef(initial_node)
                    if isinstance(initial_node, str)
                    else initial_node
                )

            instances = set()

            for row in result:
                initial_triples.add(row)
                # Extract nodes (subject and object) from each triple
                s, _, o = row
                candidate_nodes.add(s)
                candidate_nodes.add(o)
                instances.add(s)
                instances.add(o)
                if not traversal_mode:
                    if node_limit is not None and len(instances) >= node_limit:
                        print(f"Reached node limit: {node_limit}")
                        break
                    if triple_limit is not None and len(instances) >= triple_limit:
                        print(f"Reached triple limit: {triple_limit}")
                        break

            # Determine start nodes for traversal
            if initial_node is not None:
                # Verify the node exists in our graph
                if not any(
                    self._instance_graph.triples((initial_node, None, None))
                ) and not any(self._instance_graph.triples((None, None, initial_node))):
                    raise ValueError(f"Initial node {initial_node} not found in graph")
                start_nodes = {initial_node}
            else:
                # Set random seed if provided
                if random_seed is not None:
                    random.seed(random_seed)
                # Convert to SORTED list for deterministic random selection
                # (sets have non-deterministic ordering, so we sort for reproducibility)
                node_list = sorted(candidate_nodes, key=str)
                if node_list:  # Only proceed if we have nodes
                    selected_node = random.choice(node_list)
                    start_nodes = {selected_node}
                else:
                    start_nodes = set()

            # Apply traversal if requested
            if traversal_mode in ("bfs", "dfs"):
                keep_triples = self._traverse(
                    start_nodes,
                    initial_triples,
                    triple_limit,
                    node_limit,
                    mode=traversal_mode,
                )
            else:
                keep_triples = initial_triples

            # Remove triples not in keep_triples.
            # Collect first to avoid modifying the graph during iteration.
            to_remove = [
                (s, p, o)
                for s, p, o in new_graph.triples((None, None, None))
                if (s, p, o) not in keep_triples
            ]
            for triple in to_remove:
                new_graph.remove(triple)

            # Count unique nodes in final graph
            final_nodes = set()
            for s, p, o in new_graph.triples((None, None, None)):
                final_nodes.add(s)
                final_nodes.add(o)

        return new_graph

    def get_instance(
        self,
        value: Union[str, URIRef, Literal],
        datatype: Optional[URIRef] = None,
        lang: Optional[str] = None,
    ) -> SemanticObject:
        """
        Get a specific instance by URI or create a literal.

        Factory method that returns SemanticInstance for URIs or SemanticLiteral for literals.

        Args:
            value: The URI or literal value
            datatype: Optional datatype URI for literals
            lang: Optional language tag for literals

        Returns:
            SemanticInstance for URI resources, SemanticLiteral for literal values
        """
        # Handle literals
        if isinstance(value, Literal) or datatype is not None or lang is not None:
            if isinstance(value, Literal):
                return SemanticLiteral(value, self)
            else:
                return SemanticLiteral(value, self, datatype=datatype, lang=lang)

        # Handle URIs
        uri = str(value)
        if uri not in self._instances:
            self._instances[uri] = SemanticInstance(uri, self)
        return self._instances[uri]

    def get_type(self, uri: Union[str, URIRef]) -> SemanticType:
        """Get a specific type by URI"""
        # uri = URIRef(uri) if isinstance(uri, str) else uri
        uri = str(uri)
        if uri not in self._types:
            self._types[uri] = SemanticType(uri, self)
        return self._types[uri]

    def get_property(self, uri: Union[str, URIRef]) -> "SemanticPredicate":
        """Get a specific property by URI.

        Deprecated: Use get_predicate() instead. This is a backward-compatible alias.
        """
        return self.get_predicate(uri)

    def get_predicate(self, uri: Union[str, URIRef]) -> SemanticPredicate:
        """Get a specific predicate by URI

        Args:
            uri: The URI of the predicate/property

        Returns:
            SemanticPredicate instance wrapping the predicate
        """
        # uri = URIRef(uri) if isinstance(uri, str) else uri
        uri = str(uri)
        if uri not in self._predicates:
            self._predicates[uri] = SemanticPredicate(uri, self)
        return self._predicates[uri]

    def get_instances_of_type(
        self, class_uris: Union[str, URIRef, SemanticType, Tuple, List]
    ) -> List[SemanticObject]:
        """
        Get all instances that match any of the specified types (including subtypes)

        Args:
            class_uris: Single URI or tuple/list of URIs representing the types to match

        Returns:
            List of SemanticObject instances that match any of the specified types
        """
        # Convert single class_uri to tuple for consistent handling
        if not isinstance(class_uris, tuple):
            if isinstance(class_uris, (list, set)):
                class_uris = tuple(class_uris)
            else:
                class_uris = (class_uris,)

        # print("GET INSTANCES OF TYPE")

        # Check if any of the requested types are XSD datatypes
        # If so, we need to find literals with those datatypes
        xsd_datatypes = []
        uri_types = []

        for class_uri in class_uris:
            if isinstance(class_uri, str):
                uri = URIRef(class_uri)
            elif isinstance(class_uri, SemanticType):
                uri = class_uri.uri
            elif isinstance(class_uri, URIRef):
                uri = class_uri
            else:
                raise ValueError(f"Invalid class URI: {class_uri}")

            # Check if this is an XSD datatype
            if str(uri).startswith("http://www.w3.org/2001/XMLSchema#"):
                xsd_datatypes.append(uri)
            else:
                uri_types.append(uri)

        instances = []
        processed_instances = set()  # To avoid duplicates

        # First, handle regular URI instances
        if uri_types:
            # Process each type in the tuple
            for uri in uri_types:
                # Get the class and all its subclasses using SemanticType
                t = self.get_type(uri)
                # Include the type itself and all its subclasses
                types_to_check = [t] + t.sub_classes + t.equivalent_classes

                # Get instances of the class and its subclasses from instance_graph
                for t_ in types_to_check:
                    # First check direct type assertions
                    for instance in self._instance_graph.subjects(RDF.type, t_.uri):
                        if instance not in processed_instances:
                            inst_obj = self.get_instance(instance)
                            instances.append(inst_obj)
                            processed_instances.add(instance)

                    # Then check for indirect type assertions through owl:sameAs
                    for instance in self._instance_graph.subjects(RDF.type, t_.uri):
                        for same_as in self._instance_graph.objects(
                            instance, URIRef("http://www.w3.org/2002/07/owl#sameAs")
                        ):
                            if same_as not in processed_instances:
                                inst_obj = self.get_instance(same_as)
                                instances.append(inst_obj)
                                processed_instances.add(same_as)

        # Then, handle literals with the specified datatypes
        if xsd_datatypes:
            # Find all literals in the instance graph with the specified datatypes
            for s, p, o in self._instance_graph.triples((None, None, None)):
                if isinstance(o, Literal) and o.datatype in xsd_datatypes:
                    # Create a SemanticObject for this literal
                    literal_obj = self.get_instance(o)
                    instances.append(literal_obj)

        return instances

    def count_instances(self) -> int:
        return len(list(self._instance_graph.subjects(RDF.type, None)))

    def count_triples(self, s=None, p=None, o=None) -> int:
        return len(list(self._instance_graph.triples((s, p, o))))

    def visualize(
        self,
        query=None,
        include_full_uri=True,
        slice_uri=None,
        dpi=2000,
        triple_limit=None,
        node_limit=None,
        generate_subgraphs=False,
        traversal_mode=None,
        initial_node=None,
        random_seed=None,
        format="svg",
        instance_style=None,
        deduplicate_inverse=True,
        pydot_transform=None,
    ):
        """
        Visualize RDF graph with optional class and predicate filtering.
        The filter acts as an OR filter.

        Args:
            query: SPARQL CONSTRUCT query to filter the graph
            include_full_uri: If True, include the last row with the full instance URI. If False, remove it.
            slice_uri: If provided, slice the URI string in row 1 (the main URI row). Can be:
                        - An integer: keep the last slice_uri characters
                        - A tuple (start, end): slice the URI using [start:end]
            dpi: DPI of the visualization. Only used if format is "png".
            limit: Limit the number of triples to visualize
            generate_subgraphs: If True, generate subgraphs for each isolated subgraph
            traversal_mode: Traversal mode to use. Can be "bfs" for breadth-first search or "dfs" for depth-first search
            initial_node: Initial node to start traversal from. If not provided, a random node will be selected from the query results
            random_seed: Random seed to use for random traversal. Use this to get reproducible results when doing random traversal
            format: Output format for the visualization. Can be "png" or "svg". Default is "svg".
            instance_style: Optional dict mapping instance URIs (str or URIRef) to style overrides.
                Per-instance styles take priority over type-based defaults.
                Each value is a dict with any of the following keys:

                - ``fill_color``: list of 4 colors ``[header, name_row, uri_row, literals]``
                - ``font_color``: list of 4 colors (same row order)
                - ``font_size``: list of 4 ints (same row order)
                - ``font_bold``: list of 4 bools (same row order)

                ``None`` entries inherit from the type-based defaults.

                Example::

                    instance_style = {
                        "http://example.org/myDamper1": {
                            "fill_color": ["#FF0000", "#FF0000", None, None],
                            "font_bold": [True, True, None, None],
                        }
                    }

            deduplicate_inverse: If True (default), when two instances are connected
                by both a predicate and its owl:inverseOf, only one edge is kept.
                The predicate with the lexicographically smaller URI is retained.
                This reduces visual clutter for ontologies that define inverse pairs
                (e.g. s4syst:connectedThrough / s4syst:connectsSystem).
            pydot_transform: Optional callable that receives the pydotplus graph
                object after node styling and can modify it in place before rendering.
        """
        # Omit rdf:type triples by default
        if query is None:
            query = """
            CONSTRUCT {
                ?s ?p ?o 
            }
            WHERE {
                ?s ?p ?o .
                FILTER (?p != rdf:type && 
                        ?p != rdfs:subClassOf)
            }
            """

        # --- Color palette (Paired / ColorBrewer) ---
        # Full palette for cycling; unmapped types pick from non-hardcoded colors
        dark_rose = "#8B4A6B"
        pink = "#FB9A99"
        dark_teal = "#2B7A78"
        light_purple = "#CAB2D6"
        purple = "#5B5EA6"
        brown = "#7A6855"
        white = "#FFFFFF"

        # Named aliases (easy to swap)
        light_black = "#3B3838"
        dark_blue = "#44546A"
        orange = "#DC8665"
        red = "#873939"
        grey = "#666666"
        light_blue = "#8497B0"
        green = "#83AF9B"
        magenta = "#660066"

        fill_color_cycle = [
            light_black,
            light_blue,
            dark_blue,
            dark_rose,
            green,
            pink,
            red,
            dark_teal,
            orange,
            light_purple,
            purple,
            brown,
        ]

        # Style lists are indexed by visual row order:
        #   [0] = Class type header       (e.g. "Damper")
        #   [1] = Short name              (e.g. "space42")
        #   [2] = Full URI                (e.g. "http://example.org/space42")
        #   [3] = Literal property rows   (e.g. "airFlowRateMax: 0.5")

        fill_color_map = {
            core.namespace.S4BLDG.BuildingSpace: [light_black, light_black, None, None],
            core.namespace.S4BLDG.Controller: [orange, orange, None, None],
            core.namespace.S4BLDG.AirToAirHeatRecovery: [
                dark_blue,
                dark_blue,
                None,
                None,
            ],
            core.namespace.S4BLDG.Coil: [red, red, None, None],
            core.namespace.S4BLDG.Damper: [dark_blue, dark_blue, None, None],
            core.namespace.S4BLDG.Valve: [red, red, None, None],
            core.namespace.S4BLDG.Fan: [dark_blue, dark_blue, None, None],
            core.namespace.S4BLDG.SpaceHeater: [red, red, None, None],
            core.namespace.SAREF.Sensor: [green, green, None, None],
            core.namespace.SAREF.Meter: [green, green, None, None],
            core.namespace.SAREF.Property: [light_purple, light_purple, None, None],
            core.namespace.S4BLDG.Schedule: [brown, brown, None, None],
            core.namespace.S4BLDG.Pump: [red, red, None, None],
            core.namespace.S4SYST.Connection: [grey, grey, None, None],
            core.namespace.S4SYST.ConnectionPoint: [grey, grey, None, None],
            core.namespace.T4B.BuildingSpaceTorchSystem: [
                light_black,
                light_black,
                None,
                None,
            ],
            core.namespace.T4B.PIDControllerSystem: [orange, orange, None, None],
            core.namespace.T4B.AirToAirHeatRecovery: [dark_blue, dark_blue, None, None],
            core.namespace.T4B.CoilTorchSystem: [red, red, None, None],
            core.namespace.T4B.DamperTorchSystem: [dark_blue, dark_blue, None, None],
            core.namespace.T4B.ValveTorchSystem: [red, red, None, None],
            core.namespace.T4B.FanTorchSystem: [dark_blue, dark_blue, None, None],
            core.namespace.T4B.SpaceHeaterTorchSystem: [red, red, None, None],
            core.namespace.T4B.SensorSystem: [green, green, None, None],
            core.namespace.T4B.ScheduleSystem: [brown, brown, None, None],
            core.namespace.T4B.PumpTorchSystem: [red, red, None, None],
        }

        # Remove colors already used by hardcoded mappings so the cycle only uses fresh ones
        _used_colors = {entry[0] for entry in fill_color_map.values()}
        _available_cycle = [c for c in fill_color_cycle if c not in _used_colors]
        if not _available_cycle:
            _available_cycle = fill_color_cycle  # fallback if all colors are taken

        def _cycle_color(type_uri):
            """Pick a fill color from the remaining (non-hardcoded) palette colors."""
            idx = hash(str(type_uri)) % len(_available_cycle)
            c = _available_cycle[idx]
            return [c, c, None, None]

        # Defaults for font color, size, bold (used when type is not in the map)
        _default_font_color = [white, white, None, None]
        _default_font_size = [10, 8, 8, 6]
        _default_font_bold = [True, True, None, None]

        font_color_map = {
            # All types currently use white; add per-type overrides here as needed
        }

        font_size_map = {
            core.namespace.S4SYST.Connection: [7, 6, 6, 5],
            core.namespace.S4SYST.ConnectionPoint: [7, 6, 6, 5],
        }

        font_bold_map = {
            core.namespace.S4SYST.Connection: [False, False, None, None],
            core.namespace.S4SYST.ConnectionPoint: [False, False, None, None],
        }

        # Build per-instance style overrides keyed by URI string
        _instance_fill = {}
        _instance_font_color = {}
        _instance_font_size = {}
        _instance_font_bold = {}
        if instance_style is not None:
            for inst_uri, style in instance_style.items():
                key = str(inst_uri)
                if "fill_color" in style:
                    _instance_fill[key] = style["fill_color"]
                if "font_color" in style:
                    _instance_font_color[key] = style["font_color"]
                if "font_size" in style:
                    _instance_font_size[key] = style["font_size"]
                if "font_bold" in style:
                    _instance_font_bold[key] = style["font_bold"]

        # Filter graph
        graph = self.filter_graph(
            query,
            triple_limit=triple_limit,
            node_limit=node_limit,
            traversal_mode=traversal_mode,
            initial_node=initial_node,
            random_seed=random_seed,
        )

        # Deduplicate inverse predicate pairs
        if deduplicate_inverse:
            owl_inverse_of = URIRef("http://www.w3.org/2002/07/owl#inverseOf")

            # Build a set of (p, p_inv) pairs from the ontology
            inverse_pairs = set()
            for p, _, p_inv in self.ontology_graph.triples(
                (None, owl_inverse_of, None)
            ):
                inverse_pairs.add((p, p_inv))
                inverse_pairs.add((p_inv, p))

            if inverse_pairs:
                triples_to_remove = []
                for s, p, o in graph:
                    for p_kept, p_dropped in inverse_pairs:
                        if p == p_dropped and (o, p_kept, s) in graph:
                            # Both (s, p_dropped, o) and (o, p_kept, s) exist.
                            # Keep the one whose predicate URI is lexicographically smaller.
                            if str(p_dropped) > str(p_kept):
                                triples_to_remove.append((s, p, o))
                            break

                for triple in triples_to_remove:
                    graph.remove(triple)

        stream = io.StringIO()
        rdf2dot(graph, stream)

        dg = pydotplus.graph_from_dot_data(stream.getvalue())

        # Add class type to node labels
        for node in dg.get_nodes():
            if node.obj_dict["name"] == "node":
                # del node.obj_dict["attributes"]["fontname"]
                node.obj_dict["attributes"]["fontname"] = "Courier-Bold"
            if "label" in node.obj_dict["attributes"]:
                html_str = node.obj_dict["attributes"]["label"]
                soup = BeautifulSoup(html_str, "html.parser")
                # Remove any existing width attribute that might conflict
                if "width" in soup.table.attrs:
                    del soup.table.attrs["width"]
                soup.table.attrs.update(
                    {
                        "BORDER": "2",
                        "CELLSPACING": "0",
                        "CELLPADDING": "2",
                        "CELLBORDER": "0",
                    }
                )
                row = soup.find_all("tr")[1]
                col = row.find_all("td")[0]
                uri = col.string
                inst = self.get_instance(uri)
                type_ = inst.types
                most_specific_type = inst.get_most_specific_type()

                # z_ = {e for e in type_ if e.has_subclasses() == False}
                # z = {e.uri.n3(self.graph.namespace_manager) for e in z_}
                # if len(z) == 0:
                #     z = {"Unknown class"}
                # z = " | ".join(z)  # data

                if most_specific_type is None:
                    z = "Unknown class"
                else:
                    z = most_specific_type.get_short_name()
                    assert (
                        z is not None
                    ), f"get_short_name() returned None for type {most_specific_type}"

                # Resolve the matching type key for this node (most specific first, then all types)
                matched_type_uri = None
                if (
                    most_specific_type is not None
                    and most_specific_type.uri in fill_color_map
                ):
                    matched_type_uri = most_specific_type.uri
                else:
                    for t in type_:
                        if t.uri in fill_color_map:
                            matched_type_uri = t.uri
                            break

                # Resolve effective style lists for this node.
                # Priority: instance_style > type-based map > cycle palette / defaults
                uri_str = str(uri)
                type_for_cycle = (
                    most_specific_type.uri if most_specific_type is not None else uri
                )

                def _get_fill(key):
                    if key is not None and key in fill_color_map:
                        return fill_color_map[key]
                    return _cycle_color(type_for_cycle)

                def _resolve_fill(inst_map):
                    base = _get_fill(matched_type_uri)
                    if uri_str not in inst_map:
                        return base
                    override = inst_map[uri_str]
                    return [o if o is not None else b for o, b in zip(override, base)]

                def _resolve(inst_map, type_map, default):
                    """Merge instance override with type default (None entries inherit)."""
                    base = (
                        type_map.get(matched_type_uri, default)
                        if matched_type_uri
                        else default
                    )
                    if uri_str not in inst_map:
                        return base
                    override = inst_map[uri_str]
                    return [o if o is not None else b for o, b in zip(override, base)]

                eff_fill = _resolve_fill(_instance_fill)

                eff_fc = _resolve(
                    _instance_font_color, font_color_map, _default_font_color
                )
                eff_fs = _resolve(
                    _instance_font_size, font_size_map, _default_font_size
                )
                eff_fb = _resolve(
                    _instance_font_bold, font_bold_map, _default_font_bold
                )

                b = soup.new_tag("b", attrs={})
                b.string = z
                header_bgcolor = eff_fill[0]
                header_font_color = eff_fc[0]
                if header_font_color is not None:
                    header_font = soup.new_tag(
                        "font", attrs={"color": header_font_color}
                    )
                    header_font.append(b)
                else:
                    header_font = b
                new_col = soup.new_tag(
                    "td",
                    attrs={
                        "BGCOLOR": header_bgcolor or "grey",
                        "COLSPAN": "2",
                        "ALIGN": "CENTER",
                        "VALIGN": "MIDDLE",
                    },
                )
                new_col.append(header_font)
                new_row = soup.new_tag("tr", attrs={})
                new_row.append(new_col)
                first_row = soup.find_all("tr")[0]
                first_row.insert_before(new_row)

                for i, row in enumerate(
                    soup.find_all("tr")
                ):  # [:-1]: #All except the last row, which is the full inst URI (small blue text)
                    # Skip the first row (header with class type) as it's already styled above
                    if i == 0:
                        continue

                    cols_in_row = row.find_all("td")

                    for col in cols_in_row:
                        i_ = i if i < 4 else 3

                        # Build uppercase TD attrs (BS4 html.parser lowercases them,
                        # but Graphviz HTML-like labels require uppercase).
                        td_attrs = {k.upper(): v for k, v in col.attrs.items()}
                        # Ensure single-cell rows span both columns
                        if len(cols_in_row) == 1:
                            td_attrs["COLSPAN"] = "2"
                        if eff_fill[i_] is not None:
                            td_attrs["BGCOLOR"] = eff_fill[i_]

                        # Build <font> attributes
                        font_attrs = {}
                        bold = False

                        fc = eff_fc
                        if fc[i_] is not None:
                            font_attrs["color"] = fc[i_]

                        fs = eff_fs
                        if fs[i_] is not None:
                            font_attrs["point-size"] = fs[i_]

                        fb = eff_fb
                        if fb[i_] is not None:
                            bold = fb[i_]

                        font = soup.new_tag("font", attrs=font_attrs)

                        # Extract text content from cell
                        if col.find("b"):
                            s = col.find("b").string
                        elif col.string:
                            s = col.string
                        else:
                            s = col.get_text()

                        # Replace old cell with a clean new <td> (uppercase attrs)
                        new_col = soup.new_tag("td", attrs=td_attrs)
                        col.replace_with(new_col)
                        new_col.append(font)

                        if bold:
                            s_ = soup.new_tag("b", attrs={})
                            s_.string = s
                            s = s_
                        font.append(s)

                # Remove the last row if include_full_uri is False
                if not include_full_uri:
                    all_rows = soup.find_all("tr")
                    if len(all_rows) > 0:
                        all_rows[2].decompose()

                # Slice the URI string in row 1 if slice_uri is provided
                if slice_uri is not None:
                    all_rows = soup.find_all("tr")
                    if len(all_rows) > 1:
                        uri_row = all_rows[1]
                        uri_col = uri_row.find_all("td")[0]

                        # Get the current URI string
                        if uri_col.string:
                            current_uri = str(uri_col.string)
                        elif uri_col.find("font"):
                            font_tag = uri_col.find("font")
                            if font_tag.find("b"):
                                current_uri = str(font_tag.find("b").string)
                            else:
                                current_uri = str(font_tag.string)
                        else:
                            current_uri = uri_col.get_text()

                        # Apply slicing based on slice_uri type
                        if isinstance(slice_uri, int):
                            # Keep the last slice_uri characters
                            sliced_uri = (
                                current_uri[-slice_uri:]
                                if slice_uri > 0
                                else current_uri[:slice_uri]
                            )
                        elif isinstance(slice_uri, tuple) and len(slice_uri) == 2:
                            # Slice using [start:end]
                            start, end = slice_uri
                            sliced_uri = current_uri[start:end]
                        else:
                            sliced_uri = current_uri

                        # Update the URI text with the sliced version
                        if uri_col.find("font"):
                            font_tag = uri_col.find("font")
                            if font_tag.find("b"):
                                font_tag.find("b").string = sliced_uri
                            else:
                                font_tag.string = sliced_uri
                        else:
                            uri_col.string = sliced_uri

                node.obj_dict["attributes"]["label"] = (
                    str(soup).replace("&lt;", "<").replace("&gt;", ">")
                )

        if pydot_transform is not None:
            pydot_transform(dg)

        def del_dir(dirname):
            for filename in os.listdir(dirname):
                file_path = os.path.join(dirname, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)

        dirname, _ = self.get_dir(folder_list=["graphs", "temp"])

        # Delete all files in dirname
        del_dir(dirname)
        dot_filename = os.path.join(dirname, "object_graph.dot")
        dg.write(dot_filename)

        ### ccomps ###
        dirname_ccomps, _ = self.get_dir(folder_list=["graphs", "temp", "ccomps"])
        dot_filename_ccomps = os.path.join(dirname_ccomps, "object_graph_ccomps.dot")
        del_dir(dirname_ccomps)
        app_path = shutil.which("ccomps")
        assert app_path is not None, "ccomps not found"
        args = [app_path, "-x", f"-o{dot_filename_ccomps}", f"{dot_filename}"]
        subprocess.run(args=args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        ### dot ###
        # Get all filenames generated in the folder dirname
        app_path = shutil.which("dot")
        assert app_path is not None, "dot not found. Is Graphviz installed?"
        filenames = []
        for filename in os.listdir(dirname_ccomps):
            file_path = os.path.join(dirname_ccomps, filename)
            if os.path.isfile(file_path):
                dot_filename_ccomps = file_path
                dot_filename_dot = os.path.join(
                    dirname_ccomps, filename.replace("ccomps", "dot")
                )
                dot_filename_ccomps_output = dot_filename_ccomps.replace(
                    ".dot", f".{format}"
                )
                args = [
                    app_path,
                    "-q",
                    f"-o{dot_filename_dot}",
                    f"{dot_filename_ccomps}",
                ]
                subprocess.run(
                    args=args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
                )
                if generate_subgraphs:
                    args = [
                        app_path,
                        f"-T{format}",
                        "-q",
                        f"-o{dot_filename_ccomps_output}",
                        f"{dot_filename_ccomps}",
                    ]
                    subprocess.run(
                        args=args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
                    )
                filenames.append(dot_filename_dot)

        dot_filename_ccomps = os.path.join(dirname, "object_graph_ccomps_joined.dot")
        with open(dot_filename_ccomps, "wb") as wfd:
            for f in filenames:
                with open(f, "rb") as fd:
                    shutil.copyfileobj(fd, wfd)

        ### gvpack ###
        dot_filename_gvpack = os.path.join(dirname, "object_graph_gvpack.dot")
        app_path = shutil.which("gvpack")
        assert app_path is not None, "gvpack not found"
        args = [
            app_path,
            "-array3",
            f"-o{dot_filename_gvpack}",
            f"{dot_filename_ccomps}",
        ]
        subprocess.run(args=args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        ### neato ###
        # Clean stale output files from previous runs (e.g. .svg when now rendering .png)
        graphs_dir, _ = self.get_dir(folder_list=["graphs"])
        for old_file in os.listdir(graphs_dir):
            if old_file.startswith("semantic_model.") and os.path.isfile(
                os.path.join(graphs_dir, old_file)
            ):
                os.remove(os.path.join(graphs_dir, old_file))

        semantic_model_output, _ = self.get_dir(
            folder_list=["graphs"], filename=f"semantic_model.{format}"
        )
        app_path = shutil.which("neato")
        assert app_path is not None, "neato not found"
        args = [
            app_path,
            f"-T{format}",
            "-n2",
            "-Gsize=10!",
            f"-Gdpi={dpi}",
            "-Grankdir=RL",
            "-q",
            # "-v", # verbose
            f"-o{semantic_model_output}",
            f"{dot_filename_gvpack}",
        ]
        subprocess.run(args=args, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    def parse_spreadsheet(self, spreadsheet, mappings_dir=None):
        """Parse spreadsheet into RDF graph using brickify tool"""

        LOGGER.add_level()
        LOGGER.info("Parsing spreadsheet")

        # Overwrite typer progress bar to prevent it from printing to the console.
        class Overwriter:
            def __init__(self, iterable, *args, **kwargs):
                self.iterable = iterable

            def __enter__(self):
                return self.iterable

            def __exit__(self, exc_type, exc_value, traceback):
                pass

        def overwriter(iterable_object, *args, **kwargs):
            return Overwriter(iterable_object, *args, **kwargs)

        table_handler.progressbar = overwriter

        instance_graph = Graph()
        ontology_graph = Graph()

        # Get directories for storing files
        dirname, _ = self.get_dir(folder_list=["mappings"])
        if mappings_dir is None:
            mappings_dir = os.path.join(
                uppath(os.path.abspath(__file__), 1), "mappings"
            )

        # Read common macros content
        common_macros_path = os.path.join(mappings_dir, "common_macros.yml")
        with open(common_macros_path, "r") as fd:
            common_macros_content = fd.read()

        # Load workbook and process each sheet
        wb = load_workbook(spreadsheet)

        # # temp #
        # for sheet in wb.sheetnames:
        #     csv_file = os.path.join(dirname, f"{sheet}.csv")
        #     df = pd.read_excel(spreadsheet, sheet_name=sheet)
        #     df.to_csv(csv_file, index=False)
        # ########
        # aa

        for sheet in wb.sheetnames:
            # Check if there's a corresponding config file
            config_file = os.path.join(mappings_dir, f"{sheet}.yml")
            if os.path.exists(config_file):
                # Create new config file with macros in dirname
                new_config_path = os.path.join(dirname, f"{sheet}.yml")
                with open(config_file, "rb") as fd:
                    with open(new_config_path, "wb") as wfd:
                        shutil.copyfileobj(fd, wfd)
                        wfd.write(b"\n")
                        wfd.write(common_macros_content.encode())

                # Convert sheet to CSV
                csv_file = os.path.join(dirname, f"{sheet}.csv")
                df = pd.read_excel(spreadsheet, sheet_name=sheet)
                df.to_csv(csv_file, index=False)

                # Run brickify with new config file
                output_file = os.path.join(dirname, f"{sheet}.ttl")

                # cmd = [
                #     "brickify",
                #     csv_file,
                #     "--output", output_file,
                #     "--input-type", "csv",
                #     "--config", new_config_path,
                #     "--building-prefix", "bldg",
                #     "--building-namespace", "http://example.org/building#"
                # ]

                # Parse the output file into our graph
                handler = table_handler.TableHandler(
                    source=csv_file,
                    input_format="csv",
                    config_file=new_config_path,
                )

                # Convert using the handler
                result_graph = handler.convert(
                    "bldg",
                    "http://example.org/building#",
                    "site",
                    "https://example.com/site#",
                )

                # Serialize to the output file
                result_graph.serialize(destination=str(output_file), format="turtle")

                try:
                    # subprocess.run(cmd, check=True)

                    # Parse the output file into our graph

                    if sheet == "Extensions":
                        parse_wrapper(
                            ontology_graph, source=output_file, format="turtle"
                        )
                    else:
                        parse_wrapper(
                            instance_graph, source=output_file, format="turtle"
                        )
                except subprocess.CalledProcessError as e:
                    print(f"Error running brickify for sheet {sheet}: {e}")
                except Exception as e:
                    print(f"Error processing sheet {sheet}: {e}")

        LOGGER.ok("Parsing spreadsheet", change_status=True)
        LOGGER.remove_level()
        return instance_graph, ontology_graph

    def reason(self, namespaces=None):
        """Perform RDFS and OWL reasoning to infer additional triples. Currently, we infer:
        - Inverse properties (owl:inverseOf)
        - Symmetric properties (owl:SymmetricProperty)
        - Transitive properties (owl:TransitiveProperty)
        - Subclass reasoning (rdfs:subClassOf)
        - Equivalent classes (owl:equivalentClass)
        - Equivalent properties (owl:equivalentProperty)
        - SameAs reasoning (owl:sameAs)

        This method reads ontology definitions from ontology_graph and instance data from
        instance_graph, and adds inferred triples to instance_graph.

        Note: For transitive equivalence chains (e.g., A≡B≡C), you may need to call
        reason() multiple times to fully propagate all inferences.
        """

        # import sys
        # print("=" * 80, file=sys.stderr)
        # print("REASON() METHOD CALLED", file=sys.stderr)
        # print("=" * 80, file=sys.stderr)

        # print(f"CURSES MODE: {LOGGER._curses_mode}", file=sys.stderr)
        LOGGER.info("Reasoning")
        LOGGER.add_level()

        if namespaces is None:
            # Convert Namespace objects to URI strings for parse_namespaces
            # namespaces = {prefix: str(uri) for prefix, uri in self.namespaces.items()}
            namespaces = self.namespaces

        # Parse namespaces into ontology_graph if not already done
        self.parse_namespaces(namespaces=namespaces)

        # Track new triples to add
        new_triples = set()

        # Handle inverse properties (read from ontology_graph, apply to instance_graph)
        for s, p, o in self._ontology_graph.triples(
            (None, URIRef("http://www.w3.org/2002/07/owl#inverseOf"), None)
        ):
            # Find and add inverse relationships
            for subj, _, obj in self._instance_graph.triples((None, s, None)):
                new_triple = (obj, o, subj)
                new_triples.add(new_triple)

            for subj, _, obj in self._instance_graph.triples((None, o, None)):
                new_triple = (obj, s, subj)
                new_triples.add(new_triple)

        n_triples = len(new_triples)
        LOGGER.info("Added number of inverse triples: %s", n_triples)

        # Handle symmetric properties
        symmetric_props = set(
            self._ontology_graph.subjects(
                RDF.type, URIRef("http://www.w3.org/2002/07/owl#SymmetricProperty")
            )
        )
        for prop in symmetric_props:
            for subj, _, obj in self._instance_graph.triples((None, prop, None)):
                new_triples.add((obj, prop, subj))

        LOGGER.info(f"Added number of symmetric triples: {len(new_triples)-n_triples}")
        n_triples = len(new_triples)

        # Handle transitive properties
        transitive_props = set(
            self._ontology_graph.subjects(
                RDF.type, URIRef("http://www.w3.org/2002/07/owl#TransitiveProperty")
            )
        )
        for prop in transitive_props:
            # Find all pairs connected by this property
            pairs = list(self._instance_graph.triples((None, prop, None)))
            # For each pair, look for additional connections
            for s1, _, o1 in pairs:
                for s2, _, o2 in pairs:
                    if o1 == s2:  # Found a chain
                        new_triples.add((s1, prop, o2))

        LOGGER.info(
            "Added number of transitive triples: %d", len(new_triples) - n_triples
        )
        n_triples = len(new_triples)

        # Handle subclass reasoning (rdfs:subClassOf)
        # This adds missing class assertions for instances based on complete class hierarchy

        # Execute subclass reasoning (read hierarchy from ontology_graph, apply to instance_graph)
        # For each class that has instances, get all its superclasses and propagate the type assertions
        classes_with_instances = set(self._instance_graph.objects(None, RDF.type))
        for class_uri in classes_with_instances:
            semantic_type = self.get_type(class_uri)
            # Get all superclasses using the SemanticType.super_classes property (handles transitive reasoning)
            for superclass_type in semantic_type.super_classes:
                # For each instance of this class, assert it's also an instance of all superclasses
                for instance in self._instance_graph.subjects(RDF.type, class_uri):
                    new_triples.add((instance, RDF.type, superclass_type.uri))

        LOGGER.info(
            "Added number of subclass triples: %d", len(new_triples) - n_triples
        )
        n_triples = len(new_triples)

        # Handle equivalent classes (owl:equivalentClass)
        owl_equiv_class = URIRef("http://www.w3.org/2002/07/owl#equivalentClass")

        # For each rdf:type assertion, check if the class has equivalent classes
        for instance, _, class_uri in self._instance_graph.triples(
            (None, RDF.type, None)
        ):
            # Find equivalent classes in both directions
            for equiv_class in self._ontology_graph.objects(class_uri, owl_equiv_class):
                if equiv_class != class_uri:
                    new_triples.add((instance, RDF.type, equiv_class))
            for equiv_class in self._ontology_graph.subjects(
                owl_equiv_class, class_uri
            ):
                if equiv_class != class_uri:
                    new_triples.add((instance, RDF.type, equiv_class))

        LOGGER.info(
            "Added number of equivalent class triples: %d", len(new_triples) - n_triples
        )
        n_triples = len(new_triples)

        # Handle equivalent properties (owl:equivalentProperty)
        owl_equiv_prop = URIRef("http://www.w3.org/2002/07/owl#equivalentProperty")

        # For each triple in the instance graph, check if its predicate has equivalent properties
        for subj, pred, obj in self._instance_graph.triples((None, None, None)):
            # Find equivalent properties in both directions
            for equiv_pred in self._ontology_graph.objects(pred, owl_equiv_prop):
                if equiv_pred != pred:
                    new_triples.add((subj, equiv_pred, obj))
            for equiv_pred in self._ontology_graph.subjects(owl_equiv_prop, pred):
                if equiv_pred != pred:
                    new_triples.add((subj, equiv_pred, obj))

        LOGGER.info(
            "Added number of equivalent property triples: %d",
            len(new_triples) - n_triples,
        )
        n_triples = len(new_triples)

        # These have not been tested yet

        # # Handle property chains (owl:propertyChainAxiom)
        # # This handles cases where property1 o property2 -> property3
        # for chain_prop, _, chain_list in ontology_graph.triples((None, URIRef("http://www.w3.org/2002/07/owl#propertyChainAxiom"), None)):
        #     if isinstance(chain_list, URIRef):
        #         # Handle simple property chains (property1 o property2 -> property3)
        #         # Find the chain properties and target property
        #         chain_properties = list(ontology_graph.objects(chain_list, RDF.first))
        #         target_property = list(ontology_graph.objects(chain_list, RDF.rest))

        #         if len(chain_properties) == 2 and target_property:
        #             prop1, prop2 = chain_properties[0], chain_properties[1]
        #             target_prop = target_property[0]

        #             # Find all chains: if a->b via prop1 and b->c via prop2, then a->c via target_prop
        #             for s1, _, o1 in self.graph.triples((None, prop1, None)):
        #                 for s2, _, o2 in self.graph.triples((None, prop2, None)):
        #                     if o1 == s2:  # Found a chain
        #                         new_triples.add((s1, target_prop, o2))

        # Handle sameAs reasoning (owl:sameAs)
        # This propagates all properties from one individual to its sameAs individuals
        for ind1, _, ind2 in self._instance_graph.triples(
            (None, URIRef("http://www.w3.org/2002/07/owl#sameAs"), None)
        ):
            # Copy all properties from ind1 to ind2
            for subj, pred, obj in self._instance_graph.triples((ind1, None, None)):
                if pred != URIRef(
                    "http://www.w3.org/2002/07/owl#sameAs"
                ):  # Avoid infinite loops
                    new_triples.add((ind2, pred, obj))
            # Copy all properties from ind2 to ind1
            for subj, pred, obj in self._instance_graph.triples((ind2, None, None)):
                if pred != URIRef(
                    "http://www.w3.org/2002/07/owl#sameAs"
                ):  # Avoid infinite loops
                    new_triples.add((ind1, pred, obj))

        # # Handle functional properties (owl:FunctionalProperty)
        # # If a property is functional, ensure we don't have conflicting values
        # functional_props = set(ontology_graph.subjects(RDF.type, URIRef("http://www.w3.org/2002/07/owl#FunctionalProperty")))
        # for prop in functional_props:
        #     # Group by subject and keep only the first value for each subject
        #     subject_values = {}
        #     for subj, _, obj in self.graph.triples((None, prop, None)):
        #         if subj not in subject_values:
        #             subject_values[subj] = obj
        #         # Note: In a real implementation, you might want to handle conflicts differently

        # # Handle inverse functional properties (owl:InverseFunctionalProperty)
        # # If a property is inverse functional, ensure we don't have conflicting subjects
        # inverse_functional_props = set(ontology_graph.subjects(RDF.type, URIRef("http://www.w3.org/2002/07/owl#InverseFunctionalProperty")))
        # for prop in inverse_functional_props:
        #     # Group by object and keep only the first subject for each object
        #     object_subjects = {}
        #     for subj, _, obj in self.graph.triples((None, prop, None)):
        #         if obj not in object_subjects:
        #             object_subjects[obj] = subj
        #         # Note: In a real implementation, you might want to handle conflicts differently

        # Add all new triples to the instance_graph
        for s, p, o in new_triples:
            self._instance_graph.add((s, p, o))

        LOGGER.remove_level()
        LOGGER.ok("Reasoning", change_status=True)

    def serialize(
        self,
        folder_list: List[str] = None,
        filename_ontology_graph: str = "ontology_graph.ttl",
        filename_instance_graph: str = "instance_graph.ttl",
    ):
        """Serialize the instance_graph to a file"""
        dirname_ontology_graph, _ = self.get_dir(
            folder_list=folder_list, filename=filename_ontology_graph
        )
        dirname_instance_graph, _ = self.get_dir(
            folder_list=folder_list, filename=filename_instance_graph
        )
        for uri, instance in self._instances.items():

            # Added inferred types
            type_ = instance.types
            for type_ in type_:
                self._instance_graph.add((URIRef(uri), RDF.type, type_.uri))

            # Add inferred predicate-object pairs
            if instance._attributes is not None:
                for pred, obj in instance.get_predicate_object_pairs().items():
                    for obj_ in obj:
                        self._instance_graph.add((URIRef(uri), pred.uri, obj_.uri))

        self._ontology_graph.serialize(
            destination=str(dirname_ontology_graph), format="turtle"
        )
        self._instance_graph.serialize(
            destination=str(dirname_instance_graph), format="turtle"
        )
