import networkx as nx
import pandas as pd
import warnings
import shutil
import subprocess
import sys
import os
import copy
import pydot
import inspect
import numpy as np
import pandas as pd
import datetime
import torch
import json
import builtins
import pickle
import matplotlib.font_manager
from itertools import count
from prettytable import PrettyTable
from twin4build.utils.print_progress import PrintProgress
from openpyxl import load_workbook
from twin4build.utils.fmu.fmu_component import FMUComponent
from twin4build.utils.isnumeric import isnumeric
from twin4build.utils.get_object_attributes import get_object_attributes
from twin4build.utils.mkdir_in_root import mkdir_in_root
from twin4build.utils.rsetattr import rsetattr
from twin4build.utils.rgetattr import rgetattr
from twin4build.utils.rhasattr import rhasattr
from twin4build.utils.istype import istype
from twin4build.utils.data_loaders.load_spreadsheet import sample_from_df
from twin4build.saref4syst.connection import Connection 
from twin4build.saref4syst.connection_point import ConnectionPoint
from twin4build.saref4syst.system import System
import twin4build.utils.signature_pattern.signature_pattern as signature_pattern
import twin4build.base as base
import twin4build.systems as systems
import twin4build.estimator.estimator as estimator
from typing import List, Dict, Any, Optional, Tuple, Type, Callable
from twin4build.utils.simple_cycle import simple_cycles
import twin4build.utils.input_output_types as tps

def str2Class(str):
    return getattr(sys.modules[__name__], str)



class Graph:
    def __init__(self):
        self.system_graph = None
        self.object_graph = None
        self.execution_graph = None
        self.system_graph_no_cycles = None
        self.system_subgraph_dict = {}
        self.object_subgraph_dict = {}
        self.system_graph_node_attribute_dict = {}
        self.object_graph_node_attribute_dict = {}
        
    def create_graph(self, graph_type: str, attributes: Dict, subgraphs: Dict, add_brackets: bool = False) -> None:
        # Implementation from _create_graph method
        pass

    def _get_font(self) -> Tuple[str, str]:
        """
        Get the font path and name for graph labels.

        Returns:
            Tuple[str, str]: The font path and font name.
        """
        font_files = matplotlib.font_manager.findSystemFonts(fontpaths=None)
        preferred_font = "Helvetica-Bold".lower()
        preferred_font = "CMUNBTL".lower()
        found_preferred_font = False
        for font in font_files:
            s = os.path.split(font)
            if preferred_font in s[1].lower():
                found_preferred_font = True
                break

        if found_preferred_font==False:
            font = matplotlib.font_manager.findfont(matplotlib.font_manager.FontProperties(family=['sans-serif']))

        fontpath = os.path.split(font)[0]
        fontname = os.path.split(font)[1]

        fontname = "CMU Typewriter Text"

        return fontpath, fontname
        
    def draw_graph(self, filename: str, graph: pydot.Dot, args: Optional[List[str]] = None) -> None:
        """
        Draw a graph and save it to a file.

        Args:
            filename (str): The filename to save the graph to.
            graph (pydot.Dot): The graph to draw.
            args (Optional[List[str]]): Additional arguments for the graph drawing command.
        """
        fontpath, fontname = self._get_font()
        beige = "#F2EADD"
        light_grey = "#71797E"
        graph_filename = os.path.join(self.graph_path, f"{filename}.png")
        graph.write(f'{filename}.dot', prog="dot")
        app_path = shutil.which("dot")
        if args is None:
            args = [app_path,
                    "-q",
                    "-Tpng",
                    "-Kdot",
                    f"-Gfontpath={fontpath}",
                    "-Nstyle=filled",
                    "-Nshape=box",
                    "-Nfontcolor=white",
                    f"-Nfontname=Helvetica bold",
                    "-Nfixedsize=true",
                    "-Gnodesep=0.1",
                    "-Efontname=Helvetica",
                    "-Efontsize=21",
                    "-Epenwidth=2",
                    "-Eminlen=1",
                    f"-Ecolor={light_grey}",
                    "-Gcompound=true",
                    "-Grankdir=TB",
                    "-Gsplines=true", #true
                    "-Gmargin=0",
                    "-Gsize=10!",
                    "-Gratio=compress", #0.5 #auto
                    "-Gpack=true",
                    "-Gdpi=1000", #5000 for large graphs
                    "-Grepulsiveforce=0.5",
                    "-Gremincross=true",
                    "-Gstart=1",
                    "-Gbgcolor=transparent",
                    "-q",
                    f"-o{graph_filename}",
                    f"{filename}.dot"] #__unflatten
        else:
            args_ = [app_path]
            args_.extend(args)
            args_.extend([f"-o{graph_filename}", f"{filename}.dot"])
            args = args_
        subprocess.run(args=args)
        os.remove(f"{filename}.dot")

class Model:
    """
    A class representing a building system model.

    This class is responsible for creating, managing, and simulating a building system model.
    It handles component instantiation, connections between components, and execution order
    for simulation.

    Attributes:
        id (str): Unique identifier for the model.
        saveSimulationResult (bool): Flag to determine if simulation results should be saved.
        components (dict): Dictionary of all components in the model.
        object_dict (dict): Dictionary of all objects in the model.
        system_dict (dict): Dictionary of systems in the model (ventilation, heating, cooling).
        execution_order (list): Ordered list of component groups for execution.
        flat_execution_order (list): Flattened list of components in execution order.
    """

    __slots__ = (
        'id', 'saveSimulationResult', 'components', 'component_base_dict',
        'system_dict', 'object_dict', 'object_dict_reversed', 'object_counter_dict',
        'property_dict', 'instance_map', 'instance_map_reversed', 'instance_to_group_map',
        'custom_initial_dict', 'execution_order', 'flat_execution_order',
        'required_initialization_connections', '_components_no_cycles',
        'activeComponents', 'system_graph', 'object_graph', 'execution_graph',
        'system_graph_no_cycles', 'system_subgraph_dict_no_cycles',
        'object_graph_no_cycles', 'object_subgraph_dict_no_cycles',
        'system_graph_edge_counter', 'object_graph_edge_counter',
        'system_subgraph_dict', 'object_subgraph_dict',
        'system_graph_node_attribute_dict', 'object_graph_node_attribute_dict',
        'system_graph_edge_label_dict', 'object_graph_edge_label_dict',
        'system_graph_rank', 'object_graph_rank', 'is_loaded', 'result',
        'valid_chars', 'graph_path',"heatexchanger_types", "p", "validated", "validated_for_simulator", "validated_for_estimator",
        "validated_for_evaluator", "validated_for_monitor",
    )


    def __str__(self):
        t = PrettyTable(["Number of components in simulation model: ", len(self.components)])
        t.add_row(["Number of edges in simulation model: ", self.system_graph_edge_counter], divider=True)
        title = f"Model overview    id: {self.id}"
        t.title = title
        t.add_row(["Number of objects in semantic model: ", len(self.object_dict)], divider=True)
        t.add_row(["Number of triples in semantic model: ", self.object_graph_edge_counter], divider=True)
        t.add_row(["", ""])
        t.add_row(["", ""], divider=True)
        t.add_row(["id", "Class"], divider=True)
        unique_class_list = []
        for component in self.components.values():
            cls = component.__class__
            if cls not in unique_class_list:
                unique_class_list.append(cls)
        unique_class_list = sorted(unique_class_list, key=lambda x: x.__name__.lower())

        for cls in unique_class_list:
            cs = self.get_component_by_class(self.components, cls, filter=lambda v, class_: v.__class__ is class_)
            n = len(cs)
            for i,c in enumerate(cs):
                t.add_row([c.id, cls.__name__], divider=True if i==n-1 else False)
            
        return t.get_string()

    def __init__(self, id: str, saveSimulationResult: bool = True) -> None:
        """
        Initialize the Model instance.

        Args:
            id (str): Unique identifier for the model.
            saveSimulationResult (bool): Flag to determine if simulation results should be saved.

        Raises:
            AssertionError: If the id is not a string or contains invalid characters.
        """
        self.valid_chars = ["_", "-", " ", "(", ")", "[", "]"]
        assert isinstance(id, str), f"Argument \"id\" must be of type {str(type(str))}"
        isvalid = np.array([x.isalnum() or x in self.valid_chars for x in id])
        np_id = np.array(list(id))
        violated_characters = list(np_id[isvalid==False])
        assert all(isvalid), f"The model with id \"{id}\" has an invalid id. The characters \"{', '.join(violated_characters)}\" are not allowed."
        self.id = id
        self.saveSimulationResult = saveSimulationResult
        self._initialize_graph("system")
        self._initialize_graph("object")

        self.system_dict = {"ventilation": {},
                            "heating": {},
                            "cooling": {},
                            }
        self.component_base_dict = {} #Subset of object_dict
        self.components = {} #Subset of object_dict
        self.object_dict = {}
        self.object_dict_reversed = {}
        self.object_counter_dict = {}
        self.property_dict = {}
        self.custom_initial_dict = None
        self.heatexchanger_types = (base.AirToAirHeatRecovery, base.Coil)
        self.is_loaded = False
        self.validated = False

        self.instance_to_group_map = {}

        self.graph_path, isfile = self.get_dir(folder_list=["graphs"])


    def _get_font(self) -> Tuple[str, str]:
        """
        Get the font path and name for graph labels.

        Returns:
            Tuple[str, str]: The font path and font name.
        """
        font_files = matplotlib.font_manager.findSystemFonts(fontpaths=None)
        preferred_font = "Helvetica-Bold".lower()
        preferred_font = "CMUNBTL".lower()
        found_preferred_font = False
        for font in font_files:
            s = os.path.split(font)
            if preferred_font in s[1].lower():
                found_preferred_font = True
                break

        if found_preferred_font==False:
            font = matplotlib.font_manager.findfont(matplotlib.font_manager.FontProperties(family=['sans-serif']))

        fontpath = os.path.split(font)[0]
        fontname = os.path.split(font)[1]

        fontname = "CMU Typewriter Text"

        return fontpath, fontname

    @property
    def component_dict(self) -> dict:
        """
        Deprecated property that provides backward compatibility for accessing components.
        Will be removed.
        
        Returns:
            dict: Dictionary of all components in the model
        """
        warnings.warn(
            "component_dict is deprecated and will be removed."
            "Use components instead.",
            DeprecationWarning,
            stacklevel=2
        )
        return self.components

    @component_dict.setter
    def component_dict(self, value: dict) -> None:
        """
        Deprecated setter for component_dict that maintains backward compatibility.
        Will be removed.
        
        Args:
            value (dict): Dictionary of components to set
        """
        warnings.warn(
            "component_dict is deprecated and will be removed."
            "Use components instead.",
            DeprecationWarning,
            stacklevel=2
        )
    
    
    def get_dir(self, folder_list: List[str] = [], filename: Optional[str] = None) -> Tuple[str, bool]:
        """
        Get the directory path for storing model-related files.

        Args:
            folder_list (List[str]): List of folder names to create.
            filename (Optional[str]): Name of the file to create.

        Returns:
            Tuple[str, bool]: The full path to the directory or file, and a boolean indicating if the file exists.
        """
        f = ["generated_files", "models", self.id]
        f.extend(folder_list)
        folder_list = f
        filename, isfile = mkdir_in_root(folder_list=folder_list, filename=filename)
        return filename, isfile

    def _add_edge(self, 
                  graph: pydot.Dot, a: str, 
                  b: str, 
                  sender_property_name: Optional[str] = None, 
                  receiver_property_name: Optional[str] = None, 
                  edge_kwargs: Optional[Dict] = None) -> None:
        """
        Add an edge to the graph.

        Args:
            graph (pydot.Dot): The graph to add the edge to.
            a (str): Source node name.
            b (str): Target node name.
            sender_property_name (Optional[str]): Name of the sender property.
            receiver_property_name (Optional[str]): Name of the receiver property.
            edge_kwargs (Optional[Dict]): Additional edge attributes.
        """
        if edge_kwargs is None:
            edge_label = self._get_edge_label(sender_property_name, receiver_property_name)
            edge_kwargs = {"label": edge_label}
            graph.add_edge(pydot.Edge(a, b, **edge_kwargs))
        else:
            graph.add_edge(pydot.Edge(a, b, **edge_kwargs))

    def _del_edge(self, graph: pydot.Dot, a: str, b: str, label: str) -> bool:
        """
        Delete an edge from the graph.

        Args:
            graph (pydot.Dot): The graph to delete the edge from.
            a (str): Source node name.
            b (str): Target node name.
            label (str): Edge label to match.

        Returns:
            bool: True if the edge was successfully deleted, False otherwise.

        Raises:
            AssertionError: If more than one matching edge is found.
        """
        if pydot.needs_quotes(a):
            a = f"\"{a}\""
        if pydot.needs_quotes(b):
            b = f"\"{b}\""

        edges = graph.get_edge(a, b)
        is_matched = [el.obj_dict["attributes"]["label"]==label for el in edges]
        match_idx = [i for i, x in enumerate(is_matched) if x]
        assert len(match_idx)==1, "Wrong length"
        status = graph.del_edge(a, b, match_idx[0])
        return status
    
    def _del_node(self, graph: pydot.Dot, node: str) -> bool:

        if pydot.needs_quotes(node):
            node = f"\"{node}\""
        status = graph.del_node(node)
        return status


    def _add_component(self, component: System) -> None:
        """
        Add a component to the model.

        Args:
            component (System): The component to add.

        Raises:
            AssertionError: If the component is not an instance of System.
        """
        assert isinstance(component, System), f"The argument \"component\" must be of type {System.__name__}"
        if component.id not in self.components:
            self.components[component.id] = component

        self._add_object(component)

    def _get_new_object_name(self, obj: Any) -> str:
        """
        Generate a new unique name for an object.

        Args:
            obj (Any): The object to generate a name for.

        Returns:
            str: A unique name for the object.
        """
        if "id" not in get_object_attributes(obj):
            if obj.__class__.__name__ not in self.object_counter_dict:
                self.object_counter_dict[obj.__class__.__name__] = 0
            name = f"{obj.__class__.__name__.lower()} {str(self.object_counter_dict[obj.__class__.__name__])}"# [{id(obj)}]"
            self.object_counter_dict[obj.__class__.__name__] += 1
        else:
            name = obj.id
        return name

    def make_pickable(self) -> None:
        """
        Make the model instance pickable by removing unpickable references.

        This method prepares the Model instance for use with multiprocessing in the Estimator class.
        """
        self.object_dict = {} 
        self.object_dict_reversed = {}
        fmu_components = self.get_component_by_class(self.components, FMUComponent)
        for fmu_component in fmu_components:
            if "fmu" in get_object_attributes(fmu_component):
                # fmu_component.fmu.freeInstance()
                # fmu_component.fmu.terminate()
                del fmu_component.fmu
                del fmu_component.fmu_initial_state
                fmu_component.INITIALIZED = False

    def _add_object(self, obj: Any) -> None:
        """
        Add an object to the model's object dictionaries.

        Args:
            obj (Any): The object to add.
        """
        if obj in self.components.values() or obj in self.component_base_dict.values():
            name = obj.id
            self.object_dict[name] = obj
            self.object_dict_reversed[obj] = name
        elif obj not in self.object_dict_reversed:
            name = self._get_new_object_name(obj)
            self.object_dict[name] = obj
            self.object_dict_reversed[obj] = name

    def remove_component(self, component: System) -> None:
        """
        Remove a component from the model.

        Args:
            component (System): The component to remove.
        """
        for connection in component.connectedThrough:
            for connection_point in connection.connectsSystemAt:
                connected_component = connection_point.connectionPointOf
                self.remove_connection(component, connected_component, connection.senderPropertyName, connection_point.receiverPropertyName)
                connection.connectsSystem = None

        for connection_point in component.connectsAt:
            connection_point.connectPointOf = None
        
        del self.components[component.id]
        #Remove from subgraph dict
        subgraph_dict = self.system_subgraph_dict
        component_class_name = component.__class__
        if component_class_name in subgraph_dict:
            status = self._del_node(subgraph_dict[component_class_name], component.id)
            subgraph_dict[component_class_name].del_node(component.id)

    def _get_edge_label(self, sender_property_name: str, receiver_property_name: str) -> str:
        """
        Generate a label for an edge in the graph.

        Args:
            sender_property_name (str): Name of the sender property.
            receiver_property_name (str): Name of the receiver property.

        Returns:
            str: The formatted edge label.
        """
        end_space = "          "
        # edge_label = ("Out: " + sender_property_name.split("_")[0] + end_space + "\n"
        #                 "In: " + receiver_property_name.split("_")[0] + end_space)
        edge_label = ("Out: " + sender_property_name + end_space + "\n"
                        "In: " + receiver_property_name + end_space)
        return edge_label

    def add_connection(self, sender_component: System, receiver_component: System, 
                       sender_property_name: str, receiver_property_name: str) -> None:
        """
        Add a connection between two components in the system.

        Args:
            sender_component (System): The component sending the connection.
            receiver_component (System): The component receiving the connection.
            sender_property_name (str): Name of the sender property.
            receiver_property_name (str): Name of the receiver property.
        Raises:
            AssertionError: If property names are invalid for the components.
            AssertionError: If a connection already exists.
        """
        self._add_component(sender_component)
        self._add_component(receiver_component)

        found_connection_point = False
        # Check if there already is a connectionPoint with the same receiver_property_name
        for receiver_component_connection_point in receiver_component.connectsAt:
            if receiver_component_connection_point.receiverPropertyName == receiver_property_name:
                found_connection_point = True
                break
        
        
        found_connection = False
        # Check if there already is a connection with the same sender_property_name
        for sender_obj_connection in sender_component.connectedThrough:
            if sender_obj_connection.senderPropertyName == sender_property_name:
                found_connection = True
                break

        if found_connection_point and found_connection:
            message = f"Connection between \"{sender_component.id}\" and \"{receiver_component.id}\" with the properties \"{sender_property_name}\" and \"{receiver_property_name}\" already exists."
            assert receiver_component_connection_point not in sender_obj_connection.connectsSystemAt, message
                    

        if found_connection==False:
            sender_obj_connection = Connection(connectsSystem=sender_component, senderPropertyName=sender_property_name)
            sender_component.connectedThrough.append(sender_obj_connection)

        if found_connection_point==False:
            receiver_component_connection_point = ConnectionPoint(connectionPointOf=receiver_component, receiverPropertyName=receiver_property_name)
            receiver_component.connectsAt.append(receiver_component_connection_point)
        
        sender_obj_connection.connectsSystemAt.append(receiver_component_connection_point)
        receiver_component_connection_point.connectsSystemThrough.append(sender_obj_connection)# if sender_obj_connection not in receiver_component_connection_point.connectsSystemThrough else None


        # Inputs and outputs of these classes can be set dynamically. Inputs and outputs of classes not in this tuple are set as part of their class definition.
        exception_classes = (systems.TimeSeriesInputSystem,
                             systems.PiecewiseLinearSystem,
                             systems.PiecewiseLinearSupplyWaterTemperatureSystem,
                             systems.PiecewiseLinearScheduleSystem,
                             base.Sensor,
                             base.Meter,
                             systems.MaxSystem,
                             systems.NeuralPolicyControllerSystem) 
        
        if isinstance(sender_component, exception_classes):
            if sender_property_name not in sender_component.output:
                # If the property is not already an output, we assume it is a Scalar
                sender_component.output.update({sender_property_name: tps.Scalar()})
            else:
                pass
        else:
            message = f"The property \"{sender_property_name}\" is not a valid output for the component \"{sender_component.id}\" of type \"{type(sender_component)}\".\nThe valid output properties are: {','.join(list(sender_component.output.keys()))}"
            assert sender_property_name in (set(sender_component.input.keys()) | set(sender_component.output.keys())), message
        
        if isinstance(receiver_component, exception_classes):
            if receiver_property_name not in receiver_component.input:
                # If the property is not already an input, we assume it is a Scalar
                receiver_component.input.update({receiver_property_name: tps.Scalar()})
            else:
                assert isinstance(receiver_component.input[receiver_property_name], tps.Vector), f"The input property \"{receiver_property_name}\" for the component \"{receiver_component.id}\" of type \"{type(receiver_component)}\" is already set as a Scalar input."
        else:
            message = f"The property \"{receiver_property_name}\" is not a valid input for the component \"{receiver_component.id}\" of type \"{type(receiver_component)}\".\nThe valid input properties are: {','.join(list(receiver_component.input.keys()))}"
            assert receiver_property_name in receiver_component.input.keys(), message
        
        self._add_graph_relation(graph=self.system_graph, sender_component=sender_component, receiver_component=receiver_component, sender_property_name=sender_property_name, receiver_property_name=receiver_property_name)

    def _add_graph_relation(self, graph: pydot.Dot, sender_component: System, receiver_component: System, 
                            sender_property_name: Optional[str] = None, receiver_property_name: Optional[str] = None, 
                            edge_kwargs: Optional[Dict] = None, sender_node_kwargs: Optional[Dict] = None, 
                            receiver_node_kwargs: Optional[Dict] = None) -> None:
        """
        Add a relation between components to the graph.

        Args:
            graph (pydot.Dot): The graph to add the relation to.
            sender_component (System): The sending component.
            receiver_component (System): The receiving component.
            sender_property_name (Optional[str]): Name of the sender property.
            receiver_property_name (Optional[str]): Name of the receiver property.
            edge_kwargs (Optional[Dict]): Additional edge attributes.
            sender_node_kwargs (Optional[Dict]): Additional sender node attributes.
            receiver_node_kwargs (Optional[Dict]): Additional receiver node attributes.

        Raises:
            ValueError: If an unknown graph object is provided.
            TypeError: If the graph is not a pydot.Dot object.
        """
        if sender_node_kwargs is None:
            sender_node_kwargs = {}

        if receiver_node_kwargs is None:
            receiver_node_kwargs = {}

        if graph is self.system_graph:
            rank = self.system_graph_rank
            subgraph_dict = self.system_subgraph_dict
            graph_node_attribute_dict = self.system_graph_node_attribute_dict
            graph_edge_label_dict = self.system_graph_edge_label_dict
        elif graph is self.object_graph:
            rank = self.object_graph_rank
            subgraph_dict = self.object_subgraph_dict
            graph_node_attribute_dict = self.object_graph_node_attribute_dict
            graph_edge_label_dict = self.object_graph_edge_label_dict
        else:
            if isinstance(graph, pydot.Dot):
                raise ValueError("Unknown graph object. Currently implemented graph objects are \"self.system_graph\" and \"self.object_graph\"")
            else:
                raise TypeError(f"The supplied \"graph\" argument must be of type \"{pydot.Dot.__name__}\"")
                
        if sender_component not in self.components.values():
            self._add_object(sender_component)

        if receiver_component not in self.components.values():
            self._add_object(receiver_component)

        sender_class_name = sender_component.__class__
        receiver_class_name = receiver_component.__class__
        if sender_class_name not in subgraph_dict:
            subgraph_dict[sender_class_name] = pydot.Subgraph(rank=rank)
            graph.add_subgraph(subgraph_dict[sender_class_name])
        
        if receiver_class_name not in subgraph_dict:
            subgraph_dict[receiver_class_name] = pydot.Subgraph(rank=rank)
            graph.add_subgraph(subgraph_dict[receiver_class_name])
        
        sender_component_name = self.object_dict_reversed[sender_component]
        receiver_component_name = self.object_dict_reversed[receiver_component]
        self._add_edge(graph, sender_component_name, receiver_component_name, sender_property_name, receiver_property_name, edge_kwargs) ###
        if graph==self.system_graph:
            self.system_graph_edge_counter += 1
        elif graph==self.object_graph:
            self.object_graph_edge_counter += 1
        
        cond1 = not subgraph_dict[sender_class_name].get_node(sender_component_name)
        cond2 = not subgraph_dict[sender_class_name].get_node("\""+ sender_component_name +"\"")
        if cond1 and cond2:
            node = pydot.Node(sender_component_name)
            subgraph_dict[sender_class_name].add_node(node)
        
        cond1 = not subgraph_dict[receiver_class_name].get_node(receiver_component_name)
        cond2 = not subgraph_dict[receiver_class_name].get_node("\""+ receiver_component_name +"\"")
        if cond1 and cond2:
            node = pydot.Node(receiver_component_name)
            subgraph_dict[receiver_class_name].add_node(node)


        if "label" not in sender_node_kwargs:
            sender_node_kwargs.update({"label": sender_component_name})
            graph_node_attribute_dict[sender_component_name] = sender_node_kwargs
        graph_node_attribute_dict[sender_component_name] = sender_node_kwargs

        if "label" not in receiver_node_kwargs:
            receiver_node_kwargs.update({"label": receiver_component_name})
            graph_node_attribute_dict[receiver_component_name] = receiver_node_kwargs
        graph_node_attribute_dict[receiver_component_name] = receiver_node_kwargs


    def remove_connection(self, sender_component: System, receiver_component: System, 
                          sender_property_name: str, receiver_property_name: str) -> None:
        """
        Remove a connection between two components in the system.

        Args:
            sender_component (System): The component sending the connection.
            receiver_component (System): The component receiving the connection.
            sender_property_name (str): Name of the sender property.
            receiver_property_name (str): Name of the receiver property.

        Raises:
            ValueError: If the specified connection does not exist.
        """

        #print("==============================")
        #print("Removing connection between: ", sender_component.id, " and ", receiver_component.id)
        #print("==============================")

        sender_obj_connection = None
        for connection in sender_component.connectedThrough:
            if connection.senderPropertyName == sender_property_name:
                sender_obj_connection = connection
                break
        if sender_obj_connection is None:
            raise ValueError(f"The sender component \"{sender_component.id}\" does not have a connection with the property \"{sender_property_name}\"")
        sender_component.connectedThrough.remove(sender_obj_connection)

        receiver_component_connection_point = None
        for connection_point in receiver_component.connectsAt:
            if connection_point.receiverPropertyName == receiver_property_name:
                receiver_component_connection_point = connection_point
                break
        if receiver_component_connection_point is None:
            raise ValueError(f"The receiver component \"{receiver_component.id}\" does not have a connection point with the property \"{receiver_property_name}\"")
        receiver_component.connectsAt.remove(receiver_component_connection_point)

        del sender_obj_connection
        del receiver_component_connection_point
        
        self._del_edge(self.system_graph, sender_component.id, receiver_component.id, self._get_edge_label(sender_property_name, receiver_property_name))

        #Exception classes 
        exception_classes = (systems.TimeSeriesInputSystem, systems.PiecewiseLinearSystem, systems.PiecewiseLinearSupplyWaterTemperatureSystem, systems.PiecewiseLinearScheduleSystem, base.Sensor, base.Meter) # These classes are exceptions because their inputs and outputs can take any form

        if isinstance(sender_component, exception_classes):
            del sender_component.output[sender_property_name]

        if isinstance(receiver_component, exception_classes):
            del receiver_component.input[receiver_property_name]
    
    def add_outdoor_environment(self, filename: Optional[str] = None) -> None:
        """
        Add an outdoor environment to the model.

        Args:
            filename (Optional[str]): Path to the file containing outdoor environment data.
        """
        outdoor_environment = base.OutdoorEnvironment(
            filename=filename,
            saveSimulationResult = self.saveSimulationResult,
            id = "outdoor_environment")
        self.component_base_dict["outdoor_environment"] = outdoor_environment

    def add_outdoor_environment_system(self, filename: Optional[str] = None) -> None:
        """
        Add an outdoor environment system to the model.

        Args:
            filename (Optional[str]): Path to the file containing outdoor environment data.
        """
        outdoor_environment = systems.OutdoorEnvironmentSystem(
            filename=filename,
            saveSimulationResult = self.saveSimulationResult,
            id = "outdoor_environment")
        self._add_component(outdoor_environment)

    def _instantiate_objects(self, df_dict: Dict[str, pd.DataFrame]) -> None:
        """
        Instantiate all components listed in the configuration file.

        Args:
            df_dict (Dict[str, pd.DataFrame]): A dictionary of dataframes read from the configuration file.
        """
        for ventilation_system_name in df_dict["System"]["Ventilation system name"].dropna():
            ventilation_system = base.DistributionDevice(id=ventilation_system_name)
            self.system_dict["ventilation"][ventilation_system_name] = ventilation_system
        
        for heating_system_name in df_dict["System"]["Heating system name"].dropna():
            heating_system = base.DistributionDevice(id=heating_system_name)
            self.system_dict["heating"][heating_system_name] = heating_system

        for cooling_system_name in df_dict["System"]["Cooling system name"].dropna():
            cooling_system = base.DistributionDevice(id=cooling_system_name)
            self.system_dict["cooling"][cooling_system_name] = cooling_system

        for row in df_dict["BuildingSpace"].dropna(subset=["id"]).itertuples(index=False):
            space_name = row[df_dict["BuildingSpace"].columns.get_loc("id")]
            space = base.BuildingSpace(id=space_name)
            self.component_base_dict[space_name] = space
            
        for row in df_dict["Damper"].dropna(subset=["id"]).itertuples(index=False):
            damper_name = row[df_dict["Damper"].columns.get_loc("id")]
            damper = base.Damper(id=damper_name)
            self.component_base_dict[damper_name] = damper
            if row[df_dict["Damper"].columns.get_loc("isContainedIn")] not in self.component_base_dict:
                warnings.warn("Cannot find a matching BuildingSpace object for damper \"" + damper_name + "\"")                

        for row in df_dict["SpaceHeater"].dropna(subset=["id"]).itertuples(index=False):
            space_heater_name = row[df_dict["SpaceHeater"].columns.get_loc("id")]
            space_heater = base.SpaceHeater(id=space_heater_name)
            self.component_base_dict[space_heater_name] = space_heater
            if row[df_dict["SpaceHeater"].columns.get_loc("isContainedIn")] not in self.component_base_dict:
                warnings.warn("Cannot find a matching SpaceHeater object for space heater \"" + space_heater_name + "\"")                

        for row in df_dict["Valve"].dropna(subset=["id"]).itertuples(index=False):
            valve_name = row[df_dict["Valve"].columns.get_loc("id")]
            valve = base.Valve(id=valve_name)
            self.component_base_dict[valve_name] = valve
            if row[df_dict["Valve"].columns.get_loc("isContainedIn")] not in self.component_base_dict:
                warnings.warn("Cannot find a matching Valve object for valve \"" + valve_name + "\"")

        for row in df_dict["Coil"].dropna(subset=["id"]).itertuples(index=False):
            coil_name = row[df_dict["Coil"].columns.get_loc("id")]
            coil = base.Coil(id=coil_name)
            self.component_base_dict[coil_name] = coil
            
        for row in df_dict["AirToAirHeatRecovery"].dropna(subset=["id"]).itertuples(index=False):
            air_to_air_heat_recovery_name = row[df_dict["AirToAirHeatRecovery"].columns.get_loc("id")]
            air_to_air_heat_recovery = base.AirToAirHeatRecovery(id=air_to_air_heat_recovery_name)
            self.component_base_dict[air_to_air_heat_recovery_name] = air_to_air_heat_recovery

        for row in df_dict["Fan"].dropna(subset=["id"]).itertuples(index=False):
            fan_name = row[df_dict["Fan"].columns.get_loc("id")]
            fan = base.Fan(id=fan_name)
            self.component_base_dict[fan_name] = fan

        for row in df_dict["Controller"].dropna(subset=["id"]).itertuples(index=False):
            controller_name = row[df_dict["Controller"].columns.get_loc("id")]
            controller = base.Controller(id=controller_name)
            self.component_base_dict[controller_name] = controller

        for row in df_dict["SetpointController"].dropna(subset=["id"]).itertuples(index=False):
            controller_name = row[df_dict["SetpointController"].columns.get_loc("id")]
            controller = base.SetpointController(id=controller_name)
            self.component_base_dict[controller_name] = controller

        for row in df_dict["RulebasedController"].dropna(subset=["id"]).itertuples(index=False):
            controller_name = row[df_dict["RulebasedController"].columns.get_loc("id")]
            controller = base.RulebasedController(id=controller_name)
            self.component_base_dict[controller_name] = controller

        for row in df_dict["Schedule"].dropna(subset=["id"]).itertuples(index=False):
            schedule_name = row[df_dict["Schedule"].columns.get_loc("id")]
            schedule = base.Schedule(id=schedule_name)
            self.component_base_dict[schedule_name] = schedule
                
        for row in df_dict["ShadingDevice"].dropna(subset=["id"]).itertuples(index=False):
            shading_device_name = row[df_dict["ShadingDevice"].columns.get_loc("id")]
            shading_device = base.ShadingDevice(id=shading_device_name)
            self.component_base_dict[shading_device_name] = shading_device            

        for row in df_dict["Sensor"].dropna(subset=["id"]).itertuples(index=False):
            sensor_name = row[df_dict["Sensor"].columns.get_loc("id")]
            sensor = base.Sensor(id=sensor_name)
            self.component_base_dict[sensor_name] = sensor

        for row in df_dict["Meter"].dropna(subset=["id"]).itertuples(index=False):
            meter_name = row[df_dict["Meter"].columns.get_loc("id")]
            meter = base.Meter(id=meter_name)
            self.component_base_dict[meter_name] = meter

        for row in df_dict["Pump"].dropna(subset=["id"]).itertuples(index=False):
            pump_name = row[df_dict["Pump"].columns.get_loc("id")]
            pump = base.Pump(id=pump_name)
            self.component_base_dict[pump_name] = pump

        for row in df_dict["FlowJunction"].dropna(subset=["id"]).itertuples(index=False):
            flow_junction_name = row[df_dict["FlowJunction"].columns.get_loc("id")]
            flow_junction = base.FlowJunction(id=flow_junction_name)
            self.component_base_dict[flow_junction_name] = flow_junction
            
        for row in df_dict["Property"].dropna(subset=["id"]).itertuples(index=False):
            property_name = row[df_dict["Property"].columns.get_loc("id")]
            Property = getattr(base, row[df_dict["Property"].columns.get_loc("type")])
            property_ = Property()
            self.property_dict[property_name] = property_

        self._add_heatexchanger_subsystems()

    def _add_heatexchanger_subsystems(self) -> None:
        """
        Add heat exchanger subsystems to the model.
        """
        added_components = []
        for component in self.component_base_dict.values():
            if isinstance(component, self.heatexchanger_types):
                if isinstance(component, base.AirToAirHeatRecovery):
                    component_supply = base.AirToAirHeatRecovery(subSystemOf=[component], id=f"{component.id} (supply)")
                    component_exhaust = base.AirToAirHeatRecovery(subSystemOf=[component], id=f"{component.id} (return)")
                    component.hasSubSystem.append(component_supply)
                    component.hasSubSystem.append(component_exhaust)
                    added_components.append(component_supply)
                    added_components.append(component_exhaust)

                elif isinstance(component, base.Coil):
                    component_waterside = base.Coil(subSystemOf=[component], id=f"{component.id} (waterside)")
                    component_airside = base.Coil(subSystemOf=[component], id=f"{component.id} (airside)")
                    component.hasSubSystem.append(component_waterside)
                    component.hasSubSystem.append(component_airside)
                    added_components.append(component_waterside)
                    added_components.append(component_airside)

        for component in added_components:
            self.component_base_dict[component.id] = component
                
    def _populate_objects(self, df_dict: Dict[str, pd.DataFrame]) -> None:
        """
        Populate all components with data and define connections.

        Args:
            df_dict (Dict[str, pd.DataFrame]): A dictionary of dataframes read from the configuration file.
        """
        allowed_numeric_types = (float, int)
        true_list = ["True", "true", "TRUE"]

        for row in df_dict["BuildingSpace"].dropna(subset=["id"]).itertuples(index=False):
            space_name = row[df_dict["BuildingSpace"].columns.get_loc("id")]
            space = self.component_base_dict[space_name]
            if isinstance(row[df_dict["BuildingSpace"].columns.get_loc("hasProperty")], str):
                properties = [self.property_dict[property_name] for property_name in row[df_dict["BuildingSpace"].columns.get_loc("hasProperty")].split(";")]
                space.hasProperty.extend(properties)
            else:
                warnings.warn(f"The property \"hasProperty\" is not set for BuildingSpace object \"{space.id}\"")
            

            if "connectedTo" not in df_dict["BuildingSpace"].columns:
                warnings.warn("The property \"connectedTo\" is not found in \"BuildingSpace\" sheet. This is ignored for now but will raise an error in the future. It probably is caused by using an outdated configuration file.")
            elif isinstance(row[df_dict["BuildingSpace"].columns.get_loc("connectedTo")], str):
                connected_to = row[df_dict["BuildingSpace"].columns.get_loc("connectedTo")].split(";")
                connected_to = [self.component_base_dict[component_name] for component_name in connected_to]
                space.connectedTo.extend(connected_to)
            
            if "hasFluidSuppliedBy" not in df_dict["BuildingSpace"].columns:
                warnings.warn("The property \"hasFluidSuppliedBy\" is not found in \"BuildingSpace\" sheet. This is ignored for now but will raise an error in the future. It probably is caused by using an outdated configuration file.")
            else:
                if isinstance(row[df_dict["BuildingSpace"].columns.get_loc("hasFluidSuppliedBy")], str):
                    connected_after = row[df_dict["BuildingSpace"].columns.get_loc("hasFluidSuppliedBy")].split(";")
                    connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                    space.hasFluidSuppliedBy.extend(connected_after)
                else:
                    message = f"Property \"hasFluidSuppliedBy\" not set for BuildingSpace object \"{space.id}\""
                    warnings.warn(message)


            if "hasProfile" not in df_dict["BuildingSpace"].columns:
                warnings.warn("The property \"hasProfile\" is not found in \"BuildingSpace\" sheet. This is ignored for now but will raise an error in the future. It probably is caused by using an outdated configuration file.")
            elif isinstance(row[df_dict["BuildingSpace"].columns.get_loc("hasProfile")], str):
                schedule_name = row[df_dict["BuildingSpace"].columns.get_loc("hasProfile")]
                space.hasProfile = self.component_base_dict[schedule_name]

            space.airVolume = row[df_dict["BuildingSpace"].columns.get_loc("airVolume")]
            
        for row in df_dict["Damper"].dropna(subset=["id"]).itertuples(index=False):
            damper_name = row[df_dict["Damper"].columns.get_loc("id")]
            damper = self.component_base_dict[damper_name]
            systems = row[df_dict["Damper"].columns.get_loc("subSystemOf")].split(";")
            systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
            damper.subSystemOf.extend(systems)

            if isinstance(row[df_dict["Damper"].columns.get_loc("hasFluidSuppliedBy")], str):
                connected_after = row[df_dict["Damper"].columns.get_loc("hasFluidSuppliedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                damper.hasFluidSuppliedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidSuppliedBy\" not set for Damper object \"{damper.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["Damper"].columns.get_loc("hasFluidReturnedBy")], str):
                connected_after = row[df_dict["Damper"].columns.get_loc("hasFluidReturnedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                damper.hasFluidReturnedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidSuppliedBy\" not set for Damper object \"{damper.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["Damper"].columns.get_loc("hasProperty")], str):
                properties = [self.property_dict[property_name] for property_name in row[df_dict["Damper"].columns.get_loc("hasProperty")].split(";")]
                damper.hasProperty.extend(properties)
            else:
                message = f"Required property \"hasProperty\" not set for Damper object \"{damper.id}\""
                raise(ValueError(message))
            
            
            damper.isContainedIn = self.component_base_dict[row[df_dict["Damper"].columns.get_loc("isContainedIn")]]
            rsetattr(damper, "nominalAirFlowRate.hasValue", row[df_dict["Damper"].columns.get_loc("nominalAirFlowRate")])
            
        for row in df_dict["SpaceHeater"].dropna(subset=["id"]).itertuples(index=False):
            space_heater_name = row[df_dict["SpaceHeater"].columns.get_loc("id")]
            space_heater = self.component_base_dict[space_heater_name]

            if isinstance(row[df_dict["SpaceHeater"].columns.get_loc("subSystemOf")], str):
                systems = row[df_dict["SpaceHeater"].columns.get_loc("subSystemOf")].split(";")
                systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
                space_heater.subSystemOf.extend(systems)
            else:
                message = f"Required property \"subSystemOf\" not set for SpaceHeater object \"{space_heater.id}\""
                raise(ValueError(message))

            if isinstance(row[df_dict["SpaceHeater"].columns.get_loc("hasFluidSuppliedBy")], str):
                connected_after = row[df_dict["SpaceHeater"].columns.get_loc("hasFluidSuppliedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                space_heater.hasFluidSuppliedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidSuppliedBy\" not set for SpaceHeater object \"{space_heater.id}\""
                warnings.warn(message)


            if isinstance(row[df_dict["SpaceHeater"].columns.get_loc("hasProperty")], str):
                properties = [self.property_dict[property_name] for property_name in row[df_dict["SpaceHeater"].columns.get_loc("hasProperty")].split(";")]
                space_heater.hasProperty.extend(properties)
            else:
                message = f"Property \"hasProperty\" not set for SpaceHeater object \"{space_heater.id}\""
                warnings.warn(message)
            
            space_heater.isContainedIn = self.component_base_dict[row[df_dict["SpaceHeater"].columns.get_loc("isContainedIn")]]
            rsetattr(space_heater, "outputCapacity.hasValue", row[df_dict["SpaceHeater"].columns.get_loc("outputCapacity")])

            if isinstance(row[df_dict["SpaceHeater"].columns.get_loc("temperatureClassification")], str):
                rsetattr(space_heater, "temperatureClassification.hasValue", row[df_dict["SpaceHeater"].columns.get_loc("temperatureClassification")])
            else:
                message = f"Property \"temperatureClassification\" not set for SpaceHeater object \"{space_heater.id}\""
                warnings.warn(message)
                # raise(ValueError(message))
            

            if isinstance(row[df_dict["SpaceHeater"].columns.get_loc("thermalMassHeatCapacity")], str):
                rsetattr(space_heater, "thermalMassHeatCapacity.hasValue", float(row[df_dict["SpaceHeater"].columns.get_loc("thermalMassHeatCapacity")]))
            elif isinstance(row[df_dict["SpaceHeater"].columns.get_loc("thermalMassHeatCapacity")], allowed_numeric_types) and np.isnan(row[df_dict["SpaceHeater"].columns.get_loc("thermalMassHeatCapacity")])==False:
                rsetattr(space_heater, "thermalMassHeatCapacity.hasValue", row[df_dict["SpaceHeater"].columns.get_loc("thermalMassHeatCapacity")])



        for row in df_dict["Valve"].dropna(subset=["id"]).itertuples(index=False):
            valve_name = row[df_dict["Valve"].columns.get_loc("id")]
            valve = self.component_base_dict[valve_name]

            if isinstance(row[df_dict["Valve"].columns.get_loc("subSystemOf")], str):
                systems = row[df_dict["Valve"].columns.get_loc("subSystemOf")].split(";")
                systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
                valve.subSystemOf.extend(systems)
            else:
                message = f"Required property \"subSystemOf\" not set for Valve object \"{valve.id}\""
                raise(ValueError(message))

            if isinstance(row[df_dict["Valve"].columns.get_loc("hasFluidSuppliedBy")], str):
                connected_after = row[df_dict["Valve"].columns.get_loc("hasFluidSuppliedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                valve.hasFluidSuppliedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidSuppliedBy\" not set for Valve object \"{valve.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["Valve"].columns.get_loc("hasFluidReturnedBy")], str):
                connected_after = row[df_dict["Valve"].columns.get_loc("hasFluidReturnedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                valve.hasFluidReturnedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidReturnedBy\" not set for Valve object \"{valve.id}\""
                warnings.warn(message)


            if isinstance(row[df_dict["Valve"].columns.get_loc("hasProperty")], str):
                properties = [self.property_dict[property_name] for property_name in row[df_dict["Valve"].columns.get_loc("hasProperty")].split(";")]
                valve.hasProperty.extend(properties)

            if isinstance(row[df_dict["Valve"].columns.get_loc("isContainedIn")], str):
                valve.isContainedIn = self.component_base_dict[row[df_dict["Valve"].columns.get_loc("isContainedIn")]]
            
            
            if isinstance(row[df_dict["Valve"].columns.get_loc("flowCoefficient")], float) and np.isnan(row[df_dict["Valve"].columns.get_loc("flowCoefficient")])==False:
                rsetattr(valve, "flowCoefficient.hasValue", row[df_dict["Valve"].columns.get_loc("flowCoefficient")])
            
            if isinstance(row[df_dict["Valve"].columns.get_loc("testPressure")], float) and np.isnan(row[df_dict["Valve"].columns.get_loc("testPressure")])==False:
                rsetattr(valve, "testPressure.hasValue", row[df_dict["Valve"].columns.get_loc("testPressure")])

        for row in df_dict["Coil"].dropna(subset=["id"]).itertuples(index=False):
            coil_name = row[df_dict["Coil"].columns.get_loc("id")]
            coil = self.component_base_dict[coil_name]

            if isinstance(row[df_dict["Coil"].columns.get_loc("subSystemOf")], str):
                systems = row[df_dict["Coil"].columns.get_loc("subSystemOf")].split(";")
                systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
                has_ventilation_system = any([system in self.system_dict["ventilation"].values() for system in systems])
                has_heating_system = any([system in self.system_dict["heating"].values() for system in systems])
                has_cooling_system = any([system in self.system_dict["cooling"].values() for system in systems])
                assert has_ventilation_system and (has_heating_system or has_cooling_system), f"Required property \"subSystemOf\" must contain both a Ventilation system and either a Heating or Cooling system for Coil object \"{coil.id}\""
                coil.subSystemOf.extend(systems)
            else:
                message = f"Required property \"subSystemOf\" not set for Coil object \"{coil.id}\""
                raise(ValueError(message))

            if isinstance(row[df_dict["Coil"].columns.get_loc("hasFluidSuppliedBy (airside)")], str):
                connected_after = row[df_dict["Coil"].columns.get_loc("hasFluidSuppliedBy (airside)")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                coil_ = self.component_base_dict[coil_name + " (airside)"]
                coil_.hasFluidSuppliedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidSuppliedBy (airside)\" not set for Coil object \"{coil.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["Coil"].columns.get_loc("hasFluidSuppliedBy (waterside)")], str):
                connected_after = row[df_dict["Coil"].columns.get_loc("hasFluidSuppliedBy (waterside)")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                coil_ = self.component_base_dict[coil_name + " (waterside)"]
                coil_.hasFluidSuppliedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidSuppliedBy (waterside)\" not set for Coil object \"{coil.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["Coil"].columns.get_loc("hasProperty")], str):
                properties = [self.property_dict[property_name] for property_name in row[df_dict["Coil"].columns.get_loc("hasProperty")].split(";")]
                coil.hasProperty.extend(properties)
            
        for row in df_dict["AirToAirHeatRecovery"].dropna(subset=["id"]).itertuples(index=False):
            air_to_air_heat_recovery_name = row[df_dict["AirToAirHeatRecovery"].columns.get_loc("id")]
            air_to_air_heat_recovery = self.component_base_dict[air_to_air_heat_recovery_name]
            systems = row[df_dict["AirToAirHeatRecovery"].columns.get_loc("subSystemOf")].split(";")
            systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
            air_to_air_heat_recovery.subSystemOf = systems

            if isinstance(row[df_dict["AirToAirHeatRecovery"].columns.get_loc("hasFluidSuppliedBy")], str):
                connected_after = row[df_dict["AirToAirHeatRecovery"].columns.get_loc("hasFluidSuppliedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                air_to_air_heat_recovery_ = self.component_base_dict[air_to_air_heat_recovery_name + " (supply)"]
                air_to_air_heat_recovery_.hasFluidSuppliedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidSuppliedBy\" not set for AirToAirHeatRecovery object \"{air_to_air_heat_recovery.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["AirToAirHeatRecovery"].columns.get_loc("hasFluidReturnedBy")], str):
                connected_after = row[df_dict["AirToAirHeatRecovery"].columns.get_loc("hasFluidReturnedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                air_to_air_heat_recovery_ = self.component_base_dict[air_to_air_heat_recovery_name + " (return)"]
                air_to_air_heat_recovery_.hasFluidReturnedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidReturnedBy\" not set for AirToAirHeatRecovery object \"{air_to_air_heat_recovery.id}\""
                warnings.warn(message)
            

            properties = [self.property_dict[property_name] for property_name in row[df_dict["AirToAirHeatRecovery"].columns.get_loc("hasProperty")].split(";")]
            air_to_air_heat_recovery.hasProperty.extend(properties)
            rsetattr(air_to_air_heat_recovery, "primaryAirFlowRateMax.hasValue", row[df_dict["AirToAirHeatRecovery"].columns.get_loc("primaryAirFlowRateMax")])
            rsetattr(air_to_air_heat_recovery, "secondaryAirFlowRateMax.hasValue", row[df_dict["AirToAirHeatRecovery"].columns.get_loc("secondaryAirFlowRateMax")])

        for row in df_dict["Fan"].dropna(subset=["id"]).itertuples(index=False):
            fan_name = row[df_dict["Fan"].columns.get_loc("id")]
            fan = self.component_base_dict[fan_name]

            if isinstance(row[df_dict["Fan"].columns.get_loc("subSystemOf")], str):
                systems = row[df_dict["Fan"].columns.get_loc("subSystemOf")].split(";")
                systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
                fan.subSystemOf.extend(systems)
            else:
                message = f"Required property \"subSystemOf\" not set for fan object \"{fan.id}\""
                raise(ValueError(message))

            if isinstance(row[df_dict["Fan"].columns.get_loc("hasFluidSuppliedBy")], str):
                connected_after = row[df_dict["Fan"].columns.get_loc("hasFluidSuppliedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                fan.hasFluidSuppliedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidSuppliedBy\" not set for Fan object \"{fan.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["Fan"].columns.get_loc("hasFluidReturnedBy")], str):
                connected_after = row[df_dict["Fan"].columns.get_loc("hasFluidReturnedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                fan.hasFluidReturnedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidReturnedBy\" not set for Fan object \"{fan.id}\""
                warnings.warn(message)
            
            if isinstance(row[df_dict["Fan"].columns.get_loc("hasProperty")], str):
                properties = [self.property_dict[property_name] for property_name in row[df_dict["Fan"].columns.get_loc("hasProperty")].split(";")]
                fan.hasProperty.extend(properties)
            rsetattr(fan, "nominalAirFlowRate.hasValue", row[df_dict["Fan"].columns.get_loc("nominalAirFlowRate")])
            rsetattr(fan, "nominalPowerRate.hasValue", row[df_dict["Fan"].columns.get_loc("nominalPowerRate")])

            
        for row in df_dict["Controller"].dropna(subset=["id"]).itertuples(index=False):
            controller_name = row[df_dict["Controller"].columns.get_loc("id")]
            controller = self.component_base_dict[controller_name]

            if isinstance(row[df_dict["Controller"].columns.get_loc("subSystemOf")], str):
                systems = row[df_dict["Controller"].columns.get_loc("subSystemOf")].split(";")
                systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
            else:
                message = f"Required property \"subSystemOf\" not set for controller object \"{controller.id}\""
                raise(ValueError(message))
            
            controller.subSystemOf.extend(systems)

            if isinstance(row[df_dict["Controller"].columns.get_loc("isContainedIn")], str):
                controller.isContainedIn = self.component_base_dict[row[df_dict["Controller"].columns.get_loc("isContainedIn")]]

            _property = self.property_dict[row[df_dict["Controller"].columns.get_loc("observes")]]
            controller.observes = _property


            if "controls" not in df_dict["Controller"].columns:
                warnings.warn("The property \"controls\" is not found in \"Controller\" sheet. This is ignored for now but will raise an error in the future. It probably is caused by using an outdated configuration file.")
            else:
                if isinstance(row[df_dict["Controller"].columns.get_loc("controls")], str):
                    controls = row[df_dict["Controller"].columns.get_loc("controls")].split(";")
                    controls = [self.property_dict[component_name] for component_name in controls]
                    controller.controls.extend(controls)
                else:
                    message = f"Required property \"controls\" not set for controller object \"{controller.id}\""
                    raise(ValueError(message))

        for row in df_dict["SetpointController"].dropna(subset=["id"]).itertuples(index=False):
            controller_name = row[df_dict["SetpointController"].columns.get_loc("id")]
            controller = self.component_base_dict[controller_name]

            if isinstance(row[df_dict["SetpointController"].columns.get_loc("subSystemOf")], str):
                systems = row[df_dict["SetpointController"].columns.get_loc("subSystemOf")].split(";")
                systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
                controller.subSystemOf.extend(systems)
            else:
                message = f"Required property \"subSystemOf\" not set for controller object \"{controller.id}\""
                raise(ValueError(message))
            
            

            if isinstance(row[df_dict["SetpointController"].columns.get_loc("isContainedIn")], str):
                controller.isContainedIn = self.component_base_dict[row[df_dict["SetpointController"].columns.get_loc("isContainedIn")]]
            _property = self.property_dict[row[df_dict["SetpointController"].columns.get_loc("observes")]]
            controller.observes = _property

            if "controls" not in df_dict["SetpointController"].columns:
                warnings.warn("The property \"controls\" is not found in \"SetpointController\" sheet. This is ignored for now but will raise an error in the future. It probably is caused by using an outdated configuration file.")
            else:
                if isinstance(row[df_dict["SetpointController"].columns.get_loc("controls")], str):
                    controls = row[df_dict["SetpointController"].columns.get_loc("controls")].split(";")
                    controls = [self.property_dict[component_name] for component_name in controls]
                    controller.controls.extend(controls)
                else:
                    s = str(row[df_dict["SetpointController"].columns.get_loc("controls")])
                    message = f"Required property \"controls\" not set for controller object \"{controller.id}\", {s} was provided"
                    raise(ValueError(message))

            if "isReverse" not in df_dict["SetpointController"].columns:
                warnings.warn("The property \"isReverse\" is not found in \"SetpointController\" sheet. This is ignored for now but will raise an error in the future. It probably is caused by using an outdated configuration file.")
            else:
                if isinstance(row[df_dict["SetpointController"].columns.get_loc("isReverse")], bool):
                    is_reverse = row[df_dict["SetpointController"].columns.get_loc("isReverse")]
                    controller.isReverse = is_reverse
                elif isinstance(row[df_dict["SetpointController"].columns.get_loc("isReverse")], str):
                    is_reverse = row[df_dict["SetpointController"].columns.get_loc("isReverse")] in true_list
                    controller.isReverse = is_reverse
                elif isinstance(row[df_dict["SetpointController"].columns.get_loc("isReverse")], allowed_numeric_types):
                    is_reverse = int(row[df_dict["SetpointController"].columns.get_loc("isReverse")])==1
                    controller.isReverse = is_reverse
                else:
                    message = f"Required property \"isReverse\" not set to Bool value for controller object \"{controller.id}\""
                    raise(ValueError(message))

            if isinstance(row[df_dict["SetpointController"].columns.get_loc("hasProfile")], str):
                schedule_name = row[df_dict["SetpointController"].columns.get_loc("hasProfile")]
                controller.hasProfile = self.component_base_dict[schedule_name]
            else:
                message = f"Required property \"hasProfile\" not set for controller object \"{controller.id}\""
                raise(ValueError(message))

        for row in df_dict["RulebasedController"].dropna(subset=["id"]).itertuples(index=False):
            controller_name = row[df_dict["RulebasedController"].columns.get_loc("id")]
            controller = self.component_base_dict[controller_name]

            if isinstance(row[df_dict["RulebasedController"].columns.get_loc("subSystemOf")], str):
                systems = row[df_dict["RulebasedController"].columns.get_loc("subSystemOf")].split(";")
                systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
            else:
                message = f"Required property \"subSystemOf\" not set for controller object \"{controller.id}\""
                raise(ValueError(message))
            
            controller.subSystemOf.extend(systems)

            if isinstance(row[df_dict["RulebasedController"].columns.get_loc("isContainedIn")], str):
                controller.isContainedIn = self.component_base_dict[row[df_dict["RulebasedController"].columns.get_loc("isContainedIn")]]
            
            _property = self.property_dict[row[df_dict["RulebasedController"].columns.get_loc("observes")]]
            controller.observes = _property


            if "controls" not in df_dict["RulebasedController"].columns:
                warnings.warn("The property \"controls\" is not found in \"RulebasedController\" sheet. This is ignored for now but will raise an error in the future. It probably is caused by using an outdated configuration file.")
            else:
                if isinstance(row[df_dict["RulebasedController"].columns.get_loc("controls")], str):
                    controls = row[df_dict["RulebasedController"].columns.get_loc("controls")].split(";")
                    controls = [self.property_dict[component_name] for component_name in controls]
                    controller.controls.extend(controls)
                else:
                    message = f"Required property \"controls\" not set for controller object \"{controller.id}\""
                    raise(ValueError(message))

            if "isReverse" not in df_dict["RulebasedController"].columns:
                warnings.warn("The property \"isReverse\" is not found in \"RulebasedController\" sheet. This is ignored for now but will raise an error in the future. It probably is caused by using an outdated configuration file.")
            else:
                if np.isnan(row[df_dict["RulebasedController"].columns.get_loc("isReverse")])==False:
                    if isinstance(row[df_dict["RulebasedController"].columns.get_loc("isReverse")], bool):
                        is_reverse = row[df_dict["RulebasedController"].columns.get_loc("isReverse")]
                        controller.isReverse = is_reverse
                    elif isinstance(row[df_dict["RulebasedController"].columns.get_loc("isReverse")], str):
                        is_reverse = row[df_dict["RulebasedController"].columns.get_loc("isReverse")] in true_list
                        controller.isReverse = is_reverse
                    elif isinstance(row[df_dict["RulebasedController"].columns.get_loc("isReverse")], allowed_numeric_types):
                        is_reverse = int(row[df_dict["RulebasedController"].columns.get_loc("isReverse")])==1
                        controller.isReverse = is_reverse
                    else:
                        message = f"Required property \"isReverse\" not set to Bool value for controller object \"{controller.id}\""
                        raise(ValueError(message))

            if isinstance(row[df_dict["RulebasedController"].columns.get_loc("hasProfile")], str):
                schedule_name = row[df_dict["RulebasedController"].columns.get_loc("hasProfile")]
                controller.hasProfile = self.component_base_dict[schedule_name]
            else:
                message = f"Required property \"hasProfile\" not set for controller object \"{controller.id}\""
                warnings.warn(message)
 
        for row in df_dict["ShadingDevice"].dropna(subset=["id"]).itertuples(index=False):
            shading_device_name = row[df_dict["ShadingDevice"].columns.get_loc("id")]
            shading_device = self.component_base_dict[shading_device_name]

            if isinstance(row[df_dict["ShadingDevice"].columns.get_loc("subSystemOf")], str):
                systems = row[df_dict["ShadingDevice"].columns.get_loc("subSystemOf")].split(";")
                systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
                shading_device.subSystemOf.extend(systems)
            else:
                message = f"Required property \"subSystemOf\" not set for shading_device object \"{shading_device.id}\""
                raise(ValueError(message))

            properties = [self.property_dict[property_name] for property_name in row[df_dict["ShadingDevice"].columns.get_loc("hasProperty")].split(";")]
            shading_device.isContainedIn = self.component_base_dict[row[df_dict["ShadingDevice"].columns.get_loc("isContainedIn")]]
            shading_device.hasProperty.extend(properties)
      
        for row in df_dict["Sensor"].dropna(subset=["id"]).itertuples(index=False):
            sensor_name = row[df_dict["Sensor"].columns.get_loc("id")]
            sensor = self.component_base_dict[sensor_name]

            if isinstance(row[df_dict["Sensor"].columns.get_loc("observes")], str):
                properties = self.property_dict[row[df_dict["Sensor"].columns.get_loc("observes")]]
                sensor.observes = properties
            else:
                message = f"Required property \"observes\" not set for Sensor object \"{sensor.id}\""
                raise(ValueError(message))
            
            if isinstance(row[df_dict["Sensor"].columns.get_loc("isContainedIn")], str):
                sensor.isContainedIn = self.component_base_dict[row[df_dict["Sensor"].columns.get_loc("isContainedIn")]]
            
            if isinstance(row[df_dict["Sensor"].columns.get_loc("hasFluidSuppliedBy")], str):
                connected_after = row[df_dict["Sensor"].columns.get_loc("hasFluidSuppliedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                sensor.hasFluidSuppliedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidSuppliedBy\" not set for Sensor object \"{sensor.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["Sensor"].columns.get_loc("hasFluidReturnedBy")], str):
                connected_after = row[df_dict["Sensor"].columns.get_loc("hasFluidReturnedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                sensor.hasFluidReturnedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidReturnedBy\" not set for Sensor object \"{sensor.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["Sensor"].columns.get_loc("subSystemOf")], str):
                systems = row[df_dict["Sensor"].columns.get_loc("subSystemOf")].split(";")
                systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
                sensor.subSystemOf.extend(systems)
 
        for row in df_dict["Meter"].dropna(subset=["id"]).itertuples(index=False):
            meter_name = row[df_dict["Meter"].columns.get_loc("id")]
            meter = self.component_base_dict[meter_name]
            if isinstance(row[df_dict["Meter"].columns.get_loc("observes")], str):
                properties = self.property_dict[row[df_dict["Meter"].columns.get_loc("observes")]]
                meter.observes = properties
            else:
                message = f"Required property \"observes\" not set for Sensor object \"{meter.id}\""
                raise(ValueError(message))
            
            if isinstance(row[df_dict["Meter"].columns.get_loc("isContainedIn")], str):
                meter.isContainedIn = self.component_base_dict[row[df_dict["Meter"].columns.get_loc("isContainedIn")]]

            if isinstance(row[df_dict["Meter"].columns.get_loc("hasFluidSuppliedBy")], str):
                connected_after = row[df_dict["Meter"].columns.get_loc("hasFluidSuppliedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                meter.hasFluidSuppliedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidSuppliedBy\" not set for Meter object \"{meter.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["Meter"].columns.get_loc("hasFluidReturnedBy")], str):
                connected_after = row[df_dict["Meter"].columns.get_loc("hasFluidReturnedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                meter.hasFluidReturnedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidReturnedBy\" not set for Meter object \"{meter.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["Meter"].columns.get_loc("subSystemOf")], str):
                systems = row[df_dict["Meter"].columns.get_loc("subSystemOf")].split(";")
                systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
                meter.subSystemOf.extend(systems)

        for row in df_dict["Pump"].dropna(subset=["id"]).itertuples(index=False):
            pump_name = row[df_dict["Pump"].columns.get_loc("id")]
            pump = self.component_base_dict[pump_name]

            if isinstance(row[df_dict["Pump"].columns.get_loc("subSystemOf")], str):
                systems = row[df_dict["Pump"].columns.get_loc("subSystemOf")].split(";")
                systems = [system for system_dict in self.system_dict.values() for system in system_dict.values() if system.id in systems]
                pump.subSystemOf.extend(systems)
            else:
                message = f"Required property \"subSystemOf\" not set for Pump object \"{pump.id}\""
                raise(ValueError(message))
            
            if isinstance(row[df_dict["Pump"].columns.get_loc("hasFluidSuppliedBy")], str):
                connected_after = row[df_dict["Pump"].columns.get_loc("hasFluidSuppliedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                pump.hasFluidSuppliedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidSuppliedBy\" not set for Pump object \"{pump.id}\""
                warnings.warn(message)

            if isinstance(row[df_dict["Pump"].columns.get_loc("hasFluidReturnedBy")], str):
                connected_after = row[df_dict["Pump"].columns.get_loc("hasFluidReturnedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                pump.hasFluidReturnedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidReturnedBy\" not set for Pump object \"{pump.id}\""
                warnings.warn(message)
            
            if isinstance(row[df_dict["Pump"].columns.get_loc("hasProperty")], str):
                pump.hasProperty.extend(row[df_dict["Pump"].columns.get_loc("hasProperty")])

        for row in df_dict["FlowJunction"].dropna(subset=["id"]).itertuples(index=False):
            flow_junction_name = row[df_dict["Pump"].columns.get_loc("id")]
            flow_junction = self.component_base_dict[flow_junction_name]

            if isinstance(row[df_dict["FlowJunction"].columns.get_loc("hasFluidSuppliedBy")], str):
                connected_after = row[df_dict["FlowJunction"].columns.get_loc("hasFluidSuppliedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                flow_junction.hasFluidSuppliedBy.extend(connected_after)
            else:
                message = f"Required property \"hasFluidSuppliedBy\" not set for FlowJunction object \"{flow_junction.id}\""
                warnings.warn(message)
            
            if isinstance(row[df_dict["FlowJunction"].columns.get_loc("hasFluidReturnedBy")], str):
                connected_after = row[df_dict["FlowJunction"].columns.get_loc("hasFluidReturnedBy")].split(";")
                connected_after = [self.component_base_dict[component_name] for component_name in connected_after]
                flow_junction.hasFluidReturnedBy.extend(connected_after)
            else:
                message = f"Property \"hasFluidReturnedBy\" not set for FlowJunction object \"{flow_junction.id}\""
                warnings.warn(message)

                        

    def _read_datamodel_config(self, semantic_model_filename: str) -> None:
        """
        Read the configuration file and instantiate/populate objects.

        Args:
            semantic_model_filename (str): Path to the semantic model configuration file.
        """

        wb = load_workbook(semantic_model_filename, read_only=True)
        df_Systems = pd.read_excel(semantic_model_filename, sheet_name="System") if 'System' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_Space = pd.read_excel(semantic_model_filename, sheet_name="BuildingSpace") if 'BuildingSpace' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_Damper = pd.read_excel(semantic_model_filename, sheet_name="Damper") if 'Damper' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_SpaceHeater = pd.read_excel(semantic_model_filename, sheet_name="SpaceHeater") if 'SpaceHeater' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_Valve = pd.read_excel(semantic_model_filename, sheet_name="Valve") if 'Valve' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_Coil = pd.read_excel(semantic_model_filename, sheet_name="Coil") if 'Coil' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_AirToAirHeatRecovery = pd.read_excel(semantic_model_filename, sheet_name="AirToAirHeatRecovery") if 'AirToAirHeatRecovery' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_Fan = pd.read_excel(semantic_model_filename, sheet_name="Fan") if 'Fan' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_Controller = pd.read_excel(semantic_model_filename, sheet_name="Controller") if 'Controller' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_SetpointController = pd.read_excel(semantic_model_filename, sheet_name="SetpointController") if 'SetpointController' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_Schedule = pd.read_excel(semantic_model_filename, sheet_name="Schedule") if 'Schedule' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_RulebasedController = pd.read_excel(semantic_model_filename, sheet_name="RulebasedController") if 'RulebasedController' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_ShadingDevice = pd.read_excel(semantic_model_filename, sheet_name="ShadingDevice") if 'ShadingDevice' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_Sensor = pd.read_excel(semantic_model_filename, sheet_name="Sensor") if 'Sensor' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_Meter = pd.read_excel(semantic_model_filename, sheet_name="Meter") if 'Meter' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_Pump = pd.read_excel(semantic_model_filename, sheet_name="Pump") if 'Pump' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_FlowJunction = pd.read_excel(semantic_model_filename, sheet_name="FlowJunction") if 'FlowJunction' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])
        df_Property = pd.read_excel(semantic_model_filename, sheet_name="Property") if 'Property' in wb.sheetnames else pd.DataFrame([np.nan], columns=["id"])

        df_dict = {"System": df_Systems,
                   "BuildingSpace": df_Space,
                   "Damper": df_Damper,
                   "SpaceHeater": df_SpaceHeater,
                   "Valve": df_Valve,
                   "Coil": df_Coil,
                   "AirToAirHeatRecovery": df_AirToAirHeatRecovery,
                   "Fan": df_Fan,
                   "Controller": df_Controller,
                   "SetpointController": df_SetpointController,
                   "Schedule": df_Schedule,
                   "RulebasedController": df_RulebasedController,
                   "ShadingDevice": df_ShadingDevice,
                   "Sensor": df_Sensor,
                   "Meter": df_Meter,
                   "Pump": df_Pump,
                   "FlowJunction": df_FlowJunction,
                   "Property": df_Property}

        self._instantiate_objects(df_dict)
        self._populate_objects(df_dict)
        
    def _read_input_config(self, input_dict: Dict) -> None:
        """
        Read input configuration and populate corresponding objects.

        Args:
            input_dict (Dict): Dictionary containing input configuration data.
        """
        time_format = '%Y-%m-%d %H:%M:%S%z'
        startTime = datetime.datetime.strptime(input_dict["metadata"]["start_time"], time_format)
        endTime = datetime.datetime.strptime(input_dict["metadata"]["end_time"], time_format)
        stepSize = input_dict["metadata"]['stepSize']

        #print(input_dict.keys())
        
        
        if "rooms_sensor_data" in input_dict:
            """
            Reacting to different input combinations for custom ventilation system model
            """
            
            for component_id, component_data in input_dict["rooms_sensor_data"].items():
                data_available = bool(component_data["sensor_data_available"]) 
                if data_available:
                    timestamp = component_data["time"]
                    co2_values = component_data["co2"]
                    damper_values = component_data["damper_position"]
                    
                    df_co2 = pd.DataFrame({"timestamp": timestamp, "co2": co2_values})
                    df_damper = pd.DataFrame({"timestamp": timestamp, "damper_position": damper_values})
                    
                    df_co2 = sample_from_df(df_co2,
                                            stepSize=stepSize,
                                            start_time=startTime,
                                            end_time=endTime,
                                            resample=True,
                                            clip=True,
                                            tz="Europe/Copenhagen",
                                            preserve_order=True)
                    df_damper = sample_from_df(df_damper,
                                            stepSize=stepSize,
                                            start_time=startTime,
                                            end_time=endTime,
                                            resample=True,
                                            clip=True,
                                            tz="Europe/Copenhagen",
                                            preserve_order=True)
                    df_damper["damper_position"] = df_damper["damper_position"]/100
                    
                    components_ = self.components.keys()
                    #Make it an array of strings
                    components_ = list(components_)
                    # find all the components that contain the last part of the component_id, after the first dash
                    
                    #Extract the substring after the first dash
                    substring_id = component_id.split("-")[1] + "-" + component_id.split("-")[2]
                    substring_id = substring_id.lower().replace("-", "_")
                    
                    # Find all the components that contain the substring
                    filtered_components = [component_ for component_ in components_ if substring_id in component_]

                    #If the component contains "co2" in the id, add the co2 data
                    for component_ in filtered_components:
                        if "CO2_sensor" in component_:
                            co2_sensor = self.components[component_]
                            co2_sensor.df_input = df_co2
                        elif "Damper_position_sensor" in component_:
                            damper_position_sensor = self.components[component_]
                            damper_position_sensor.df_input = df_damper
                else:
                    """
                    Reading schedules and reacting to different controller configurations
                    """
                    rooms_volumes = {
                        "22_601b_00": 500,
                        "22_601b_0": 417,
                        "22_601b_1": 417,
                        "22_601b_2": 417,
                        "22_603_0" : 240,
                        "22_603_1" : 240,
                        "22_604_0" : 486,
                        "22_604_1" : 375,
                        "22_603b_2" : 159,
                        "22_603a_2" : 33,
                        "22_604a_2" : 33,
                        "22_604b_2" : 33,
                        "22_605a_2" : 33,
                        "22_605b_2" : 33,
                        "22_604e_2" : 33,
                        "22_604d_2" : 33,
                        "22_604c_2" : 33,
                        "22_605e_2" : 33,
                        "22_605d_2" : 33,
                        "22_605c_2" : 30,

                    }

                    components_ = self.components.keys()
                    components_ = list(components_)
                    substring_id = component_id.split("-")[1] + "-" + component_id.split("-")[2]
                    substring_id = substring_id.lower().replace("-", "_")
                    substring_id = "22_" + substring_id
                    room_filtered_components = [component_ for component_ in components_ if substring_id in component_]

                    if substring_id == "22_601b_0":
                        components_601b_00 = {
                            "CO2_controller_sensor_22_601b_00",                             
                            "Damper_position_sensor_22_601b_00", 
                            "Supply_damper_22_601b_00", 
                            "Return_damper_22_601b_00", 
                            "CO2_sensor_22_601b_00"
                        }
                        filtered_components = []
                        for component_ in room_filtered_components:
                            if component_ not in components_601b_00:
                                filtered_components.append(component_)
                        room_filtered_components = filtered_components

                    controller_type = component_data["controller_type"]

                    if controller_type == "PID":
                        #Assert that a co2 setpoint schedule is available, if not raise an error with a message
                        assert "co2_setpoint_schedule" in component_data["schedules"], "No CO2 setpoint schedule in input dict. available for PID controller"
                        schedule_input = component_data["schedules"]["co2_setpoint_schedule"]
                        #Assert that the controller constants are available, if not raise an error with a message
                        assert "kp" in component_data, "No kp value in input dict. available for PID controller"
                        assert "ki" in component_data, "No ki value in input dict. available for PID controller"
                        assert "kd" in component_data, "No kd value in input dict. available for PID controller"
                        #get the id of the sensor from the filtered components
                        co2_sensor_component_id = next((component_ for component_ in room_filtered_components if "CO2_sensor" in component_), None)
                        co2_sensor_component = self.components[co2_sensor_component_id]
                        co2_sensor_observed_property = co2_sensor_component.observes

                        #Create the schedule object
                        co2_setpoint_schedule = systems.ScheduleSystem(
                            **schedule_input,
                            add_noise = False,
                            saveSimulationResult = True,
                            id = f"{substring_id}_co2_setpoint_schedule")
                        #Create the PID controller object
                        pid_controller = systems.ControllerSystem(
                            observes = co2_sensor_observed_property,
                            K_p = component_data["kp"],
                            K_i = component_data["ki"],
                            K_d = component_data["kd"],
                            saveSimulationResult = True,
                            id = f"{substring_id}_CO2_PID_controller")
                 
                        #Remove the connections to the previous controller and delete it
                        ann_controller_component_id = next((component_ for component_ in room_filtered_components if "CO2_controller" in component_), None)
                        ann_controller_component = self.components[ann_controller_component_id]
                        co2_sensor_component_id = next((component_ for component_ in room_filtered_components if "CO2_sensor" in component_), None)
                        co2_sensor_component = self.components[co2_sensor_component_id]
                        supply_damper_id = next((component_ for component_ in room_filtered_components if "Supply_damper" in component_), None)
                        return_damper_id = next((component_ for component_ in room_filtered_components if "Return_damper" in component_), None)
                        supply_damper = self.components[supply_damper_id]
                        return_damper = self.components[return_damper_id]

                        self.remove_connection(co2_sensor_component, ann_controller_component, "indoorCo2Concentration", "actualValue")
                        self.remove_connection(ann_controller_component, return_damper, "inputSignal", "damperPosition")
                        self.remove_connection(ann_controller_component, supply_damper, "inputSignal", "damperPosition")
                        self.remove_component(ann_controller_component)

                        #Add the components to the model
                        self._add_component(co2_setpoint_schedule)
                        self._add_component(pid_controller)
                        #Add the connection between the schedule and the controller
                        self.add_connection(co2_setpoint_schedule, pid_controller, "scheduleValue", "setpointValue")
                        self.add_connection(co2_sensor_component, pid_controller, "indoorCo2Concentration", "actualValue")
                        #Add the connection between the controller and the dampers
                        self.add_connection(pid_controller, supply_damper, "inputSignal", "damperPosition")
                        self.add_connection(pid_controller, return_damper, "inputSignal", "damperPosition")
                        pid_controller.observes.isPropertyOf = co2_sensor_component
                
                        #Recalculate the filtered components
                        components_ = self.components.keys()
                        components_ = list(components_)
                        substring_id = component_id.split("-")[1] + "-" + component_id.split("-")[2]
                        substring_id = substring_id.lower().replace("-", "_")
                        substring_id = "22_" + substring_id
                        room_filtered_components = [component_ for component_ in components_ if substring_id in component_]

                        if substring_id == "22_601b_0":
                            components_601b_00 = {
                                "CO2_controller_sensor_22_601b_00",                             
                                "Damper_position_sensor_22_601b_00", 
                                "Supply_damper_22_601b_00", 
                                "Return_damper_22_601b_00", 
                                "CO2_sensor_22_601b_00"
                            }
                            filtered_components = []
                            for component_ in room_filtered_components:
                                if component_ not in components_601b_00:
                                    filtered_components.append(component_)
                            room_filtered_components = filtered_components
                    elif controller_type == "RBC":
                        #Assert that a co2 setpoint schedule is available, if not raise an error with a message
                        assert "co2_setpoint_schedule" in component_data["schedules"], "No CO2 setpoint schedule in input dict. available for PID controller"
                        schedule_input = component_data["schedules"]["co2_setpoint_schedule"]
                        #get the id of the sensor from the filtered components
                        co2_sensor_component_id = next((component_ for component_ in room_filtered_components if "CO2_sensor" in component_), None)
                        co2_sensor_component = self.components[co2_sensor_component_id]
                        co2_sensor_observed_property = co2_sensor_component.observes

                        #Create the schedule object
                        co2_setpoint_schedule = systems.ScheduleSystem(
                            **schedule_input,
                            add_noise = False,
                            saveSimulationResult = True,
                            id = f"{substring_id}_co2_setpoint_schedule")
                        #Create the RBC controller object
                        rbc_controller = systems.RulebasedSetpointInputControllerSystem(
                            observes = co2_sensor_observed_property,
                            saveSimulationResult = True,
                            id = f"{substring_id}_CO2_RBC_controller")
                        
                 
                        #Remove the connections to the previous controller and delete it
                        ann_controller_component_id = next((component_ for component_ in room_filtered_components if "CO2_controller" in component_), None)
                        ann_controller_component = self.components[ann_controller_component_id]
                        co2_sensor_component_id = next((component_ for component_ in room_filtered_components if "CO2_sensor" in component_), None)
                        co2_sensor_component = self.components[co2_sensor_component_id]
                        supply_damper_id = next((component_ for component_ in room_filtered_components if "Supply_damper" in component_), None)
                        return_damper_id = next((component_ for component_ in room_filtered_components if "Return_damper" in component_), None)
                        supply_damper = self.components[supply_damper_id]
                        return_damper = self.components[return_damper_id]

                        self.remove_connection(co2_sensor_component, ann_controller_component, "indoorCo2Concentration", "actualValue")
                        self.remove_connection(ann_controller_component, return_damper, "inputSignal", "damperPosition")
                        self.remove_connection(ann_controller_component, supply_damper, "inputSignal", "damperPosition")
                        self.remove_component(ann_controller_component)

                        #Add the components to the model
                        self._add_component(co2_setpoint_schedule)
                        self._add_component(rbc_controller)
                        #Add the connection between the schedule and the controller
                        self.add_connection(co2_setpoint_schedule, rbc_controller, "scheduleValue", "setpointValue")
                        self.add_connection(co2_sensor_component, rbc_controller, "indoorCo2Concentration", "actualValue")
                        #Add the connection between the controller and the dampers
                        self.add_connection(rbc_controller, supply_damper, "inputSignal", "damperPosition")
                        self.add_connection(rbc_controller, return_damper, "inputSignal", "damperPosition")
                        rbc_controller.observes.isPropertyOf = co2_sensor_component
                
                        #Recalculate the filtered components
                        components_ = self.components.keys()
                        components_ = list(components_)
                        substring_id = component_id.split("-")[1] + "-" + component_id.split("-")[2]
                        substring_id = substring_id.lower().replace("-", "_")
                        substring_id = "22_" + substring_id
                        room_filtered_components = [component_ for component_ in components_ if substring_id in component_]

                        if substring_id == "22_601b_0":
                            components_601b_00 = {
                                "CO2_controller_sensor_22_601b_00",                             
                                "Damper_position_sensor_22_601b_00", 
                                "Supply_damper_22_601b_00", 
                                "Return_damper_22_601b_00", 
                                "CO2_sensor_22_601b_00"
                            }
                            filtered_components = []
                            for component_ in room_filtered_components:
                                if component_ not in components_601b_00:
                                    filtered_components.append(component_)
                            room_filtered_components = filtered_components
                    
                    
                    for component_ in room_filtered_components:
                        if "CO2_sensor" in component_:
                            sender_component = self.components[component_]
                            receiver_component_id = next((component_ for component_ in room_filtered_components if "_controller" in component_), None)
                            receiver_component = self.components[receiver_component_id]
                            self.remove_connection(sender_component, receiver_component, "indoorCo2Concentration", "actualValue")
                            self.remove_component(sender_component)
                            
                            schedule_input = component_data["schedules"]["occupancy_schedule"] 
                            #Create the schedule object
                            room_occupancy_schedule = systems.ScheduleSystem(
                                **schedule_input,
                                add_noise = False,
                                saveSimulationResult = True,
                                id = f"{substring_id}_occupancy_schedule")
                            #Create the space co2 object
                            room_volume = int(rooms_volumes[substring_id]) 
                            room_space_co2 = systems.BuildingSpaceCo2System(
                                airVolume = room_volume,
                                saveSimulationResult = True,
                                id = f"{substring_id}_CO2_space")
                            
                            self._add_component(room_occupancy_schedule)
                            self._add_component(room_space_co2)

                            supply_damper_id = next((component_ for component_ in room_filtered_components if "Supply_damper" in component_), None)
                            return_damper_id = next((component_ for component_ in room_filtered_components if "Return_damper" in component_), None)
                            supply_damper = self.components[supply_damper_id]
                            return_damper = self.components[return_damper_id]

                            self.add_connection(room_occupancy_schedule, room_space_co2,
                                "scheduleValue", "numberOfPeople")
                            self.add_connection(supply_damper, room_space_co2,
                                "airFlowRate", "supplyAirFlowRate")
                            self.add_connection(return_damper, room_space_co2,
                                "airFlowRate", "returnAirFlowRate")
                            self.add_connection(room_space_co2, receiver_component,
                                "indoorCo2Concentration", "actualValue")
                            
                            receiver_component.observes.isPropertyOf = room_space_co2           


            ## ADDED FOR DAMPER CONTROL of 601b_00, missing data

            oe_601b_00_component = self.components["CO2_controller_sensor_22_601b_00"]
            freq = f"{stepSize}S"
            df_damper_control = pd.DataFrame({"timestamp": pd.date_range(start=startTime, end=endTime, freq=freq)})
            df_damper_control["damper_position"] = 1
            
            df_damper_control = sample_from_df(df_damper_control,
                        stepSize=stepSize,
                        start_time=startTime,
                        end_time=endTime,
                        resample=True,
                        clip=True,
                        tz="Europe/Copenhagen",
                        preserve_order=True)
            oe_601b_00_component.df_input = df_damper_control

        else:
            sensor_inputs = input_dict["inputs_sensor"] #Change naming to be consistent
            schedule_inputs = input_dict["input_schedules"] #Change naming to be consistent
            weather_inputs = sensor_inputs["ml_inputs_dmi"]
            
            df_raw = pd.DataFrame()
            df_raw.insert(0, "datetime", weather_inputs["observed"])
            df_raw.insert(1, "outdoorTemperature", weather_inputs["temp_dry"])
            df_raw.insert(2, "globalIrradiation", weather_inputs["radia_glob"])
            df_sample = sample_from_df(df_raw,
                                        stepSize=stepSize,
                                        start_time=startTime,
                                        end_time=endTime,
                                        resample=True,
                                        clip=True,
                                        tz="Europe/Copenhagen",
                                        preserve_order=True)

            outdoor_environment = systems.OutdoorEnvironmentSystem(df_input=df_sample,
                                                            saveSimulationResult = self.saveSimulationResult,
                                                            id = "outdoor_environment")

            

            '''
                MODIFIED BY NEC-INDIA

                temperature_list = sensor_inputs["ml_inputs"]["temperature"]
            '''
            if sensor_inputs["ml_inputs"]["temperature"][0]=="None":
                initial_temperature = 21
            else:
                initial_temperature = float(sensor_inputs["ml_inputs"]["temperature"][0])
            custom_initial_dict = {"OE20-601b-2": {"indoorTemperature": initial_temperature}}
            self.set_custom_initial_dict(custom_initial_dict)

            indoor_temperature_setpoint_schedule = systems.ScheduleSystem(
                **schedule_inputs["temperature_setpoint_schedule"],
                add_noise = False,
                saveSimulationResult = True,
                id = "OE20-601b-2_temperature_setpoint_schedule")

            occupancy_schedule = systems.ScheduleSystem(
                **schedule_inputs["occupancy_schedule"],
                add_noise = True,
                saveSimulationResult = True,
                id = "OE20-601b-2_occupancy_schedule")

            supply_water_temperature_setpoint_schedule = systems.PiecewiseLinearScheduleSystem(
                **schedule_inputs["supply_water_temperature_schedule_pwlf"],
                saveSimulationResult = True,
                id = "Heating system_supply_water_temperature_schedule")
            
            supply_air_temperature_schedule = systems.ScheduleSystem(
                **schedule_inputs["supply_air_temperature_schedule"],
                saveSimulationResult = True,
                id = "Ventilation system_supply_air_temperature_schedule")

            # self._add_component(outdoor_environment)
            # self._add_component(occupancy_schedule)
            # self._add_component(indoor_temperature_setpoint_schedule)
            # self._add_component(supply_water_temperature_setpoint_schedule)
            # self._add_component(supply_air_temperature_schedule)


            space = self.components["OE20-601b-2"]
            space_heater = self.components["Space heater"]
            temperature_controller = self.components["Temperature controller"]
            # outdoor_environment = self.components["outdoor_environment"]
            # supply_water_temperature_setpoint_schedule = self.components["Heating system_supply_water_temperature_schedule"]
            # supply_air_temperature_schedule = self.components["Ventilation system_supply_air_temperature_schedule"]
            # indoor_temperature_setpoint_schedule = self.components["OE20-601b-2_temperature_setpoint_schedule"]
            # occupancy_schedule = self.components["OE20-601b-2_occupancy_schedule"]
    

            self.add_connection(outdoor_environment, space, "outdoorTemperature", "outdoorTemperature")
            self.add_connection(outdoor_environment, space, "globalIrradiation", "globalIrradiation")
            self.add_connection(outdoor_environment, supply_water_temperature_setpoint_schedule, "outdoorTemperature", "outdoorTemperature")
            self.add_connection(supply_water_temperature_setpoint_schedule, space_heater, "scheduleValue", "supplyWaterTemperature")
            self.add_connection(supply_water_temperature_setpoint_schedule, space, "scheduleValue", "supplyWaterTemperature")
            self.add_connection(supply_air_temperature_schedule, space, "scheduleValue", "supplyAirTemperature")
            self.add_connection(indoor_temperature_setpoint_schedule, temperature_controller, "scheduleValue", "setpointValue")
            self.add_connection(occupancy_schedule, space, "scheduleValue", "numberOfPeople")

    def update_attribute(self, component: Any, attribute: str, value: Any) -> None:
        """
        Update the value of an attribute of a component.

        Args:
            component (Any): The component to update.
            attribute (str): The name of the attribute to update.
            value (Any): The new value for the attribute.
        """
        if component is None:
            return
            
        attr = rgetattr(component, attribute)
        if isinstance(attr, list):
            if isinstance(value, list):
                for v in value:
                    if v not in attr:
                        attr.append(v)
            if value not in attr:
                attr.append(value)
        else:
            rsetattr(component, attribute, value)

    def _parse_semantic_model(self) -> None:
        """
        Parse the semantic model and update component relationships.
        """
        space_instances = self.get_component_by_class(self.component_base_dict, base.BuildingSpace)
        damper_instances = self.get_component_by_class(self.component_base_dict, base.Damper)
        space_heater_instances = self.get_component_by_class(self.component_base_dict, base.SpaceHeater)
        valve_instances = self.get_component_by_class(self.component_base_dict, base.Valve)
        coil_instances = self.get_component_by_class(self.component_base_dict, base.Coil)
        air_to_air_heat_recovery_instances = self.get_component_by_class(self.component_base_dict, base.AirToAirHeatRecovery)
        fan_instances = self.get_component_by_class(self.component_base_dict, base.Fan)
        controller_instances = self.get_component_by_class(self.component_base_dict, base.Controller)
        setpoint_controller_instances = self.get_component_by_class(self.component_base_dict, base.SetpointController)
        rulebased_controller_instances = self.get_component_by_class(self.component_base_dict, base.RulebasedController)
        shading_device_instances = self.get_component_by_class(self.component_base_dict, base.ShadingDevice)
        sensor_instances = self.get_component_by_class(self.component_base_dict, base.Sensor)
        meter_instances = self.get_component_by_class(self.component_base_dict, base.Meter)
        pump_instances = self.get_component_by_class(self.component_base_dict, base.Pump)
        flow_junction_instances = self.get_component_by_class(self.component_base_dict, base.FlowJunction)

        for space in space_instances:
            for property_ in space.hasProperty:
                self.update_attribute(property_, "isPropertyOf", space)
            for component in space.hasFluidSuppliedBy:
                self.update_attribute(component, "suppliesFluidTo", space)
            space.hasFluidFedBy = space.hasFluidSuppliedBy
            for component in space.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", space)
            for component in space.connectedTo:
                self.update_attribute(component, "connectedTo", space)
            
        for damper in damper_instances:
            self.update_attribute(damper, "isContainedIn.contains", damper)
            for system in damper.subSystemOf:
                self.update_attribute(system, "hasSubSystem", damper)
            for property_ in damper.hasProperty:
                self.update_attribute(property_, "isPropertyOf", damper)
            for component in damper.hasFluidSuppliedBy:
                self.update_attribute(component, "suppliesFluidTo", damper)
            for component in damper.hasFluidReturnedBy:
                self.update_attribute(component, "returnsFluidTo", damper)

            damper.hasFluidFedBy = damper.hasFluidSuppliedBy
            for component in damper.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", damper)

        for space_heater in space_heater_instances:
            for component in space_heater.hasFluidSuppliedBy:
                self.update_attribute(component, "suppliesFluidTo", space_heater)
            space_heater.hasFluidFedBy = space_heater.hasFluidSuppliedBy
            for component in space_heater.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", space_heater)
            self.update_attribute(space_heater, "isContainedIn.contains", space_heater)
            for system in space_heater.subSystemOf:
                self.update_attribute(system, "hasSubSystem", space_heater)
            for property_ in space_heater.hasProperty:
                self.update_attribute(property_, "isPropertyOf", space_heater)

        for valve in valve_instances:
            self.update_attribute(valve.isContainedIn, "contains", valve)
            for system in valve.subSystemOf:
                self.update_attribute(system, "hasSubSystem", valve)
            for property_ in valve.hasProperty:
                self.update_attribute(property_, "isPropertyOf", valve)
            for component in valve.hasFluidSuppliedBy:
                self.update_attribute(component, "suppliesFluidTo", valve)
            for component in valve.hasFluidReturnedBy:
                self.update_attribute(component, "returnsFluidTo", valve)
            valve.hasFluidFedBy = valve.hasFluidSuppliedBy
            for component in valve.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", valve)

        for coil in coil_instances:
            for system in coil.subSystemOf:
                self.update_attribute(system, "hasSubSystem", coil)
            for property_ in coil.hasProperty:
                self.update_attribute(property_, "isPropertyOf", coil)

            for component in coil.hasFluidSuppliedBy:
                self.update_attribute(component, "suppliesFluidTo", coil)
            for component in coil.hasFluidReturnedBy:
                self.update_attribute(component, "returnsFluidTo", coil)

            coil.hasFluidFedBy = coil.hasFluidSuppliedBy
            for component in coil.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", coil)

        for air_to_air_heat_recovery in air_to_air_heat_recovery_instances:
            for system in air_to_air_heat_recovery.subSystemOf:
                self.update_attribute(system, "hasSubSystem", air_to_air_heat_recovery)
            for property_ in air_to_air_heat_recovery.hasProperty:
                self.update_attribute(property_, "isPropertyOf", air_to_air_heat_recovery)

            for component in air_to_air_heat_recovery.hasFluidSuppliedBy:
                self.update_attribute(component, "suppliesFluidTo", air_to_air_heat_recovery)
            for component in air_to_air_heat_recovery.hasFluidReturnedBy:
                self.update_attribute(component, "returnsFluidTo", air_to_air_heat_recovery)

            air_to_air_heat_recovery.hasFluidFedBy = air_to_air_heat_recovery.hasFluidSuppliedBy
            for component in air_to_air_heat_recovery.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", air_to_air_heat_recovery)

        for fan in fan_instances:
            for system in fan.subSystemOf:
                self.update_attribute(system, "hasSubSystem", fan)
            for property_ in fan.hasProperty:
                self.update_attribute(property_, "isPropertyOf", fan)
            for component in fan.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", fan)

            for component in fan.hasFluidSuppliedBy:
                self.update_attribute(component, "suppliesFluidTo", fan)
            for component in fan.hasFluidReturnedBy:
                self.update_attribute(component, "returnsFluidTo", fan)

            fan.hasFluidFedBy = fan.hasFluidSuppliedBy
            for component in fan.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", fan)

        for controller in controller_instances:
            if controller.isContainedIn is not None:
                self.update_attribute(controller.isContainedIn, "contains", controller)
            self.update_attribute(controller.observes, "isObservedBy", controller)
            for property_ in controller.controls:
                self.update_attribute(property_, "isControlledBy", controller)
            for system in controller.subSystemOf:
                self.update_attribute(system, "hasSubSystem", controller)

        for setpoint_controller in setpoint_controller_instances:
            self.update_attribute(setpoint_controller.isContainedIn, "contains", setpoint_controller)
            self.update_attribute(setpoint_controller, "observes.isObservedBy", setpoint_controller)
            for property_ in setpoint_controller.controls:
                self.update_attribute(property_, "isControlledBy", setpoint_controller)
            for system in setpoint_controller.subSystemOf:
                self.update_attribute(system, "hasSubSystem", setpoint_controller)


        for rulebased_controller in rulebased_controller_instances:
            self.update_attribute(rulebased_controller.isContainedIn, "contains", rulebased_controller)
            self.update_attribute(rulebased_controller.observes, "isObservedBy", rulebased_controller)
            for property_ in rulebased_controller.controls:
                self.update_attribute(property_, "isControlledBy", rulebased_controller)
            for system in rulebased_controller.subSystemOf:
                self.update_attribute(system, "hasSubSystem", rulebased_controller)

        for shading_device in shading_device_instances:
            self.update_attribute(shading_device.isContainedIn, "contains", shading_device)
            for system in shading_device.subSystemOf:
                self.update_attribute(system, "hasSubSystem", shading_device)
            for property_ in shading_device.hasProperty:
                self.update_attribute(property_, "isPropertyOf", shading_device)

        for sensor in sensor_instances:
            self.update_attribute(sensor.isContainedIn, "contains", sensor)
            self.update_attribute(sensor.observes, "isObservedBy", sensor)
            for system in sensor.subSystemOf:
                self.update_attribute(system, "hasSubSystem", sensor)
            for component in sensor.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", sensor)

            for component in sensor.hasFluidSuppliedBy:
                self.update_attribute(component, "suppliesFluidTo", sensor)
            for component in sensor.hasFluidReturnedBy:
                self.update_attribute(component, "returnsFluidTo", sensor)

            sensor.hasFluidFedBy = sensor.hasFluidSuppliedBy
            for component in sensor.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", sensor)


        for meter in meter_instances:
            self.update_attribute(meter.isContainedIn, "contains", meter)
            self.update_attribute(meter.observes, "isObservedBy", meter)
            for system in meter.subSystemOf:
                self.update_attribute(system, "hasSubSystem", meter)
            for component in meter.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", meter)

            for component in meter.hasFluidSuppliedBy:
                self.update_attribute(component, "suppliesFluidTo", meter)
            for component in meter.hasFluidReturnedBy:
                self.update_attribute(component, "returnsFluidTo", meter)

            
            meter.hasFluidFedBy = meter.hasFluidSuppliedBy
            for component in meter.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", meter)

        for pump in pump_instances:
            for system in pump.subSystemOf:
                self.update_attribute(system, "hasSubSystem", pump)
            # for component in pump.hasFluidFedBy:
            #     self.update_attribute(component, "feedsFluidTo", pump)
            for property_ in pump.hasProperty:
                self.update_attribute(property_, "isPropertyOf", pump)

            for component in pump.hasFluidSuppliedBy:
                self.update_attribute(component, "suppliesFluidTo", pump)
            for component in pump.hasFluidReturnedBy:
                self.update_attribute(component, "returnsFluidTo", pump)

            pump.hasFluidFedBy = pump.hasFluidSuppliedBy
            for component in pump.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", pump)

        for flow_junction in flow_junction_instances:
            for system in flow_junction.subSystemOf:
                self.update_attribute(system, "hasSubSystem", flow_junction)
            for component in flow_junction.hasFluidSuppliedBy:
                self.update_attribute(component, "suppliesFluidTo", flow_junction)
            for component in flow_junction.hasFluidReturnedBy:
                self.update_attribute(component, "returnsFluidTo", flow_junction)

            flow_junction.hasFluidFedBy = flow_junction.hasFluidSuppliedBy
            for component in flow_junction.hasFluidFedBy:
                self.update_attribute(component, "feedsFluidTo", flow_junction)


        for heating_system in self.system_dict["heating"].values():
            self._add_object(heating_system)

        for cooling_system in self.system_dict["cooling"].values():
            self._add_object(cooling_system)

        for ventilation_system in self.system_dict["ventilation"].values():
            self._add_object(ventilation_system)

    def get_object_properties(self, object_: Any) -> Dict:
        """
        Get all properties of an object.

        Args:
            object_ (Any): The object to get properties from.

        Returns:
            Dict: A dictionary of object properties.
        """
        return {key: value for (key, value) in vars(object_).items()}
        
    def get_component_by_class(self, dict_: Dict, class_: Type, filter: Optional[Callable] = None) -> List:
        """
        Get components of a specific class from a dictionary.

        Args:
            dict_ (Dict): The dictionary to search.
            class_ (Type): The class to filter by.
            filter (Optional[Callable]): Additional filter function.

        Returns:
            List: List of components matching the class and filter.
        """
        if filter is None:
            filter = lambda v, class_: True
        return [v for v in dict_.values() if (isinstance(v, class_) and filter(v, class_))]

    def _connect(self):
        def copy_nodemap(nodemap):
            return {k: v for k, v in nodemap.items()}

        def copy_nodemap_list(nodemap_list):
            return [copy_nodemap(nodemap) for nodemap in nodemap_list]
        
        def _prune_recursive(match_node, sp_node, node_map, node_map_list, feasible, comparison_table, ruleset):
            if sp_node not in feasible: feasible[sp_node] = set()
            if sp_node not in comparison_table: comparison_table[sp_node] = set()
            feasible[sp_node].add(match_node)
            comparison_table[sp_node].add(match_node)
            match_name_attributes = get_object_attributes(match_node)
            sp_node_pairs = sp_node.attributes
            sp_node_pairs_ = sp_node._attributes
            sp_node_pairs_list = sp_node._list_attributes
            
            for sp_attr_name, sp_node_child in sp_node_pairs_.items(): #iterate the required attributes/predicates of the signature node
                if sp_attr_name in match_name_attributes: #is there a match with the semantic node?
                    match_node_child = rgetattr(match_node, sp_attr_name)
                    if match_node_child is not None:
                        rule = ruleset[(sp_node, sp_node_child, sp_attr_name)]
                        pairs, rule_applies, ruleset = rule.apply(match_node, match_node_child, ruleset, node_map_list=node_map_list)
                        if len(pairs)==1:
                            node_map_list__, filtered_match_node_child, filtered_sp_node_child = next(iter(pairs))
                            if filtered_match_node_child not in comparison_table[sp_node_child]:
                                node_map_list, node_map, feasible, comparison_table, prune = _prune_recursive(filtered_match_node_child, filtered_sp_node_child, node_map, node_map_list__, feasible, comparison_table, ruleset)
                                if prune and isinstance(rule, signature_pattern.Optional)==False:
                                    feasible[sp_node].discard(match_node)
                                    return node_map_list, node_map, feasible, comparison_table, True
                            elif filtered_match_node_child not in feasible[sp_node_child]:
                                feasible[sp_node].discard(match_node)
                                return node_map_list, node_map, feasible, comparison_table, True
                        else:
                            feasible[sp_node].discard(match_node)
                            return node_map_list, node_map, feasible, comparison_table, True
                    else:
                        if isinstance(sp_node_child, list):
                            for sp_node_child_ in sp_node_child:
                                rule = ruleset[(sp_node, sp_node_child_, sp_attr_name)]
                                if isinstance(rule, signature_pattern.Optional)==False:
                                    feasible[sp_node].discard(match_node)
                                    return node_map_list, node_map, feasible, comparison_table, True
                        else:
                            rule = ruleset[(sp_node, sp_node_child, sp_attr_name)]
                            if isinstance(rule, signature_pattern.Optional)==False:
                                feasible[sp_node].discard(match_node)
                                return node_map_list, node_map, feasible, comparison_table, True
                else:
                    if isinstance(sp_node_child, list):
                        for sp_node_child_ in sp_node_child:
                            rule = ruleset[(sp_node, sp_node_child_, sp_attr_name)]
                            if isinstance(rule, signature_pattern.Optional)==False:
                                feasible[sp_node].discard(match_node)
                                return node_map_list, node_map, feasible, comparison_table, True
                    else:
                        rule = ruleset[(sp_node, sp_node_child, sp_attr_name)]
                        if isinstance(rule, signature_pattern.Optional)==False:
                            feasible[sp_node].discard(match_node)
                            return node_map_list, node_map, feasible, comparison_table, True
            for sp_attr_name, sp_node_child in sp_node_pairs_list.items(): #iterate the required attributes/predicates of the signature node
                if sp_attr_name in match_name_attributes: #is there a match with the semantic node?
                    match_node_child = rgetattr(match_node, sp_attr_name)
                    if match_node_child is not None:
                        for sp_node_child_ in sp_node_child:
                            rule = ruleset[(sp_node, sp_node_child_, sp_attr_name)]
                            pairs, rule_applies, ruleset = rule.apply(match_node, match_node_child, ruleset, node_map_list=node_map_list)
                            found = False
                            new_node_map_list = []
                            for node_map_list__, filtered_match_node_child, filtered_sp_node_child in pairs:
                                if filtered_match_node_child not in comparison_table[sp_node_child_]:
                                    comparison_table[sp_node_child_].add(filtered_match_node_child)
                                    node_map_list_, node_map_, feasible, comparison_table, prune = _prune_recursive(filtered_match_node_child, filtered_sp_node_child, node_map, node_map_list__, feasible, comparison_table, ruleset)
                                        
                                    if found and prune==False:
                                        name = match_node.id if "id" in get_object_attributes(match_node) else match_node.__class__.__name__
                                        warnings.warn(f"Multiple matches found for context signature node \"{sp_node.id}\" and semantic model node \"{name}\".")
                                    
                                    if prune==False:
                                        new_node_map_list.extend(node_map_list_)
                                        found = True

                                elif filtered_match_node_child in feasible[sp_node_child_]:
                                    for node_map__ in node_map_list__:
                                        node_map__[sp_node_child_] = filtered_match_node_child
                                    new_node_map_list.extend(node_map_list__)
                                    found = True

                            if found==False and isinstance(rule, signature_pattern.Optional)==False:
                                feasible[sp_node].discard(match_node)
                                return node_map_list, node_map, feasible, comparison_table, True
                            else:
                                node_map_list = new_node_map_list

                    else:
                        if isinstance(sp_node_child, list):
                            for sp_node_child_ in sp_node_child:
                                rule = ruleset[(sp_node, sp_node_child_, sp_attr_name)]
                                if isinstance(rule, signature_pattern.Optional)==False:
                                    feasible[sp_node].discard(match_node)
                                    return node_map_list, node_map, feasible, comparison_table, True
                        else:
                            rule = ruleset[(sp_node, sp_node_child, sp_attr_name)]
                            if isinstance(rule, signature_pattern.Optional)==False:
                                feasible[sp_node].discard(match_node)
                                return node_map_list, node_map, feasible, comparison_table, True
                else:
                    if isinstance(sp_node_child, list):
                        for sp_node_child_ in sp_node_child:
                            rule = ruleset[(sp_node, sp_node_child_, sp_attr_name)]
                            if isinstance(rule, signature_pattern.Optional)==False:
                                feasible[sp_node].discard(match_node)
                                return node_map_list, node_map, feasible, comparison_table, True
                    else:
                        rule = ruleset[(sp_node, sp_node_child, sp_attr_name)]
                        if isinstance(rule, signature_pattern.Optional)==False:
                            feasible[sp_node].discard(match_node)
                            return node_map_list, node_map, feasible, comparison_table, True
            if len(node_map_list)==0:
                node_map_list = [node_map]

            node_map_list = copy_nodemap_list(node_map_list)
            for node_map__ in node_map_list:
                # CHANGED: Direct assignment instead of set operation
                node_map__[sp_node] = match_node
            
            return node_map_list, node_map, feasible, comparison_table, False


        def match(group, node_map_, sp, cg, new_ig):
            # CHANGED: Modified comparison to check direct values instead of set lengths
            can_match = all([group[sp_node_] == node_map_[sp_node_]
                            if group[sp_node_] is not None and node_map_[sp_node_] is not None
                            else True for sp_node_ in sp.nodes])
            is_match = False
            if can_match:
                # CHANGED: Filter None values instead of empty sets
                node_map_no_None = {sp_node_: match_node
                                    for sp_node_, match_node in node_map_.items()
                                    if match_node is not None}

                for sp_node_, match_node_nm in node_map_no_None.items():
                    attributes = sp_node_.attributes
                    for attr, subject in attributes.items():
                        # CHANGED: Single node access instead of iterating over set
                        node_map_child = getattr(match_node_nm, attr)
                        if node_map_child is not None and (isinstance(node_map_child, list) and len(node_map_child) == 0) == False:
                            if isinstance(node_map_child, list) == False:
                                node_map_child_ = [node_map_child]
                            else:
                                node_map_child_ = node_map_child
                            if isinstance(subject, list) == False:
                                subject_ = [subject]
                            else:
                                subject_ = subject

                            for subject__ in subject_:
                                # CHANGED: Direct node access instead of set
                                group_child = group[subject__]
                                if group_child is not None and len(node_map_child_) != 0:
                                    # CHANGED: Simple list membership check instead of set operations
                                    if group_child in node_map_child_:
                                        is_match = True
                                        break
                        if is_match:
                            break
                    if is_match:
                        break

                if is_match:
                    for sp_node_, match_node_ in node_map_no_None.items():
                        feasible = {sp_node: set() for sp_node in sp.nodes}
                        comparison_table = {sp_node: set() for sp_node in sp.nodes}
                        sp.reset_ruleset()
                        group_prune = copy_nodemap(group)
                        group_prune = {sp_node___: group_prune[sp_node___] for sp_node___ in sp.nodes}
                        _, _, _, _, prune = _prune_recursive(match_node_, sp_node_, copy_nodemap(group_prune), [group_prune], feasible, comparison_table, sp.ruleset)
                        if prune:
                            is_match = False
                            break

                    if is_match:
                        for sp_node__, match_node__ in node_map_no_None.items():
                            group[sp_node__] = match_node__  # CHANGED: Direct assignment instead of set operations
                        if all([group[sp_node_] is not None for sp_node_ in sp.nodes]):  # CHANGED: Check for None instead of empty sets
                            cg.append(group)
                            new_ig.remove(group)

            if not is_match:
                group_no_None = {sp_node_: match_node for sp_node_, match_node in group.items() if match_node is not None}
                for sp_node_, match_node_group in group_no_None.items():
                    attributes = sp_node_.attributes
                    for attr, subject in attributes.items():
                        group_child = getattr(match_node_group, attr)
                        if group_child is not None and (isinstance(group_child, list) and len(group_child) == 0) == False:
                            if isinstance(group_child, list) == False:
                                group_child_ = [group_child]
                            else:
                                group_child_ = group_child
                            if isinstance(subject, list) == False:
                                subject_ = [subject]
                            else:
                                subject_ = subject

                            for subject__ in subject_:
                                node_map_child_ = node_map_[subject__]
                                if node_map_child_ is not None and group_child_ is not None:
                                    if group_child_ == node_map_child_:
                                        is_match = True
                                        break
                        if is_match:
                            break
                    if is_match:
                        break

                if is_match:
                    for sp_node_, match_node_ in node_map_no_None.items():
                        feasible = {sp_node: set() for sp_node in sp.nodes}
                        comparison_table = {sp_node: set() for sp_node in sp.nodes}
                        sp.reset_ruleset()
                        group_prune = copy_nodemap(group)
                        group_prune = {sp_node___: group_prune[sp_node___] for sp_node___ in sp.nodes}
                        _, _, _, _, prune = _prune_recursive(match_node_, sp_node_, copy_nodemap(group_prune), [group_prune], feasible, comparison_table, sp.ruleset)
                        if prune:
                            is_match = False
                            break

                    if is_match:
                        for sp_node__, match_node__ in node_map_no_None.items():
                            group[sp_node__] = match_node__  # CHANGED: Direct assignment instead of set operations
                        if all([group[sp_node_] is not None for sp_node_ in sp.nodes]):  # CHANGED: Check for None instead of empty sets
                            cg.append(group)
                            new_ig.remove(group)

            return is_match, group, cg, new_ig

        classes = [cls[1] for cls in inspect.getmembers(systems, inspect.isclass) if (issubclass(cls[1], (System, )) and hasattr(cls[1], "sp"))]
        complete_groups = {}
        incomplete_groups = {}
        counter = 0
        for component_cls in classes:
            complete_groups[component_cls] = {}
            incomplete_groups[component_cls] = {}
            sps = component_cls.sp
            for sp in sps:
                complete_groups[component_cls][sp] = []
                incomplete_groups[component_cls][sp] = []
                cg = complete_groups[component_cls][sp]
                ig = incomplete_groups[component_cls][sp]
                for sp_node in sp.nodes:
                    l = list(sp_node.cls)
                    l.remove(signature_pattern.NodeBase)
                    l = tuple(l)
                    match_nodes = [c for c in self.object_dict.values() if (isinstance(c, l)) and isinstance(c, signature_pattern.NodeBase)==False] ######################################
                        
                    for match_node in match_nodes:
                        # CHANGED: Initialize with None instead of empty set
                        node_map = {sp_node_: None for sp_node_ in sp.nodes}
                        feasible = {sp_node: set() for sp_node in sp.nodes}
                        comparison_table = {sp_node: set() for sp_node in sp.nodes}
                        node_map_list = [copy_nodemap(node_map)]
                        prune = True
                        if match_node not in comparison_table[sp_node]:
                            sp.reset_ruleset()
                            node_map_list, node_map, feasible, comparison_table, prune = _prune_recursive(match_node, sp_node, node_map, node_map_list, feasible, comparison_table, sp.ruleset)

                        elif match_node in feasible[sp_node]:
                            node_map[sp_node] = match_node
                            node_map_list = [node_map]
                            prune = False
                        
                        if prune==False:
                            # We check that the obtained node_map_list contains node maps with different modeled nodes.
                            # If an SP does not contain a MultipleMatches rule, we can prune the node_map_list to only contain node maps with different modeled nodes.
                            modeled_nodes = []
                            for node_map_ in node_map_list:
                                node_map_set = set()
                                for sp_modeled_node in sp.modeled_nodes:
                                    node_map_set.add(node_map_[sp_modeled_node])
                                modeled_nodes.append(node_map_set)
                            node_map_list_new = []
                            for i,(node_map_, node_map_set) in enumerate(zip(node_map_list, modeled_nodes)):
                                active_set = node_map_set
                                passive_set = set().union(*[v for k,v in enumerate(modeled_nodes) if k!=i])
                                if len(active_set.intersection(passive_set))>0 and any([isinstance(v, signature_pattern.MultipleMatches) for v in sp._ruleset.values()])==False:
                                    warnings.warn(f"Multiple matches found for {sp_node.id} and {sp_node.cls}.")
                                node_map_list_new.append(node_map_) # This constraint has been removed to allow for multiple matches. Note that multiple
                            node_map_list = node_map_list_new
                            
                            # Cross matching could maybe stop early if a match is found. For SP with multiple allowed matches it might be necessary to check all matches 
                            for node_map_ in node_map_list:
                                if all([node_map_[sp_node_] is not None for sp_node_ in sp.nodes]):
                                    cg.append(node_map_)
                                else:
                                    if len(ig)==0: #If there are no groups in the incomplete group list, add the node map
                                        ig.append(node_map_)
                                    else:
                                        new_ig = ig.copy()
                                        is_match_ = False
                                        for group in ig: #Iterate over incomplete groups
                                            is_match, group, cg, new_ig = match(group, node_map_, sp, cg, new_ig)
                                            if is_match:
                                                is_match_ = True
                                        if is_match_==False:
                                            new_ig.append(node_map_)
                                        ig = new_ig
                
                ig_len = np.inf
                while len(ig)<ig_len:
                    ig_len = len(ig)
                    new_ig = ig.copy()
                    is_match = False
                    for group_i in ig:
                        for group_j in ig:
                            if group_i!=group_j:
                                is_match, group, cg, new_ig = match(group_i, group_j, sp, cg, new_ig)
                            if is_match:
                                break
                        if is_match:
                            break
                    ig = new_ig
                    
                
                # if True:#component_cls is components.BuildingSpace1AdjBoundaryOutdoorFMUSystem:
                #     print("INCOMPLETE GROUPS================================================================================")
                #     for group in ig:
                #         print("GROUP------------------------------")
                #         for sp_node_, match_node_set in group.items():
                #             id_sp = sp_node_.id if "id" in get_object_attributes(sp_node_) else sp_node_.__class__.__name__ + " [" + str(id(sp_node_)) +"]"
                #             id_sp = id_sp.replace(r"\n", "")
                #             id_m = [match_node.id if "id" in get_object_attributes(match_node) else match_node.__class__.__name__ + " [" + str(id(match_node)) +"]" for match_node in match_node_set]
                #             print(id_sp, id_m, "comparison: ", [m in comparison_table[sp_node_] if sp_node_ in comparison_table else None for m in match_node_set], "feasible: ", [m in feasible[sp_node_] if sp_node_ in comparison_table else None for m in match_node_set])


                #     print("COMPLETE GROUPS================================================================================")
                #     for group in cg:
                #         print("GROUP------------------------------")
                #         for sp_node_, match_node_set in group.items():
                #             id_sp = sp_node_.id if "id" in get_object_attributes(sp_node_) else sp_node_.__class__.__name__ + " [" + str(id(sp_node_)) +"]"
                #             id_sp = id_sp.replace(r"\n", "")
                #             id_m = [match_node.id if "id" in get_object_attributes(match_node) else match_node.__class__.__name__ + " [" + str(id(match_node)) +"]" for match_node in match_node_set]
                #             print(id_sp, id_m)      
                
                
                new_ig = ig.copy()
                for group in ig: #Iterate over incomplete groups
                    if all([group[sp_node_] is not None for sp_node_ in sp.required_nodes]):  # CHANGED: Check for None instead of empty sets
                        cg.append(group)
                        new_ig.remove(group)
                ig = new_ig


        for component_cls, sps in complete_groups.items():
            complete_groups[component_cls] = {sp: groups for sp, groups in sorted(complete_groups[component_cls].items(), key=lambda item: item[0].priority, reverse=True)}
        complete_groups = {k: v for k, v in sorted(complete_groups.items(), key=lambda item: max(sp.priority for sp in item[1]), reverse=True)}
        self.instance_map = {}
        self.instance_map_reversed = {}
        self.instance_to_group_map = {} ############### if changed to self.instance_to_group_map, it cannot be pickled
        modeled_components = set()
        for i, (component_cls, sps) in enumerate(complete_groups.items()):
            for sp, groups in sps.items():
                for group in groups:
                    modeled_match_nodes = {group[sp_node] for sp_node in sp.modeled_nodes} # CHANGED: Access single node directly
                    if len(modeled_components.intersection(modeled_match_nodes))==0 or any([isinstance(v, signature_pattern.MultipleMatches) for v in sp._ruleset.values()]):
                        modeled_components |= modeled_match_nodes #Union/add set
                        if len(modeled_match_nodes)==1:
                            component = next(iter(modeled_match_nodes))
                            id_ = component.id
                            base_kwargs = self.get_object_properties(component)
                            extension_kwargs = {"id": id_,
                                                "saveSimulationResult": self.saveSimulationResult,}
                        else:
                            id_ = ""
                            modeled_match_nodes_sorted = sorted(modeled_match_nodes, key=lambda x: x.id)
                            for component in modeled_match_nodes_sorted:
                                id_ += f"[{component.id}]"
                            base_kwargs = {}
                            extension_kwargs = {"id": id_,
                                                "saveSimulationResult": self.saveSimulationResult,
                                                "base_components": list(modeled_match_nodes_sorted)}
                            for component in modeled_match_nodes_sorted:
                                kwargs = self.get_object_properties(component)
                                base_kwargs.update(kwargs)

                        if id_ not in [c.id for c in self.instance_map.keys()]: #Check if the instance is already created. For components with Multiple matches, the model might already have been created.
                            base_kwargs.update(extension_kwargs)
                            component = component_cls(**base_kwargs)
                            self.instance_to_group_map[component] = (modeled_match_nodes, (component_cls, sp, [group]))
                            self.instance_map[component] = modeled_match_nodes
                            for modeled_match_node in modeled_match_nodes:
                                self.instance_map_reversed[modeled_match_node] = component
                        else:
                            component = self.instance_map_reversed[next(iter(modeled_match_nodes))] # Just index with the first element in the set as all elements should return the same component
                            (modeled_match_nodes_, (_, _, groups)) = self.instance_to_group_map[component]
                            modeled_match_nodes_ |= modeled_match_nodes
                            groups.append(group)
                            self.instance_to_group_map[component] = (modeled_match_nodes_, (component_cls, sp, groups))
                            self.instance_map[component] = modeled_match_nodes_
                            for modeled_match_node in modeled_match_nodes_:
                                self.instance_map_reversed[modeled_match_node] = component


        for component, (modeled_match_nodes, (component_cls, sp, groups)) in self.instance_to_group_map.items():
            # Get all required inputs for the component
            for key, (sp_node, source_keys) in sp.inputs.items():
                match_node_list = [group[sp_node] for group in groups]  # CHANGED: Access single node directly
                match_node_set = {group[sp_node] for group in groups}
                if match_node_set.issubset(modeled_components):
                    for match_node in match_node_list:
                        component_inner = self.instance_map_reversed[match_node]
                        source_key = [source_key for c, source_key in source_keys.items() if isinstance(component_inner, c)][0]
                        self.add_connection(component_inner, component, source_key, key)
                else:
                    for match_node in match_node_list:
                        warnings.warn(f"\nThe component with class \"{match_node.__class__.__name__}\" and id \"{match_node.id}\" is not modeled. The input \"{key}\" of the component with class \"{component_cls.__name__}\" and id \"{component.id}\" is not connected.\n")
            
            # Get all parameters for the component
            for key, node in sp.parameters.items():
                if groups[0][node] is not None:
                    value = groups[0][node]
                    rsetattr(component, key, value)
        


    def set_custom_initial_dict(self, custom_initial_dict: Dict[str, Dict[str, Any]]) -> None:
        """
        Set custom initial values for components.

        Args:
            custom_initial_dict (Dict[str, Dict[str, Any]]): Dictionary of custom initial values.

        Raises:
            AssertionError: If unknown component IDs are provided.
        """
        np_custom_initial_dict_ids = np.array(list(custom_initial_dict.keys()))
        legal_ids = np.array([dict_id in self.components for dict_id in custom_initial_dict])
        assert np.all(legal_ids), f"Unknown component id(s) provided in \"custom_initial_dict\": {np_custom_initial_dict_ids[legal_ids==False]}"
        self.custom_initial_dict = custom_initial_dict

    def set_initial_values(self) -> None:
        """
        Set initial values for all components in the model.
        """
        default_initial_dict = {
            systems.OutdoorEnvironmentSystem.__name__: {},
            systems.OccupancySystem.__name__: {"scheduleValue": tps.Scalar(0)},
            systems.ScheduleSystem.__name__: {},
            systems.BuildingSpaceCo2System.__name__: {"indoorCo2Concentration": tps.Scalar(500)},
            systems.BuildingSpaceOccSystem.__name__: {"numberOfPeople": tps.Scalar(0)},
            systems.BuildingSpace0AdjBoundaryFMUSystem.__name__: {"indoorTemperature": tps.Scalar(21),
                                                        "indoorCo2Concentration": tps.Scalar(500)},
            systems.BuildingSpace0AdjBoundaryOutdoorFMUSystem.__name__: {"indoorTemperature": tps.Scalar(21),
                                                        "indoorCo2Concentration": tps.Scalar(500)},
            systems.BuildingSpace1AdjFMUSystem.__name__: {"indoorTemperature": tps.Scalar(21),
                                                        "indoorCo2Concentration": tps.Scalar(500)},
            systems.BuildingSpace2AdjFMUSystem.__name__: {"indoorTemperature": tps.Scalar(21),
                                                        "indoorCo2Concentration": tps.Scalar(500)},
            systems.BuildingSpace1AdjBoundaryFMUSystem.__name__: {"indoorTemperature": tps.Scalar(21),
                                                        "indoorCo2Concentration": tps.Scalar(500)},
            systems.BuildingSpace2SH1AdjBoundaryOutdoorFMUSystem.__name__: {"indoorTemperature": tps.Scalar(21),
                                                        "indoorCo2Concentration": tps.Scalar(500)},
            systems.BuildingSpace2AdjBoundaryFMUSystem.__name__: {"indoorTemperature": tps.Scalar(21),
                                                        "indoorCo2Concentration": tps.Scalar(500)},
            systems.BuildingSpace2AdjBoundaryOutdoorFMUSystem.__name__: {"indoorTemperature": tps.Scalar(21),
                                                        "indoorCo2Concentration": tps.Scalar(500)},
            systems.BuildingSpaceNoSH1AdjBoundaryOutdoorFMUSystem.__name__: {"indoorTemperature": tps.Scalar(21),
                                                        "indoorCo2Concentration": tps.Scalar(500)},
            systems.BuildingSpace1AdjBoundaryOutdoorFMUSystem.__name__: {"indoorTemperature": tps.Scalar(21),
                                                        "indoorCo2Concentration": tps.Scalar(500)},      
            systems.BuildingSpace11AdjBoundaryOutdoorFMUSystem.__name__: {"indoorTemperature": tps.Scalar(21),
                                                        "indoorCo2Concentration": tps.Scalar(500)},                                                                                   
            systems.PIControllerFMUSystem.__name__: {"inputSignal": tps.Scalar(0)},
            systems.PIDControllerSystem.__name__: {"inputSignal": tps.Scalar(0)},
            systems.RulebasedControllerSystem.__name__: {"inputSignal": tps.Scalar(0)},
            systems.RulebasedSetpointInputControllerSystem.__name__: {"inputSignal": tps.Scalar(0)},
            systems.ClassificationAnnControllerSystem.__name__: {"inputSignal": tps.Scalar(0)},
            systems.PIControllerFMUSystem.__name__: {"inputSignal": tps.Scalar(0)},
            systems.SequenceControllerSystem.__name__: {"inputSignal": tps.Scalar(0)},  
            systems.OnOffControllerSystem.__name__: {"inputSignal": tps.Scalar(0)},  
            systems.AirToAirHeatRecoverySystem.__name__: {"primaryTemperatureOut": tps.Scalar(21)},
            systems.CoilPumpValveFMUSystem.__name__: {},
            systems.CoilHeatingSystem.__name__: {"outletAirTemperature": tps.Scalar(21)},
            systems.CoilCoolingSystem.__name__: {},
            systems.CoilHeatingCoolingSystem.__name__: {"outletAirTemperature": tps.Scalar(21)},
            systems.DamperSystem.__name__: {"airFlowRate": tps.Scalar(0),
                                                "damperPosition": tps.Scalar(0)},
            systems.ValveSystem.__name__: {"waterFlowRate": tps.Scalar(0),
                                                "valvePosition": tps.Scalar(0)},
            systems.ValveFMUSystem.__name__: {"waterFlowRate": tps.Scalar(0),
                                                "valvePosition": tps.Scalar(0)},
            systems.FanSystem.__name__: {}, #Energy
            systems.FanFMUSystem.__name__: {"outletAirTemperature": tps.Scalar(21)}, #Energy
            systems.SpaceHeaterSystem.__name__: {"outletWaterTemperature": tps.Scalar(21),
                                                    "Energy": tps.Scalar(0)},
            systems.SupplyFlowJunctionSystem.__name__: {"airFlowRateIn": tps.Scalar(0)},
            systems.ReturnFlowJunctionSystem.__name__: {"airFlowRateOut": tps.Scalar(0),
                                                           "airTemperatureOut": tps.Scalar(21)},
            systems.SensorSystem.__name__: {"measuredValue": tps.Scalar(0)},
            systems.ShadingDeviceSystem.__name__: {},
            systems.NeuralPolicyControllerSystem.__name__: {},
            systems.MeterSystem.__name__: {},
            systems.PiecewiseLinearSystem.__name__: {},
            systems.PiecewiseLinearSupplyWaterTemperatureSystem.__name__: {},
            systems.PiecewiseLinearScheduleSystem.__name__: {},
            systems.TimeSeriesInputSystem.__name__: {},
            systems.OnOffSystem.__name__: {},
            
        }
        initial_dict = {}
        for component in self.components.values():
            initial_dict[component.id] = {k: v.copy() for k, v in default_initial_dict[type(component).__name__].items()}
        if self.custom_initial_dict is not None:
            for key, value in self.custom_initial_dict.items():
                initial_dict[key].update(value)

        for component in self.components.values():
            component.output.update(initial_dict[component.id])

    def set_parameters_from_array(self, parameters: List[Any], component_list: List[System], attr_list: List[str]) -> None:
        """
        Set parameters for components from an array.

        Args:
            parameters (List[Any]): List of parameter values.
            component_list (List[System]): List of components to set parameters for.
            attr_list (List[str]): List of attribute names corresponding to the parameters.

        Raises:
            AssertionError: If a component doesn't have the specified attribute.
        """
        for i, (p, obj, attr) in enumerate(zip(parameters, component_list, attr_list)):
            assert rhasattr(obj, attr), f"The component with class \"{obj.__class__.__name__}\" and id \"{obj.id}\" has no attribute \"{attr}\"."
            rsetattr(obj, attr, p)

    def set_parameters_from_dict(self, parameters: Dict[str, Any], component_list: List[System], attr_list: List[str]) -> None:
        """
        Set parameters for components from a dictionary.

        Args:
            parameters (Dict[str, Any]): Dictionary of parameter values.
            component_list (List[System]): List of components to set parameters for.
            attr_list (List[str]): List of attribute names corresponding to the parameters.

        Raises:
            AssertionError: If a component doesn't have the specified attribute.
        """
        for (obj, attr) in zip(component_list, attr_list):
            assert rhasattr(obj, attr), f"The component with class \"{obj.__class__.__name__}\" and id \"{obj.id}\" has no attribute \"{attr}\"."
            rsetattr(obj, attr, parameters[attr])

    def cache(self, startTime: Optional[datetime.datetime] = None, endTime: Optional[datetime.datetime] = None, stepSize: Optional[int] = None) -> None:
        """
        Cache data and create folder structure for time series data.

        Args:
            startTime (Optional[datetime.datetime]): Start time for caching.
            endTime (Optional[datetime.datetime]): End time for caching.
            stepSize (Optional[int]): Time step size for caching.
        """
        c = self.get_component_by_class(self.components, (systems.SensorSystem, systems.MeterSystem, systems.OutdoorEnvironmentSystem, systems.TimeSeriesInputSystem))
        for component in c:
            component.initialize(startTime=startTime,
                                endTime=endTime,
                                stepSize=stepSize)
            
    def initialize(self,
                   startTime: Optional[datetime.datetime] = None,
                   endTime: Optional[datetime.datetime] = None,
                   stepSize: Optional[int] = None) -> None:
        """
        Initialize the model for simulation.

        Args:
            startTime (Optional[datetime.datetime]): Start time for the simulation.
            endTime (Optional[datetime.datetime]): End time for the simulation.
            stepSize (Optional[int]): Time step size for the simulation.
        """
        self.set_initial_values()
        self.check_for_for_missing_initial_values()
        for component in self.flat_execution_order:
            component.clear_results()
            component.initialize(startTime=startTime,
                                endTime=endTime,
                                stepSize=stepSize,
                                model=self)

            for v in component.input.values():
                v.reset()
                
            for v in component.output.values():
                v.reset()

            # Make the inputs and outputs aware of the execution order.
            # This is important to ensure that input tps.Vectors have the same order, allowing for instance element-wise operations.
            for i, connection_point in enumerate(component.connectsAt):
                for j, connection in enumerate(connection_point.connectsSystemThrough):
                    connected_component = connection.connectsSystem
                    if isinstance(component.input[connection_point.receiverPropertyName], tps.Vector):
                        if component in self.instance_to_group_map:
                            (modeled_match_nodes, (component_cls, sp, groups)) = self.instance_to_group_map[component]

                            # Find the group of the connected component
                            modeled_match_nodes_ = self.instance_map[connected_component]
                            groups_matched = [g for g in groups if len(modeled_match_nodes_.intersection(set(g.values())))>0]
                            assert len(groups_matched)==1, "Only one group is allowed for each component."
                            group = groups_matched[0]
                            group_id = id(group)
                            component.input[connection_point.receiverPropertyName].update(group_id=group_id)
                        else:
                            component.input[connection_point.receiverPropertyName].update()


            for v in component.input.values():
                v.initialize()
                
            for v in component.output.values():
                v.initialize()



    def validate(self) -> None:
        """
        Validate the model by checking IDs and connections.
        """
        self.p.add_level()
        (validated_for_simulator1, validated_for_estimator1, validated_for_evaluator1, validated_for_monitor1) = self.validate_parameters()
        (validated_for_simulator2, validated_for_estimator2, validated_for_evaluator2, validated_for_monitor2) = self.validate_ids()
        (validated_for_simulator3, validated_for_estimator3, validated_for_evaluator3, validated_for_monitor3) = self.validate_connections()

        self.validated_for_simulator = validated_for_simulator1 and validated_for_simulator2 and validated_for_simulator3
        self.validated_for_estimator = validated_for_estimator1 and validated_for_estimator2 and validated_for_estimator3
        self.validated_for_evaluator = validated_for_evaluator1 and validated_for_evaluator2 and validated_for_evaluator3
        self.validated_for_monitor = validated_for_monitor1 and validated_for_monitor2 and validated_for_monitor3
        self.validated = self.validated_for_simulator and self.validated_for_estimator and self.validated_for_evaluator and self.validated_for_monitor
        self.p.remove_level()


        self.p("Validated for Simulator")
        if self.validated_for_simulator:
            status = "OK"
        else:
            status = "FAILED"
        
        self.p("Validated for Estimator", status=status)
        if self.validated_for_estimator:
            status = "OK"
        else:
            status = "FAILED"

        self.p("Validated for Evaluator", status=status)
        if self.validated_for_evaluator:
            status = "OK"
        else:
            status = "FAILED"

        self.p("Validated for Monitor", status=status)
        if self.validated_for_monitor:
            status = "OK"
        else:
            status = "FAILED"

        self.p("", plain=True, status=status)


        # assert validated, "The model is not valid. See the warnings above."

    def validate_parameters(self) -> None:
        """
        Validate the parameters of all components in the model.

        Raises:
            AssertionError: If any component has invalid parameters.
        """
        component_instances = list(self.components.values())
        validated_for_simulator = True
        validated_for_estimator = True
        validated_for_evaluator = True
        validated_for_monitor = True
        for component in component_instances:
            if hasattr(component, "validate"): #Check if component has validate method
                (validated_for_simulator_, validated_for_estimator_, validated_for_evaluator_, validated_for_monitor_) = component.validate(self.p)
                validated_for_simulator = validated_for_simulator and validated_for_simulator_
                validated_for_estimator = validated_for_estimator and validated_for_estimator_
                validated_for_evaluator = validated_for_evaluator and validated_for_evaluator_
                validated_for_monitor = validated_for_monitor and validated_for_monitor_
            else:
                config = component.config.copy()
                parameters = {attr: rgetattr(component, attr) for attr in config["parameters"]}
                is_none = [k for k,v in parameters.items() if v is None]
                if any(is_none):
                    message = f"|CLASS: {component.__class__.__name__}|ID: {component.id}|: Missing values for the following parameter(s) to enable use of Simulator, Evaluator, and Monitor:"
                    self.p(message, plain=True, status="[WARNING]")
                    self.p.add_level()
                    for par in is_none:
                        self.p(par, plain=True, status="")
                    self.p.remove_level()
                    # 
                    validated_for_simulator = False
                    validated_for_evaluator = False
                    validated_for_monitor = False
        return (validated_for_simulator, validated_for_estimator, validated_for_evaluator, validated_for_monitor)
                
    def validate_ids(self) -> None:
        """
        Validate the IDs of all components in the model.

        Raises:
            AssertionError: If any component has an invalid ID.
        """
        validated = True
        component_instances = list(self.components.values())
        for component in component_instances:
            isvalid = np.array([x.isalnum() or x in self.valid_chars for x in component.id])
            np_id = np.array(list(component.id))
            violated_characters = list(np_id[isvalid==False])
            if not all(isvalid):
                message = f"|CLASS: {component.__class__.__name__}|ID: {component.id}|: Invalid id. The characters \"{', '.join(violated_characters)}\" are not allowed."
                self.p(message)
                validated = False
        return (validated, validated, validated, validated)


    def validate_connections(self) -> None:
        """
        Validate the connections between components in the model.

        Raises:
            AssertionError: If any required connections are missing.
        """
        component_instances = list(self.components.values())
        validated = True
        for component in component_instances:
            if len(component.connectedThrough)==0 and len(component.connectsAt)==0:
                warnings.warn(f"|CLASS: {component.__class__.__name__}|ID: {component.id}|: No connections. The component has been removed from the model.")
                self.remove_component(component)

            if hasattr(component, "optional_inputs"):
                optional_inputs = component.optional_inputs
            else:
                optional_inputs = []
            input_labels = [cp.receiverPropertyName for cp in component.connectsAt]
            first_input = True
            for req_input_label in component.input.keys():
                if req_input_label not in input_labels and req_input_label not in optional_inputs:
                    if first_input:
                        message = f"|CLASS: {component.__class__.__name__}|ID: {component.id}|: Missing connections for the following input(s) to enable use of Simulator, Estimator, Evaluator, and Monitor:"
                        self.p(message, plain=True, status="[WARNING]")
                        first_input = False
                        self.p.add_level()
                    self.p(req_input_label, plain=True)
                    validated = False
            if first_input==False:
                self.p.remove_level()
        return (validated, validated, validated, validated)

    def _load_parameters(self, force_config_update: bool = False) -> None:
        """
        Load parameters for all components from configuration files.

        Args:
            force_config_update (bool): If True, all parameters are read from the config file. If False, only the parameters that are None are read from the config file. If you want to use the fcn function
            to set the parameters, you should set force_config_update to False to avoid it being overwritten.
        """
        for component in self.components.values():
            assert hasattr(component, "config"), f"The class \"{component.__class__.__name__}\" has no \"config\" attribute."
            config = component.config.copy()
            assert "parameters" in config, f"The \"config\" attribute of class \"{component.__class__.__name__}\" has no \"parameters\" key."
            filename, isfile = self.get_dir(folder_list=["model_parameters", component.__class__.__name__], filename=f"{component.id}.json")
            config["parameters"] = {attr: rgetattr(component, attr) for attr in config["parameters"]}
            if isfile==False:
                with open(filename, 'w') as f:
                    json.dump(config, f, indent=4)
            else:
                with open(filename) as f:
                    config = json.load(f)

                if force_config_update:
                    attr_list = [k for k in config["parameters"].keys()]
                    component_list = [component for k in config["parameters"].keys()]
                else:
                    attr_list = [k for k in config["parameters"].keys() if rgetattr(component, k) is None]
                    component_list = [component for k in config["parameters"].keys() if rgetattr(component, k) is None]
                
                parameters = {k: float(v) if isnumeric(v) else v for k, v in config["parameters"].items()}
                self.set_parameters_from_dict(parameters, component_list, attr_list)

                if "readings" in config:
                    filename_ = config["readings"]["filename"]
                    datecolumn = config["readings"]["datecolumn"]
                    valuecolumn = config["readings"]["valuecolumn"]
                    if filename_ is not None:
                        component.filename = filename_

                        if datecolumn is not None:
                            component.datecolumn = datecolumn
                        elif isinstance(component, systems.OutdoorEnvironmentSystem)==False:
                            raise(ValueError(f"\"datecolumn\" is not defined in the \"readings\" key of the config file: {filename}"))

                        if valuecolumn is not None:
                            component.valuecolumn = valuecolumn
                        elif isinstance(component, systems.OutdoorEnvironmentSystem)==False:
                            raise(ValueError(f"\"valuecolumn\" is not defined in the \"readings\" key of the config file: {filename}"))
                    
    def load(self, semantic_model_filename: Optional[str] = None, input_config: Optional[Dict] = None, 
                   fcn: Optional[Callable] = None, create_object_graph: bool = True, 
                   create_signature_graphs: bool = False, create_system_graph: bool = True, verbose: bool = False, 
                   validate_model: bool = True, force_config_update: bool = False) -> None:
        """
        Load and set up the model for simulation.

        Args:
            semantic_model_filename (Optional[str]): Path to the semantic model configuration file.
            input_config (Optional[Dict]): Input configuration dictionary.
            fcn (Optional[Callable]): Custom function to be applied during model loading.
            create_object_graph (bool): Whether to create and save the object graph.
            create_signature_graphs (bool): Whether to create and save signature graphs.
            create_system_graph (bool): Whether to create and save the system graph.
            verbose (bool): Whether to print verbose output during loading.
            validate_model (bool): Whether to perform model validation.
        """
        if verbose:
            self._load(semantic_model_filename=semantic_model_filename, 
                                 input_config=input_config, 
                                 fcn=fcn, create_object_graph=create_object_graph, 
                                 create_signature_graphs=create_signature_graphs, 
                                 create_system_graph=create_system_graph, 
                                 validate_model=validate_model, force_config_update=force_config_update)
        else:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                self._load(semantic_model_filename=semantic_model_filename, 
                                     input_config=input_config, 
                                     fcn=fcn,
                                     create_object_graph=create_object_graph,
                                     create_signature_graphs=create_signature_graphs,
                                     create_system_graph=create_system_graph,
                                     validate_model=validate_model, force_config_update=force_config_update)

    def _load(self, semantic_model_filename: Optional[str] = None, input_config: Optional[Dict] = None, 
                    fcn: Optional[Callable] = None, create_object_graph: bool = True, 
                    create_signature_graphs: bool = False, create_system_graph: bool = True, 
                    validate_model: bool = True, force_config_update: bool = False) -> None:
        """
        Internal method to load and set up the model for simulation.

        This method is called by load and performs the actual loading process.

        Args:
            semantic_model_filename (Optional[str]): Path to the semantic model configuration file.
            input_config (Optional[Dict]): Input configuration dictionary.
            fcn (Optional[Callable]): Custom function to be applied during model loading.
            create_object_graph (bool): Whether to create and save the object graph.
            create_signature_graphs (bool): Whether to create and save signature graphs.
            create_system_graph (bool): Whether to create and save the system graph.
            validate_model (bool): Whether to perform model validation.
        """

        if self.is_loaded:
            warnings.warn("The model is already loaded. Resetting model.")
            self.reset()

        self.is_loaded = True

        self.p = PrintProgress()
        self.p("Loading model")
        self.p.add_level()
        self.add_outdoor_environment()
        if semantic_model_filename is not None:
            infer_connections = True
            self.p(f"Reading semantic model", status="")
            self._read_datamodel_config(semantic_model_filename)
            
            self._create_object_graph(self.component_base_dict)
            if create_object_graph:
                self.p(f"Drawing input object graph")
                self.draw_object_graph(filename="object_graph_input")

            self.p(f"Parsing semantic model")
            self._parse_semantic_model()
        else:
            infer_connections = False


        if input_config is not None:
            self.p(f"Reading input config")
            self._read_input_config(input_config)

        
        self._create_object_graph(self.component_base_dict)
        if create_object_graph:
            self.p(f"Drawing parsed object graph")
            self.draw_object_graph(filename="object_graph_parsed")

        if create_signature_graphs:
            self.p(f"Drawing signature graphs")
            self._create_signature_graphs()
        
        if infer_connections:
            self.p(f"Connecting components")
            self._connect()

            
        
        if fcn is not None:
            assert callable(fcn), "The function to be applied during model loading is not callable."
            self.p(f"Applying user defined function")
            # self.fcn = fcn.__get__(self, Model) # This is done to avoid the fcn to be shared between instances (https://stackoverflow.com/questions/28127874/monkey-patching-python-an-instance-method)
            fcn(self)
        # self.fcn()

        
        self._create_system_graph()
        if create_system_graph:
            self.p(f"Drawing system graph")
            self.draw_system_graph()



        self.p("Removing cycles")
        self._get_components_no_cycles()
        if create_system_graph:
            self.p("Drawing system graph without cycles")
            self.draw_system_graph_no_cycles()

        self.p("Determining execution order")
        self._get_execution_order()

        self.p("Creating execution graph")
        self._create_flat_execution_graph()

        if create_system_graph:
            self.p("Drawing execution graph")
            self.draw_execution_graph()

        
        self.p("Loading parameters")
        self._load_parameters(force_config_update=force_config_update)

        if validate_model:
            self.p("Validating model")
            self.validate()
        self.p()
        print(self)

    def fcn(self) -> None:
        """
        Placeholder for a custom function to be applied during model loading.
        """

    def save_simulation_result(self, flag: bool=True, c: list=None):
        assert isinstance(flag, bool), "The flag must be a boolean."
        if c is not None:
            assert isinstance(c, list), "The c must be a list."
            for component in c:
                component.saveSimulationResult = flag
        else:
            for component in self.components.values():
                component.saveSimulationResult = flag


    def reset(self) -> None:
        """
        Reset the model to its initial state.
        """
        self.id = self.id  # Keep the original id
        self.saveSimulationResult = self.saveSimulationResult  # Keep the original saveSimulationResult setting

        # Reset all the dictionaries and lists
        self.components = {} ###
        self.component_base_dict = {} ###
        self.system_dict = {"ventilation": {},
                            "heating": {},
                            "cooling": {},
                            } ###
        self.object_dict = {} ###
        self.object_dict_reversed = {} ###
        self.object_counter_dict = {} ###
        self.property_dict = {} ###
        self.instance_map = {} ###
        self.instance_map_reversed = {} ###
        self.instance_to_group_map = {} ###
        self.custom_initial_dict = None ###
        self.execution_order = [] ###
        self.flat_execution_order = [] ###
        self.required_initialization_connections = [] ###
        self._components_no_cycles = {} ###
        self.activeComponents = [] ###
        self.heatexchanger_types = (base.AirToAirHeatRecovery, base.Coil)

        # Reset graphs
        self._initialize_graph("system")
        self._initialize_graph("object")

        # Reset the loaded state
        self.is_loaded = False ###
        self.validated = False ###

        # Reset any estimation results
        self.result = None ###

    def _split_name(self, name: str, linesep: str = "\n") -> str:
        """
        Split a long name into multiple lines for better display.

        Args:
            name (str): The name to split.
            linesep (str): The line separator to use.

        Returns:
            str: The split name.
        """
        split_delimiters = [" ", ")(", "_", "]", "|"]
        new_name = name
        char_len = len(name)
        char_limit = 20
        if any([s in name for s in split_delimiters]):
            if char_len>char_limit:
                name_splits = [name]
                for split_delimiter_ in split_delimiters:
                    new_name_splits = []
                    for name_split in name_splits:
                        splitted = name_split.split(split_delimiter_)
                        n = [e+split_delimiter_ if e and i<len(splitted)-1 else e for i,e in enumerate(splitted)]
                        new_name_splits.extend(n)                    
                    name_splits = new_name_splits

                char_cumsum = np.cumsum(np.array([len(s) for s in name_splits]))
                add_space_char = np.arange(char_cumsum.shape[0])
                char_cumsum = char_cumsum + add_space_char
                idx_arr = np.where(char_cumsum>char_limit)[0]
                if idx_arr.size!=0 and (idx_arr[0]==0 and idx_arr.size==1)==False:
                    if idx_arr[0]==0:
                        idx = idx_arr[1]
                    else:
                        idx = idx_arr[0]
                    name_before_line_break = "".join(name_splits[0:idx])
                    name_after_line_break = "".join(name_splits[idx:])
                    if len(name_after_line_break)>char_limit:
                        name_after_line_break = self._split_name(name_after_line_break, linesep=linesep)
                    
                    if name_before_line_break!="" and name_after_line_break!="":
                        new_name = name_before_line_break + linesep + name_after_line_break
                    else:
                        new_name = name
        return new_name
    
    def _initialize_graph(self, graph_type: str) -> None:
        """
        Initialize a graph of the specified type.

        Args:
            graph_type (str): The type of graph to initialize ('system' or 'object').

        Raises:
            ValueError: If an unknown graph type is specified.
        """
        if graph_type=="system":
            self.system_graph_edge_counter = 0
            self.system_graph = pydot.Dot()
            self.system_subgraph_dict = {}
            self.system_graph_node_attribute_dict = {}
            self.system_graph_edge_label_dict = {}
            self.system_graph_rank=None
        elif graph_type=="object":
            self.object_graph_edge_counter = 0
            self.object_graph = pydot.Dot()
            self.object_subgraph_dict = {}
            self.object_graph_node_attribute_dict = {}
            self.object_graph_edge_label_dict = {}
            self.object_graph_rank=None
        else:
            raise(ValueError(f"Unknown graph type: \"{graph_type}\""))

    def _create_signature_graphs(self) -> None:
        """
        Create signature graphs for all component classes.
        """
        classes = [cls[1] for cls in inspect.getmembers(systems, inspect.isclass) if (issubclass(cls[1], (System, )) and hasattr(cls[1], "sp"))]
        for component_cls in classes:
            sps = component_cls.sp
            for sp in sps:
                d = {s.id: s for s in sp.nodes}
                filename = self.get_dir(folder_list=["graphs", "signatures"], filename=f"signature_{component_cls.__name__}_{sp.id}")[0]
                self._create_object_graph(d, sp.ruleset, add_brackets=False)
                self.draw_graph(filename, self.object_graph)

    def _create_object_graph(self, object_dict: Dict, ruleset: Optional[Dict] = None, add_brackets: bool = True) -> None:
        """
        Create an object graph from the given object dictionary.

        Args:
            object_dict (Dict): Dictionary of objects to include in the graph.
            ruleset (Optional[Dict]): Ruleset for graph creation.
            add_brackets (bool): Whether to add brackets to node labels.
        """
        self._initialize_graph("object")
        exceptions = []
        builtin_types = [getattr(builtins, d) for d in dir(builtins) if isinstance(getattr(builtins, d), type)]
        for exception in exceptions: builtin_types.remove(exception)
        exception_classes = (Connection, ConnectionPoint, np.ndarray, torch.device, pd.DataFrame) # These classes are excluded from the graph 
        exception_classes_exact = (base.DistributionDevice, *builtin_types, count)
        visited = []


        ruleset_applies = True if ruleset is not None else False
        shape_dict = {signature_pattern.IgnoreIntermediateNodes: "circle",
                      signature_pattern.Optional: "diamond",
                      signature_pattern.MultipleMatches: "box",}
        dummy_dim = 0.3
        
        for component in object_dict.values():
            if component not in visited:
                visited = self._depth_first_search_recursive(component, visited, exception_classes, exception_classes_exact)
        
        ignore_nodes = [v.isValueOfProperty for v in visited if isinstance(v, base.PropertyValue) and v.hasValue is None]
        ignore_nodes.extend([v for v in visited if isinstance(v, base.PropertyValue) and v.hasValue is None])
        include_nodes = [v.hasValue for v in visited if isinstance(v, base.PropertyValue) and v.hasValue is not None]
        ignore_edges = ["feedsFluidTo", "hasFluidFedBy"]
        # ignore_edges = []
        end_space = "  "
        for component in visited:
            attributes = get_object_attributes(component)
                
            for attr in attributes:
                if attr not in ignore_edges:
                    edge_label = attr+end_space
                    obj = rgetattr(component, attr)

                    if hasattr(component.__class__, attr) and isinstance(rgetattr(component.__class__, attr), property):
                        class_property = True
                    else:
                        class_property = False

                    if obj is not None and inspect.ismethod(obj)==False and component not in ignore_nodes and class_property==False:
                        if isinstance(obj, list):
                            for receiver_component in obj:
                                cond1 = receiver_component in include_nodes
                                cond2 = isinstance(receiver_component, exception_classes)==False and istype(receiver_component, exception_classes_exact)==False and receiver_component not in ignore_nodes
                                if cond1 or cond2:
                                    if ruleset_applies and attr in component.attributes and receiver_component in component.attributes[attr] and type(ruleset[(component, receiver_component, attr)]) in shape_dict:
                                        dummy = signature_pattern.Node(tuple())
                                        shape = shape_dict[type(ruleset[(component, receiver_component, attr)])]
                                        self._add_graph_relation(self.object_graph, component, dummy, edge_kwargs={"label": edge_label, "arrowhead":"none", "fontname":"CMU Typewriter Text"}, receiver_node_kwargs={"label": "", "shape":shape, "width":dummy_dim, "height":dummy_dim})#, "fontname":"Helvetica", "fontsize":"21", "fontcolor":"black"})
                                        self._add_graph_relation(self.object_graph, dummy, receiver_component, edge_kwargs={"label": ""}, sender_node_kwargs={"label": "", "shape":shape, "width":dummy_dim, "height":dummy_dim})#, "fontname":"Helvetica", "fontsize":"21", "fontcolor":"black"})
                                    else:
                                        self._add_graph_relation(self.object_graph, component, receiver_component, edge_kwargs={"label": edge_label, "fontname":"CMU Typewriter Text"})
                        else:
                            receiver_component = obj
                            cond1 = receiver_component in include_nodes
                            cond2 = isinstance(receiver_component, exception_classes)==False and istype(receiver_component, exception_classes_exact)==False and receiver_component not in ignore_nodes

                            if cond1 or cond2:
                                if ruleset_applies and attr in component.attributes and receiver_component == component.attributes[attr] and type(ruleset[(component, receiver_component, attr)]) in shape_dict:
                                    dummy = signature_pattern.Node(tuple())
                                    shape = shape_dict[type(ruleset[(component, receiver_component, attr)])]
                                    self._add_graph_relation(self.object_graph, component, dummy, edge_kwargs={"label": edge_label, "arrowhead":"none", "fontname":"CMU Typewriter Text"}, receiver_node_kwargs={"label": "", "shape":shape, "width":dummy_dim, "height":dummy_dim})#, "fontname":"Helvetica", "fontsize":"21", "fontcolor":"black"})
                                    self._add_graph_relation(self.object_graph, dummy, receiver_component, edge_kwargs={"label": ""}, sender_node_kwargs={"label": "", "shape":shape, "width":dummy_dim, "height":dummy_dim})#, "fontname":"Helvetica", "fontsize":"21", "fontcolor":"black"})
                                else:
                                    self._add_graph_relation(self.object_graph, component, receiver_component, edge_kwargs={"label": edge_label, "fontname":"CMU Typewriter Text"})
        graph = self.object_graph
        attributes = self.object_graph_node_attribute_dict
        subgraphs = self.object_subgraph_dict
        self._create_graph(graph, attributes, subgraphs, add_brackets=add_brackets)

    def _create_flat_execution_graph(self) -> None:
        """
        Create a flat execution graph based on the execution order.
        """
        self.execution_graph = pydot.Dot()
        prev_node=None
        for i,component_group in enumerate(self.execution_order):
            subgraph = pydot.Subgraph()#graph_name=f"cluster_{i}", style="dotted", penwidth=8)
            for component in component_group:
                node = pydot.Node('"' + component.id + '"')
                if component.id in self.system_graph_node_attribute_dict:
                    node.obj_dict["attributes"].update(self.system_graph_node_attribute_dict[component.id])
                subgraph.add_node(node)
                if prev_node:
                    self._add_edge(self.execution_graph, prev_node.obj_dict["name"], node.obj_dict["name"], edge_kwargs={"label": ""})
                prev_node = node
            self.execution_graph.add_subgraph(subgraph)

    def _create_system_graph(self) -> None:
        """
        Create the system graph.
        """
        graph = self.system_graph
        attributes = self.system_graph_node_attribute_dict
        subgraphs = self.system_subgraph_dict
        self._create_graph(graph, attributes, subgraphs)

    def _is_html(self, name: str) -> bool:
        """
        Check if a name is in HTML format.

        Args:
            name (str): The name to check.

        Returns:
            bool: True if the name is in HTML format, False otherwise.
        """
        if len(name)>=2 and name[0]=="<" and name[-1]==">":
            return True
        else:
            return False

    def _create_graph(self, graph: pydot.Dot, attributes: Dict, subgraphs: Dict, add_brackets: bool = False) -> None:
        """
        Create a graph with the given attributes and subgraphs.

        Args:
            graph (pydot.Dot): The graph to create.
            attributes (Dict): Node attributes.
            subgraphs (Dict): Subgraphs to include.
            add_brackets (bool): Whether to add brackets to node labels.
        """
        light_black = "#3B3838"
        dark_blue = "#44546A"
        orange = "#DC8665"#"#C55A11"
        red = "#873939"
        grey = "#666666"
        light_grey = "#71797E"
        light_blue = "#8497B0"
        green = "#83AF9B"#"#BF9000"
        buttercream = "#B89B72"
        green = "#83AF9B"        

        fill_colors = {base.BuildingSpace: light_black,
                            base.Controller: orange,
                            base.AirToAirHeatRecovery: dark_blue,
                            base.Coil: red,
                            base.Damper: dark_blue,
                            base.Valve: red,
                            base.Fan: dark_blue,
                            base.SpaceHeater: red,
                            base.Sensor: green,
                            base.Meter: green,
                            base.Schedule: grey,
                            base.Pump: red}
        fill_default = light_grey
        border_colors = {base.BuildingSpace: "black",
                            base.Controller: "black",
                            base.AirToAirHeatRecovery: "black",
                            base.Coil: "black",
                            base.Damper: "black",
                            base.Valve: "black",
                            base.Fan: "black",
                            base.SpaceHeater: "black",
                            base.Sensor: "black",
                            base.Meter: "black"}
        border_default = "black"
        K = 1
        min_fontsize = 22*K
        max_fontsize = 28*K

        fontpath, fontname = self._get_font()


        delta_box_width = 0.2
        delta_box_height = 0.5
        width_pad = 2*delta_box_width
        height_pad = 0.1
        nx_graph = nx.drawing.nx_pydot.from_pydot(graph)
        labelwidths = []
        labelheights = []
        for node in nx_graph.nodes():
            name = attributes[node]["label"]
            linesep = "\n"
            _is_html = self._is_html(name)
            name = self._split_name(name, linesep=linesep)
            html_chars = ["<", ">", "SUB," ,"/SUB"]
            no_html_name = name
            for s in html_chars:
                no_html_name = no_html_name.replace(s, "")
            names = no_html_name.split(linesep)

            if _is_html==False: #Convert to html 
                name ="<"+name+">"
            name = name.replace(linesep, "<br />")
            if add_brackets:
                name ="<&#60;" + name[1:]
                name = name[:-1] + "&#62;>"
            char_count_list = [len(s) for s in names if s]
            if len(char_count_list)>0:
                char_count = max(char_count_list)
                linecount = len(names)
            else:
                char_count = 0
                linecount = 0
            attributes[node]["label"] = name
            attributes[node]["labelcharcount"] = char_count
            attributes[node]["labellinecount"] = linecount
            attributes[node]["fontname"] = fontname


        a_fontsize = 0
        b_fontsize = max_fontsize

        a_char_width = delta_box_width
        b_char_width = width_pad

        a_line_height = delta_box_height
        b_line_height = height_pad

        for node in nx_graph.nodes():
            deg = nx_graph.degree(node)
            fontsize = a_fontsize*deg + b_fontsize
            name = attributes[node]["label"]
            labelwidth = attributes[node]["labelcharcount"]
            labelheight = attributes[node]["labellinecount"]
            
            width = a_char_width*labelwidth + b_char_width
            height = a_line_height*labelheight + b_line_height


            if node not in attributes:
                attributes[node] = {}

            attributes[node]["fontsize"] = fontsize

            if "width" not in attributes[node]:
                attributes[node]["width"] = width
            if "height" not in attributes[node]:
                attributes[node]["height"] = height
            cls = self.object_dict[node].__class__
            if cls in fill_colors:
                attributes[node]["fillcolor"] = fill_colors[cls] 
            elif issubclass(cls, tuple(fill_colors.keys())):
                c = [color for c, color in fill_colors.items() if issubclass(cls, c)]
                if len(c)>1:
                    colors = c[0]
                    for color in c[1:]:
                        colors += ":" + color #Currently, the gradient colors are limited to 2
                    c = f"\"{colors}\""
                else:
                    c = c[0]
                attributes[node]["fillcolor"] = c
                
            else:
                attributes[node]["fillcolor"] = fill_default
            
            if cls in border_colors:
                attributes[node]["color"] = border_colors[cls] 
            elif issubclass(cls, tuple(border_colors.keys())):
                c = [c for c in border_colors.keys() if issubclass(cls, c)][0]
                attributes[node]["color"] = border_colors[c]
            else:
                attributes[node]["color"] = border_default

            c = [c for c in subgraphs.keys() if cls is c][0]
            subgraph = subgraphs[c]
            name = node
            if len(subgraph.get_node(name))==1:
                subgraph.get_node(name)[0].obj_dict["attributes"].update(attributes[node])
            elif len(subgraph.get_node(name))==0: #If the name is not present, try with quotes
                 name = "\"" + node + "\""
                 subgraph.get_node(name)[0].obj_dict["attributes"].update(attributes[node])
            else:
                print([el.id for el in self.object_dict.values()])
                raise Exception(f"Multiple identical node names found in subgraph")

    def draw_object_graph(self, filename: str = "object_graph") -> None:
        """
        Draw the object graph and save it to a file.

        Args:
            filename (str): The filename to save the graph to.
        """
        graph = self.object_graph
        self.draw_graph(filename, graph)

    def draw_system_graph_no_cycles(self) -> None:
        """
        Draw the system graph without cycles and save it to a file.
        """
        filename = "system_graph_no_cycles"
        graph = self.system_graph_no_cycles
        self.draw_graph(filename, graph)

    def draw_system_graph(self) -> None:
        """
        Draw the system graph and save it to a file.
        """
        filename = "system_graph"
        graph = self.system_graph
        self.draw_graph(filename, graph)

    def draw_execution_graph(self) -> None:
        """
        Draw the execution graph and save it to a file.
        """
        filename = "execution_graph"
        graph = self.execution_graph
        self.draw_graph(filename, graph)



    def _unflatten(self, filename: str) -> None:
        """
        Unflatten a dot file using the unflatten tool.

        Args:
            filename (str): The filename of the dot file to unflatten.
        """
        app_path = shutil.which("_unflatten")
        args = [app_path, "-f", f"-l 3", f"-o{filename}__unflatten.dot", f"{filename}.dot"]
        subprocess.run(args=args)

    def draw_graph(self, filename: str, graph: pydot.Dot, args: Optional[List[str]] = None) -> None:
        """
        Draw a graph and save it to a file.

        Args:
            filename (str): The filename to save the graph to.
            graph (pydot.Dot): The graph to draw.
            args (Optional[List[str]]): Additional arguments for the graph drawing command.
        """
        fontpath, fontname = self._get_font()
        
        light_grey = "#71797E"
        graph_filename = os.path.join(self.graph_path, f"{filename}.png")
        graph.write(f'{filename}.dot', prog="dot")
        app_path = shutil.which("dot")
        if args is None:
            args = [app_path,
                    "-q",
                    "-Tpng",
                    "-Kdot",
                    f"-Gfontpath={fontpath}",
                    "-Nstyle=filled",
                    "-Nshape=box",
                    "-Nfontcolor=white",
                    f"-Nfontname=Helvetica bold",
                    "-Nfixedsize=true",
                    "-Gnodesep=0.1",
                    "-Efontname=Helvetica",
                    "-Efontsize=21",
                    "-Epenwidth=2",
                    "-Eminlen=1",
                    f"-Ecolor={light_grey}",
                    "-Gcompound=true",
                    "-Grankdir=TB",
                    "-Gsplines=true", #true
                    "-Gmargin=0",
                    "-Gsize=10!",
                    "-Gratio=compress", #0.5 #auto
                    "-Gpack=true",
                    "-Gdpi=5000", #5000 for large graphs
                    "-Grepulsiveforce=0.5",
                    "-Gremincross=true",
                    "-Gstart=1",
                    "-Gbgcolor=transparent",
                    "-q",
                    f"-o{graph_filename}",
                    f"{filename}.dot"] #__unflatten
        else:
            args_ = [app_path]
            args_.extend(args)
            args_.extend([f"-o{graph_filename}", f"{filename}.dot"])
            args = args_
        subprocess.run(args=args)
        os.remove(f"{filename}.dot")

    def _depth_first_search_recursive(self, component: Any, visited: List, exception_classes: Tuple, 
                                      exception_classes_exact: Tuple) -> List:
        """
        Perform a depth-first search on the component graph.

        Args:
            component (Any): The current component being visited.
            visited (List): List of already visited components.
            exception_classes (Tuple): Tuple of classes to be excluded from the search.
            exception_classes_exact (Tuple): Tuple of classes to be excluded from the search (exact match).

        Returns:
            List: Updated list of visited components.
        """
        visited.append(component)
        attributes = dir(component)
        attributes = [attr for attr in attributes if attr[:2]!="__"]#Remove callables
        for attr in attributes:
            obj = rgetattr(component, attr)
            if obj is not None and inspect.ismethod(obj)==False:
                if isinstance(obj, list):
                    for receiver_component in obj:
                        if isinstance(receiver_component, exception_classes)==False and receiver_component not in visited and istype(receiver_component, exception_classes_exact)==False:
                            visited = self._depth_first_search_recursive(receiver_component, visited, exception_classes, exception_classes_exact)
                else:
                    receiver_component = obj
                    if isinstance(receiver_component, exception_classes)==False and receiver_component not in visited and istype(receiver_component, exception_classes_exact)==False:
                        visited = self._depth_first_search_recursive(receiver_component, visited, exception_classes, exception_classes_exact)
        return visited

    def _depth_first_search(self, obj: Any) -> List:
        """
        Perform a depth-first search starting from the given object.

        Args:
            obj (Any): The starting object for the search.

        Returns:
            List: List of visited components.
        """
        visited = []
        visited = self._depth_first_search_recursive(obj, visited)
        return visited

    


    def _shortest_path(self, component: System) -> Dict[System, int]:
        """
        Find the shortest path from the given component to all other components.

        Args:
            component (System): The starting component.

        Returns:
            Dict[System, int]: Dictionary mapping components to their shortest path length.
        """
        def _shortest_path_recursive(shortest_path, exhausted, unvisited):
            while len(unvisited)>0:
                component = unvisited[0]
                current_path_length = shortest_path[component]
                for connection in component.connectedThrough:
                    for connection_point in connection.connectsSystemAt:    
                        receiver_component = connection_point.connectionPointOf

                        if receiver_component not in exhausted:
                            unvisited.append(receiver_component)
                            if receiver_component not in shortest_path: shortest_path[receiver_component] = np.inf
                            if current_path_length+1<shortest_path[receiver_component]:
                                shortest_path[receiver_component] = current_path_length+1
                exhausted.append(component)
                unvisited.remove(component)
            return shortest_path
                
        shortest_path = {}
        shortest_path[component] = 0
        exhausted = []
        unvisited = [component]
        shortest_path = _shortest_path_recursive(shortest_path, exhausted, unvisited)
        return shortest_path


    def get_simple_graph(self, components) -> Dict:
        """
        Get a simple graph representation of the system graph.

        Returns:
            Dict: The simple graph representation.
        """
        simple_graph = {c: [] for c in components.values()}
        for component in components.values():
            for connection in component.connectedThrough:
                for connection_point in connection.connectsSystemAt:
                    receiver_component = connection_point.connectionPointOf
                    simple_graph[component].append(receiver_component)
        return simple_graph

    def get_simple_cycles(self, components: Dict) -> List[List[System]]:
        """
        Get the simple cycles in the system graph.

        Args:
            components (Dict): Dictionary of components.

        Returns:
            List[List[System]]: List of simple cycles.
        """
        G = self.get_simple_graph(components)
        cycles = simple_cycles(G)
        return cycles
 
    def _depth_first_search_system(self, component: System) -> List[System]:
        """
        Perform a depth-first search on the system graph.

        Args:
            component (System): The starting component.

        Returns:
            List[System]: List of visited components.
        """
        def _depth_first_search_recursive_system(component, visited):
            visited.append(component)
            for connection in component.connectedThrough:
                for connection_point in connection.connectsSystemAt:
                    receiver_component = connection_point.connectionPointOf
                    if len(receiver_component.connectedThrough)==0:
                        return visited
                    if receiver_component not in visited:
                        visited = _depth_first_search_recursive_system(receiver_component, visited)
            return visited
        visited = []
        visited = _depth_first_search_recursive_system(component, visited)
        return visited
    
    
 
    def _depth_first_search_cycle_system(self, component: System) -> List[System]:
        """
        Perform a depth-first search on the system graph, stopping at cycles.

        Args:
            component (System): The starting component.

        Returns:
            List[System]: List of visited components.
        """
        def _depth_first_search_recursive_system(component, visited):
            visited.append(component)
            for connection in component.connectedThrough:
                for connection_point in connection.connectsSystemAt:
                    receiver_component = connection_point.connectionPointOf
                    if len(receiver_component.connectedThrough)==0:
                        return visited
                    if receiver_component not in visited:
                        visited = _depth_first_search_recursive_system(receiver_component, visited.copy())
            return visited
        visited = []
        visited = _depth_first_search_recursive_system(component, visited)
        return visited

    def get_subgraph_dict_no_cycles(self) -> None:
        """
        Create a dictionary of subgraphs without cycles.
        """
        self.system_subgraph_dict_no_cycles = copy.deepcopy(self.system_subgraph_dict)
        subgraphs = self.system_graph_no_cycles.get_subgraphs()
        for subgraph in subgraphs:
            subgraph.get_nodes()
            if len(subgraph.get_nodes())>0:
                node = subgraph.get_nodes()[0].obj_dict["name"].replace('"',"")
                self.system_subgraph_dict_no_cycles[self._components_no_cycles[node].__class__] = subgraph

    def get_base_component(self, key: str) -> System:
        """
        Get the base component for a given key.

        Args:
            key (str): The key of the component.

        Returns:
            System: The base component.

        Raises:
            AssertionError: If the mapping is not 1-to-1.
        """
        assert len(self.instance_map[self.components[key]])==1, f"The mapping for component \"{key}\" is not 1-to-1"
        return next(iter(self.instance_map[self.components[key]]))

    def _get_components_no_cycles(self) -> None:
        """
        Create a dictionary of components without cycles.
        """
        self._components_no_cycles = copy.deepcopy(self.components)
        self.system_graph_no_cycles = copy.deepcopy(self.system_graph)
        cycles = self.get_simple_cycles(self._components_no_cycles)
        self.get_subgraph_dict_no_cycles()
        self.required_initialization_connections = []
        for cycle in cycles:
            c_from = [(i, c) for i, c in enumerate(cycle) if isinstance(c, base.Controller)]
            if len(c_from)==1:
                idx = c_from[0][0]
                c_from = c_from[0][1]
            else:
                idx = 0
                c_from = cycle[0]

            if idx==len(cycle)-1:
                c_to = cycle[0]
            else:
                c_to = cycle[idx+1]


            for connection in c_from.connectedThrough.copy():
                for connection_point in connection.connectsSystemAt.copy():
                    if c_to==connection_point.connectionPointOf:
                        connection.connectsSystemAt.remove(connection_point)
                        connection_point.connectsSystemThrough.remove(connection)
                        edge_label = self._get_edge_label(connection.senderPropertyName, connection_point.receiverPropertyName)
                        status = self._del_edge(self.system_graph_no_cycles, c_from.id, c_to.id, label=edge_label)
                        assert status, "del_edge returned False. Check if additional characters should be added to \"disallowed_characters\"."
                        self.required_initialization_connections.append(connection)

                        if len(connection_point.connectsSystemThrough)==0:
                            c_to.connectsAt.remove(connection_point)
                
                if len(connection.connectsSystemAt)==0:
                    c_from.connectedThrough.remove(connection)

    def load_estimation_result(self, filename: Optional[str] = None, result: Optional[Dict] = None) -> None:
        """
        Load a chain log from a file or dictionary.

        Args:
            filename (Optional[str]): The filename to load the chain log from.
            result (Optional[Dict]): The chain log dictionary to load.

        Raises:
            AssertionError: If invalid arguments are provided.
        """
        if result is not None:
            assert isinstance(result, dict), "Argument d must be a dictionary"
            cls_ = result.__class__
            self.result = cls_()
            for key, value in result.items():
                if "chain." not in key:
                    self.result[key] = copy.deepcopy(value)
                else:
                    self.result[key] = value
        else:
            assert isinstance(filename, str), "Argument filename must be a string"
            _, ext = os.path.splitext(filename)
            if ext==".pickle":
                with open(filename, 'rb') as handle:
                    self.result = pickle.load(handle)
                    
            elif ext==".npz":
                if "_ls.npz" in filename:
                    d = dict(np.load(filename, allow_pickle=True))
                    d = {k.replace(".", "_"): v for k,v in d.items()} # For backwards compatibility
                    self.result = estimator.LSEstimationResult(**d)
                elif "_mcmc.npz" in filename:
                    d = dict(np.load(filename, allow_pickle=True))
                    d = {k.replace(".", "_"): v for k,v in d.items()} # For backwards compatibility
                    self.result = estimator.MCMCEstimationResult(**d)
                else:
                    raise Exception(f"The estimation result file is not of a supported type. The file must be a .pickle, .npz file with the name containing \"_ls\" or \"_mcmc\".")
                

                for key, value in self.result.items():
                    self.result[key] = 1/self.result["chain_betas"] if key=="chain_T" else value
                    if self.result[key].size==1 and (len(self.result[key].shape)==0 or len(self.result[key].shape)==1):
                        self.result[key] = value.tolist()

                    elif key=="startTime_train" or key=="endTime_train" or key=="stepSize_train":
                        self.result[key] = value.tolist()
            else:
                raise Exception(f"The estimation result is of type {type(self.result)}. This type is not supported by the model class.")

            

        if isinstance(self.result, estimator.LSEstimationResult):
            theta = self.result["result_x"]
        elif isinstance(self.result, estimator.MCMCEstimationResult):
            parameter_chain = self.result["chain_x"][:,0,:,:]
            parameter_chain = parameter_chain.reshape((parameter_chain.shape[0]*parameter_chain.shape[1], parameter_chain.shape[2]))
            best_index = np.argmax(self.result["chain_logl"], axis=0)[0][0]
            theta = parameter_chain[best_index]
        else:
            raise Exception(f"The estimation result is of type {type(self.result)}. This type is not supported by the model class.")

        flat_component_list = [self.components[com_id] for com_id in self.result["component_id"]]
        flat_attr_list = self.result["component_attr"]
        theta_mask = self.result["theta_mask"]
        theta = theta[theta_mask]
        self.set_parameters_from_array(theta, flat_component_list, flat_attr_list)

    def check_for_for_missing_initial_values(self) -> None:
        """
        Check for missing initial values in components.

        Raises:
            Exception: If any component is missing an initial value.
        """
        for connection in self.required_initialization_connections:
            component = connection.connectsSystem
            if connection.senderPropertyName not in component.output:
                raise Exception(f"The component with id: \"{component.id}\" and class: \"{component.__class__.__name__}\" is missing an initial value for the output: {connection.senderPropertyName}")
            elif component.output[connection.senderPropertyName].get() is None:
                raise Exception(f"The component with id: \"{component.id}\" and class: \"{component.__class__.__name__}\" is missing an initial value for the output: {connection.senderPropertyName}")
                
    def _get_execution_order(self) -> None:
        """
        Determine the execution order of components.

        Raises:
            AssertionError: If cycles are detected in the model.
        """
        def _flatten(_list: List) -> List:
            """
            Flatten a nested list.

            Args:
                _list (List): The nested list to flatten.

            Returns:
                List: The flattened list.
            """
            return [item for sublist in _list for item in sublist]

        def _traverse(self) -> None:
            """
            Traverse the component graph to determine execution order.
            """
            activeComponentsNew = []
            component_group = []
            for component in self.activeComponents:
                component_group.append(component)
                for connection in component.connectedThrough:
                    for connection_point in connection.connectsSystemAt:
                        # connection_point = connection.connectsSystemAt
                        receiver_component = connection_point.connectionPointOf
                        connection_point.connectsSystemThrough.remove(connection)
                        if len(connection_point.connectsSystemThrough)==0:
                            receiver_component.connectsAt.remove(connection_point)

                        if len(receiver_component.connectsAt)==0:
                            activeComponentsNew.append(receiver_component)
            self.activeComponents = activeComponentsNew
            self.execution_order.append(component_group)

        initComponents = [v for v in self._components_no_cycles.values() if len(v.connectsAt)==0]
        self.activeComponents = initComponents
        self.execution_order = []
        while len(self.activeComponents)>0:
            _traverse(self)

        # Map the execution order from the no cycles component dictionary to the full component dictionary.
        self.execution_order = [[self.components[component.id] for component in component_group] for component_group in self.execution_order]

        # Map required initialization connections from the no cycles component dictionary to the full component dictionary.
        self.required_initialization_connections = [connection for no_cycle_connection in self.required_initialization_connections for connection in self.components[no_cycle_connection.connectsSystem.id].connectedThrough if connection.senderPropertyName==no_cycle_connection.senderPropertyName]

        self.flat_execution_order = _flatten(self.execution_order)
        assert len(self.flat_execution_order)==len(self._components_no_cycles), f"Cycles detected in the model. Inspect the generated file \"system_graph.png\" to see where."

    
    


    






    

